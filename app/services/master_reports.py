"""
Сервис генерации персональных отчетов для мастеров
"""

import logging
import os
from datetime import UTC, datetime
from pathlib import Path

import aiofiles  # type: ignore[import-untyped]
from aiogram.types import BufferedInputFile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config import OrderStatus
from app.database import Database
from app.utils import format_datetime


logger = logging.getLogger(__name__)


class MasterReportsService:
    """Сервис для генерации персональных отчетов мастеров"""

    def __init__(self, db: Database):
        """
        Инициализация

        Args:
            db: Экземпляр базы данных
        """
        self.db = db
        self.reports_dir = Path("reports/masters")
        self.reports_dir.mkdir(parents=True, exist_ok=True)

    async def generate_master_report_excel(
        self,
        master_id: int,
        save_to_archive: bool = False,
        period_start: datetime | None = None,
        period_end: datetime | None = None,
    ) -> BufferedInputFile:
        """
        Генерация Excel отчета для мастера с двумя листами

        Файл сохраняется на диск с постоянным именем и обновляется при каждом запросе.

        Args:
            master_id: ID мастера
            save_to_archive: Создать архивную копию с timestamp
            period_start: Начало периода (для архивных отчетов)
            period_end: Конец периода (для архивных отчетов)

        Returns:
            BufferedInputFile: Excel файл
        """
        master = await self.db.get_master_by_id(master_id)
        if not master:
            raise ValueError(f"Мастер с ID {master_id} не найден")

        # Получаем все заявки мастера
        all_orders = await self.db.get_orders_by_master(master_id, exclude_closed=False)

        # Фильтруем по периоду, если указан
        if period_start and period_end:
            all_orders = [o for o in all_orders if period_start <= o.created_at <= period_end]

        # Разделяем на активные и завершенные (включаем REFUSED в завершенные)
        active_orders = [
            o for o in all_orders if o.status not in [OrderStatus.CLOSED, OrderStatus.REFUSED]
        ]
        completed_orders = [
            o for o in all_orders if o.status in [OrderStatus.CLOSED, OrderStatus.REFUSED]
        ]

        # Создаем Excel workbook
        wb = Workbook()

        # Создаем листы
        ws_active = wb.create_sheet("Активные заявки", 0)
        ws_completed = wb.create_sheet("Завершенные заявки", 1)

        # Удаляем стандартный лист (если он есть)
        for sheet_name in wb.sheetnames:
            if sheet_name in ["Sheet", "Sheet1"]:
                del wb[sheet_name]

        # Заполняем листы
        await self._fill_active_orders_sheet(ws_active, active_orders, master)
        await self._fill_completed_orders_sheet(ws_completed, completed_orders, master)

        # Постоянное имя файла (без timestamp) - будет обновляться при каждом запросе
        filename = f"master_{master_id}_report.xlsx"
        file_path = self.reports_dir / filename

        # Удаляем старый файл если существует (для принудительного обновления)
        if file_path.exists():
            file_path.unlink()
            logger.info(f"Старый отчет удален: {filename}")

        # Сохраняем/обновляем файл на диске
        wb.save(file_path)
        logger.info(f"Отчет для мастера {master_id} обновлен: {filename}")

        # Если нужна архивная копия (для ежемесячных отчетов)
        if save_to_archive:
            timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
            archive_filename = f"master_{master_id}_report_{timestamp}.xlsx"
            archive_path = self.reports_dir / archive_filename
            wb.save(archive_path)

            # Подсчитываем статистику
            total_revenue = sum(o.total_amount or 0 for o in completed_orders)

            # Сохраняем запись в БД
            archive_data = {
                "master_id": master_id,
                "period_start": period_start or all_orders[-1].created_at
                if all_orders
                else datetime.now(UTC),
                "period_end": period_end or datetime.now(UTC),
                "file_path": str(archive_path),
                "file_name": archive_filename,
                "file_size": os.path.getsize(archive_path),
                "total_orders": len(all_orders),
                "active_orders": len(active_orders),
                "completed_orders": len(completed_orders),
                "total_revenue": total_revenue,
                "created_at": datetime.now(UTC),
                "notes": "Автоматически созданный архивный отчет за период",
            }

            await self.db.save_master_report_archive(archive_data)

            logger.info(
                f"Архивная копия отчета для мастера {master_id} сохранена: {archive_filename}"
            )

        # Читаем файл с диска и возвращаем
        async with aiofiles.open(file_path, "rb") as f:
            file_data = await f.read()

        return BufferedInputFile(file_data, filename=filename)

    async def _fill_active_orders_sheet(self, ws, orders, master):
        """Заполнение листа активных заявок"""

        # Стили
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Заголовок
        ws.merge_cells("A1:H1")
        ws["A1"] = f"АКТИВНЫЕ ЗАЯВКИ - {master.get_display_name()}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # Дата последнего обновления
        ws.merge_cells("A2:H2")
        ws["A2"] = f"📅 Последнее обновление: {datetime.now(UTC).strftime('%d.%m.%Y %H:%M')} (UTC)"
        ws["A2"].alignment = Alignment(horizontal="center")
        ws["A2"].font = Font(italic=True, color="666666")

        # Заголовки столбцов
        headers = [
            "№ Заявки",
            "Статус",
            "Оборудование",
            "Клиент",
            "Телефон",
            "Адрес",
            "Время прибытия",
            "Дата создания",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Данные
        row_num = 5
        for order in orders:
            ws.cell(row=row_num, column=1, value=order.id).border = border
            ws.cell(
                row=row_num, column=2, value=OrderStatus.get_status_name(order.status)
            ).border = border
            ws.cell(row=row_num, column=3, value=order.equipment_type or "").border = border
            ws.cell(row=row_num, column=4, value=order.client_name or "").border = border
            ws.cell(row=row_num, column=5, value=order.client_phone or "").border = border
            ws.cell(row=row_num, column=6, value=order.client_address or "").border = border
            ws.cell(
                row=row_num, column=7, value=order.scheduled_time if order.scheduled_time else ""
            ).border = border
            ws.cell(
                row=row_num,
                column=8,
                value=format_datetime(order.created_at) if order.created_at else "",
            ).border = border
            row_num += 1

        # Итого
        ws.cell(row=row_num + 1, column=1, value="ИТОГО:").font = Font(bold=True)
        ws.cell(row=row_num + 1, column=2, value=f"{len(orders)} активных заявок").font = Font(
            bold=True
        )

        # Автоширина столбцов
        for col_num in range(1, 9):
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        ws.column_dimensions["A"].width = 20  # № Заявки - делаем шире
        ws.column_dimensions["C"].width = 20  # Оборудование
        ws.column_dimensions["F"].width = 30  # Адрес

    async def _fill_completed_orders_sheet(self, ws, orders, master):
        """Заполнение листа завершенных заявок"""

        # Стили
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Заголовок
        ws.merge_cells("A1:L1")
        ws["A1"] = f"ЗАВЕРШЕННЫЕ ЗАЯВКИ - {master.get_display_name()}"
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill = header_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # Дата последнего обновления
        ws.merge_cells("A2:L2")
        ws["A2"] = f"📅 Последнее обновление: {datetime.now(UTC).strftime('%d.%m.%Y %H:%M')} (UTC)"
        ws["A2"].alignment = Alignment(horizontal="center")
        ws["A2"].font = Font(italic=True, color="666666")

        # Заголовки столбцов
        headers = [
            "№ Заявки",
            "Статус",
            "Оборудование",
            "Клиент",
            "Телефон",
            "Адрес",
            "Общая сумма",
            "Материалы",
            "Прибыль мастера",
            "Прибыль компании",
            "Дата завершения",
            "Причина отказа",
        ]

        table_header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = table_header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Данные
        row_num = 5
        total_amount = 0
        total_materials = 0
        total_master_profit = 0
        total_company_profit = 0

        for order in orders:
            ws.cell(row=row_num, column=1, value=order.id).border = border

            # Статус
            from app.config import OrderStatus

            status_name = OrderStatus.get_status_name(order.status)
            status_cell = ws.cell(row=row_num, column=2, value=status_name)
            status_cell.border = border
            # Подсветка для отказов
            if order.status == OrderStatus.REFUSED:
                status_cell.fill = PatternFill(
                    start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                )

            ws.cell(row=row_num, column=3, value=order.equipment_type or "").border = border
            ws.cell(row=row_num, column=4, value=order.client_name or "").border = border
            ws.cell(row=row_num, column=5, value=order.client_phone or "").border = border
            ws.cell(row=row_num, column=6, value=order.client_address or "").border = border

            amount = order.total_amount or 0
            materials = order.materials_cost or 0
            master_profit = order.master_profit or 0
            company_profit = order.company_profit or 0

            ws.cell(row=row_num, column=7, value=f"{amount:.2f}").border = border
            ws.cell(row=row_num, column=8, value=f"{materials:.2f}").border = border
            ws.cell(row=row_num, column=9, value=f"{master_profit:.2f}").border = border
            ws.cell(row=row_num, column=10, value=f"{company_profit:.2f}").border = border
            ws.cell(
                row=row_num,
                column=11,
                value=format_datetime(order.updated_at) if order.updated_at else "",
            ).border = border

            # Причина отказа
            refuse_reason = order.refuse_reason if hasattr(order, "refuse_reason") else None
            refuse_cell = ws.cell(row=row_num, column=12, value=refuse_reason or "")
            refuse_cell.border = border
            if refuse_reason:
                refuse_cell.alignment = Alignment(wrap_text=True)

            total_amount += amount
            total_materials += materials
            total_master_profit += master_profit
            total_company_profit += company_profit

            row_num += 1

        # Итого
        row_num += 1
        ws.cell(row=row_num, column=1, value="ИТОГО:").font = Font(bold=True)
        ws.cell(row=row_num, column=3, value=f"{len(orders)} завершенных").font = Font(bold=True)
        ws.cell(row=row_num, column=7, value=f"{total_amount:.2f} ₽").font = Font(bold=True)
        ws.cell(row=row_num, column=8, value=f"{total_materials:.2f} ₽").font = Font(bold=True)
        ws.cell(row=row_num, column=9, value=f"{total_master_profit:.2f} ₽").font = Font(
            bold=True, color="4472C4"
        )
        ws.cell(row=row_num, column=10, value=f"{total_company_profit:.2f} ₽").font = Font(
            bold=True
        )

        # Статистика по отказам
        refused_orders_list = [o for o in orders if o.status == OrderStatus.REFUSED]
        refused_with_reason = [
            o for o in refused_orders_list if hasattr(o, "refuse_reason") and o.refuse_reason
        ]

        row_stats = row_num + 2
        ws.cell(row=row_stats, column=1, value="СТАТИСТИКА ПО ОТКАЗАМ:").font = Font(bold=True)
        row_stats += 1

        ws.cell(row=row_stats, column=1, value=f"Всего отказов: {len(refused_orders_list)}")
        row_stats += 1

        ws.cell(
            row=row_stats,
            column=1,
            value=f"С указанием причины: {len(refused_with_reason)}",
        )
        row_stats += 1

        if refused_orders_list:
            percent = (len(refused_with_reason) / len(refused_orders_list)) * 100
            ws.cell(
                row=row_stats,
                column=1,
                value=f"Процент заполнения: {percent:.1f}%",
            )

        # Автоширина столбцов
        for col_num in range(1, 13):
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        ws.column_dimensions["A"].width = 12  # № Заявки
        ws.column_dimensions["B"].width = 12  # Статус
        ws.column_dimensions["C"].width = 18  # Оборудование
        ws.column_dimensions["F"].width = 30  # Адрес
        ws.column_dimensions["L"].width = 35  # Причина отказа - делаем шире

    async def get_master_archived_reports(self, master_id: int, limit: int = 10):
        """
        Получение списка архивных отчетов мастера

        Args:
            master_id: ID мастера
            limit: Лимит записей

        Returns:
            Список архивных отчетов
        """
        return await self.db.get_master_archived_reports(master_id, limit)

    async def get_archived_report_file(self, report_id: int, master_id: int):
        """
        Получение архивного отчета по ID

        Args:
            report_id: ID отчета
            master_id: ID мастера (для проверки прав)

        Returns:
            BufferedInputFile или None
        """
        report = await self.db.get_master_report_archive_by_id(report_id)

        if not report or report.master_id != master_id:
            return None

        file_path = Path(report.file_path)
        if not file_path.exists():
            logger.error(f"Файл отчета не найден: {file_path}")
            return None

        async with aiofiles.open(file_path, "rb") as f:
            file_data = await f.read()

        return BufferedInputFile(file_data, filename=report.file_name)
