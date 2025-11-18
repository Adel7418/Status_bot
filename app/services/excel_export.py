"""
Сервис для экспорта финансовых отчетов в Excel
"""

import logging
from collections.abc import Mapping
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any

import aiosqlite
from openpyxl import Workbook

from app.database import DatabaseType, get_database
from app.repositories.order_repository_extended import OrderRepositoryExtended
from app.services.excel.styles import ExcelStyles
from app.utils.helpers import get_now


if TYPE_CHECKING:
    from app.database.db import Database as LegacyDatabase


logger = logging.getLogger(__name__)


class ExcelExportService:
    """Сервис для экспорта отчетов в Excel"""

    def __init__(self) -> None:
        self.db: DatabaseType = get_database()
        self._order_repo_extended: OrderRepositoryExtended | None = None

    def _get_legacy_db(self) -> "LegacyDatabase":
        """
        Получить legacy-реализацию БД.

        Excel-отчеты используют прямой доступ к SQLite, поэтому
        поддерживается только legacy Database, а не ORMDatabase.
        """
        from app.database.db import Database as LegacyDatabaseRuntime

        if not isinstance(self.db, LegacyDatabaseRuntime):
            raise RuntimeError("ExcelExportService поддерживает только legacy Database (SQLite)")

        return self.db

    def _get_connection(self) -> aiosqlite.Connection:
        """
        Типобезопасно получить соединение с БД.

        Предполагается, что до вызова уже выполнен self.db.connect().
        """
        legacy_db = self._get_legacy_db()
        return legacy_db.get_connection()

    async def _get_extended_repo(self) -> OrderRepositoryExtended:
        """Получить расширенный репозиторий"""
        if self._order_repo_extended is None:
            connection = self._get_connection()
            self._order_repo_extended = OrderRepositoryExtended(connection)
        return self._order_repo_extended

    async def export_report_to_excel(self, report_id: int) -> str | None:
        """
        Экспорт финансового отчета в Excel

        Args:
            report_id: ID отчета

        Returns:
            Путь к созданному файлу или None
        """
        await self.db.connect()

        try:
            # Получаем данные отчета
            report = await self.db.get_financial_report_by_id(report_id)
            if not report:
                logger.error(f"Report {report_id} not found")
                return None

            # Валидация финансовых показателей
            try:
                total_amount = float(report.total_amount or 0)
                materials_cost = float(report.total_materials_cost or 0)
                net_profit_reported = float(report.total_net_profit or 0)
                company_profit = float(report.total_company_profit or 0)
                master_profit = float(report.total_master_profit or 0)

                calculated_net = total_amount - materials_cost
                if abs(calculated_net - net_profit_reported) > 0.01:
                    logger.warning(
                        "Несоответствие в расчете чистой прибыли в отчете %s: "
                        "total_amount - materials_cost = %s, total_net_profit = %s",
                        report_id,
                        calculated_net,
                        report.total_net_profit,
                    )

                calculated_total_profit = company_profit + master_profit
                expected_total = total_amount - materials_cost
                if abs(calculated_total_profit - expected_total) > 0.01:
                    logger.warning(
                        "Несоответствие формулы прибыли в отчете %s: "
                        "company_profit + master_profit != total_amount - materials_cost "
                        "(%s != %s)",
                        report_id,
                        calculated_total_profit,
                        expected_total,
                    )
            except Exception as e:
                logger.error(
                    "Ошибка при валидации финансовых показателей отчета %s: %s", report_id, e
                )

            # Для отчета "ДЕТАЛИЗАЦИЯ ЗАЯВОК ПО МАСТЕРАМ" получаем всех мастеров
            if report.report_type == "masters_detailed":
                all_masters = await self.db.get_all_masters(only_approved=True)
                # Преобразуем в формат master_reports
                master_reports = []
                for master in all_masters:
                    # Создаем объект с нужными атрибутами
                    master_report = type(
                        "MasterReport",
                        (),
                        {
                            "master_id": master.id,
                            "master_name": master.get_display_name(),
                            "total_orders": 0,
                            "total_amount": 0,
                            "total_master_profit": 0,
                        },
                    )()
                    master_reports.append(master_report)
            else:
                master_reports = await self.db.get_master_reports_by_report_id(report_id)

            # Создаем Excel файл
            wb = Workbook()
            ws = wb.active
            ws.title = "Финансовый отчет"

            # Стили из ExcelStyles
            header_font = ExcelStyles.HEADER_FONT
            header_fill = ExcelStyles.HEADER_FILL
            subheader_font = ExcelStyles.SUBHEADER_FONT
            subheader_fill = ExcelStyles.SUBHEADER_FILL

            center_alignment = ExcelStyles.CENTER_ALIGNMENT
            left_alignment = ExcelStyles.LEFT_ALIGNMENT
            right_alignment = ExcelStyles.RIGHT_ALIGNMENT
            thin_border = ExcelStyles.THIN_BORDER

            # Определяем тип отчета
            period_text = ""
            if report.report_type == "DAILY" and report.period_start:
                period_text = f"{report.period_start.strftime('%d.%m.%Y')}"
            elif report.report_type == "WEEKLY" and report.period_start and report.period_end:
                period_text = (
                    f"{report.period_start.strftime('%d.%m')} - "
                    f"{report.period_end.strftime('%d.%m.%Y')}"
                )
            elif report.report_type == "MONTHLY" and report.period_start:
                period_text = f"{report.period_start.strftime('%B %Y')}"

            # Заголовок
            row = 1
            ws.merge_cells(f"A{row}:H{row}")
            cell = ws[f"A{row}"]
            cell.value = f"ФИНАНСОВЫЙ ОТЧЕТ - {report.report_type}"
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            ws.row_dimensions[row].height = ExcelStyles.HEADER_ROW_HEIGHT

            row += 1
            ws.merge_cells(f"A{row}:H{row}")
            cell = ws[f"A{row}"]
            cell.value = f"Период: {period_text}"
            cell.font = subheader_font
            cell.alignment = center_alignment
            ws.row_dimensions[row].height = ExcelStyles.SUBHEADER_ROW_HEIGHT

            row += 2

            # Общие показатели
            ws.merge_cells(f"A{row}:H{row}")
            cell = ws[f"A{row}"]
            cell.value = "ОБЩИЕ ПОКАЗАТЕЛИ"
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.alignment = center_alignment
            cell.border = thin_border

            row += 1
            summary_data: list[list[str | int | float]] = [
                ["Показатель", "Значение"],
                ["Всего заказов", int(report.total_orders or 0)],
                ["Общая сумма", f"{report.total_amount:,.2f} ₽"],
                ["Расходный материал", f"{report.total_materials_cost:,.2f} ₽"],
                ["Чистая прибыль", f"{report.total_net_profit:,.2f} ₽"],
                ["Средний чек", f"{report.average_check:,.2f} ₽"],
                ["", ""],
                ["Прибыль компании", f"{report.total_company_profit:,.2f} ₽"],
                ["Прибыль мастеров", f"{report.total_master_profit:,.2f} ₽"],
            ]

            for row_data in summary_data:
                for col_idx, cell_value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=cell_value)
                    cell.border = thin_border
                    if row_data == summary_data[0]:  # Заголовок
                        cell.font = ExcelStyles.SIMPLE_BOLD_FONT
                        cell.fill = ExcelStyles.TABLE_HEADER_FILL
                    if col_idx == 2:
                        cell.alignment = right_alignment
                    else:
                        cell.alignment = left_alignment
                row += 1

            row += 1

            # Отчеты по мастерам
            if master_reports:
                ws.merge_cells(f"A{row}:H{row}")
                cell = ws[f"A{row}"]
                cell.value = "ОТЧЁТ ПО МАСТЕРАМ"
                cell.font = subheader_font
                cell.fill = subheader_fill
                cell.alignment = center_alignment
                cell.border = thin_border

                row += 1
                headers = [
                    "Мастер",
                    "Заказов",
                    "Сумма",
                    "К сдаче",
                    "Средний чек",
                    "Отзывы",
                    "Выезды",
                    "Прибыль компании",
                ]
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=header)
                    cell.font = ExcelStyles.SIMPLE_BOLD_FONT
                    cell.fill = ExcelStyles.TABLE_HEADER_FILL
                    cell.alignment = center_alignment
                    cell.border = thin_border

                row += 1

                # Данные по мастерам
                for master_report in sorted(
                    master_reports, key=lambda x: x.total_master_profit, reverse=True
                ):
                    data = [
                        master_report.master_name,
                        master_report.orders_count,
                        f"{master_report.total_amount:,.2f} ₽",
                        f"{master_report.total_master_profit:,.2f} ₽",
                        f"{master_report.average_check:,.2f} ₽",
                        master_report.reviews_count,
                        master_report.out_of_city_count,
                        f"{master_report.total_company_profit:,.2f} ₽",
                    ]
                    for col_idx, cell_value in enumerate(data, start=1):
                        cell = ws.cell(row=row, column=col_idx, value=cell_value)
                        cell.border = thin_border
                        if col_idx == 1:
                            cell.alignment = left_alignment
                        else:
                            cell.alignment = (
                                right_alignment
                                if isinstance(cell_value, str) and "₽" in cell_value
                                else center_alignment
                            )
                    row += 1

            # Устанавливаем ширину столбцов
            column_widths: dict[str, int] = {
                "A": 25,  # Мастер/Показатель
                "B": 12,  # Заказов/Значение
                "C": 15,  # Сумма
                "D": 15,  # К сдаче
                "E": 15,  # Средний чек
                "F": 12,  # Отзывы
                "G": 12,  # Выезды
                "H": 18,  # Прибыль компании
            }
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            # ✨ Добавляем лист "Заявки по мастерам"
            if master_reports:
                await self._add_orders_by_master_sheet(
                    wb,
                    master_reports,
                    thin_border,
                    header_font,
                    header_fill,
                    center_alignment,
                    left_alignment,
                    right_alignment,
                )

                # ✨ Добавляем отдельные листы для каждого мастера
                await self._add_individual_master_sheets(
                    wb,
                    master_reports,
                    thin_border,
                    header_font,
                    header_fill,
                    center_alignment,
                    left_alignment,
                    right_alignment,
                )

            # Создаем директорию для отчетов
            reports_dir = Path("reports")
            reports_dir.mkdir(exist_ok=True)

            # Фиксированное имя файла (обновляется каждый раз)
            filename = f"financial_report_{report.report_type.lower()}.xlsx"
            filepath = reports_dir / filename

            # Сохраняем файл (перезаписываем)
            wb.save(filepath)
            logger.info(f"Excel report saved: {filepath}")

            return str(filepath)

        except Exception as e:
            logger.error(f"Error exporting report to Excel: {e}")
            return None

        finally:
            await self.db.disconnect()

    async def _add_orders_by_master_sheet(
        self,
        wb,
        master_reports,
        thin_border,
        header_font,
        header_fill,
        center_alignment,
        left_alignment,
        right_alignment,
    ):
        """Добавляет лист с детализацией заявок по мастерам"""
        ws = wb.create_sheet(title="Заявки по мастерам")

        # Заголовок
        row = 1
        ws.merge_cells(f"A{row}:N{row}")
        cell = ws[f"A{row}"]
        cell.value = "ДЕТАЛИЗАЦИЯ ЗАЯВОК ПО МАСТЕРАМ"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        ws.row_dimensions[row].height = ExcelStyles.HEADER_ROW_HEIGHT

        row += 1

        # Заголовки колонок
        headers = [
            "Мастер",
            "ID",
            "Статус",
            "Тип техники",
            "Клиент",
            "Адрес",
            "Телефон",
            "Создана",
            "Обновлена",
            "Сумма",
            "Материалы",
            "Прибыль мастера",
            "Сдача в кассу",
            "Примечания",
            "Причина отказа",
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = ExcelStyles.TABLE_HEADER_FONT
            cell.fill = ExcelStyles.TABLE_HEADER_FILL
            cell.alignment = center_alignment
            cell.border = thin_border

        row += 1

        # Получаем расширенный репозиторий
        await self._get_extended_repo()

        # Для каждого мастера
        connection = self._get_connection()
        for master_report in master_reports:
            master_id = master_report.master_id
            master_name = master_report.master_name

            if not master_id:
                continue

            # Получаем все заявки мастера
            cursor = await connection.execute(
                """
                SELECT
                    o.id,
                    o.status,
                    o.equipment_type,
                    o.client_name,
                    o.client_address,
                    o.client_phone,
                    o.created_at,
                    o.updated_at,
                    o.total_amount,
                    o.materials_cost,
                    o.master_profit,
                    o.company_profit,
                    o.notes,
                    o.scheduled_time,
                    o.out_of_city,
                    o.has_review,
                    o.refuse_reason
                FROM orders o
                WHERE o.assigned_master_id = ?
                    AND o.status IN ('ASSIGNED', 'ACCEPTED', 'IN_PROGRESS', 'COMPLETED', 'CLOSED', 'REFUSED')
                    AND o.deleted_at IS NULL
                ORDER BY
                    CASE o.status
                        WHEN 'IN_PROGRESS' THEN 1
                        WHEN 'ACCEPTED' THEN 2
                        WHEN 'ASSIGNED' THEN 3
                        WHEN 'COMPLETED' THEN 4
                        WHEN 'CLOSED' THEN 5
                        WHEN 'REFUSED' THEN 6
                        ELSE 7
                    END,
                    o.created_at DESC
                """,
                (master_id,),
            )

            orders = await cursor.fetchall()

            if not orders:
                continue

            # Заголовок мастера
            cell_master = ws[f"A{row}"]
            cell_master.value = f"👨‍🔧 {master_name}"
            cell_master.font = ExcelStyles.MASTER_NAME_FONT
            cell_master.fill = ExcelStyles.MASTER_HEADER_FILL
            cell_master.alignment = left_alignment
            ws.merge_cells(f"A{row}:N{row}")
            ws.row_dimensions[row].height = ExcelStyles.SUBHEADER_ROW_HEIGHT
            row += 1

            # Заявки мастера
            for order in orders:
                status_emoji = {
                    "ASSIGNED": "🆕",
                    "ACCEPTED": "✅",
                    "IN_PROGRESS": "⚙️",
                    "COMPLETED": "✔️",
                    "CLOSED": "🔒",
                    "REFUSED": "❌",
                }.get(order["status"], "❓")

                notes = []
                if order["out_of_city"]:
                    notes.append("Выезд")
                if order["has_review"]:
                    notes.append("Отзыв")
                if order["scheduled_time"] and order["scheduled_time"] != "None":
                    notes.append(f"Время: {order['scheduled_time']}")

                # Безопасное форматирование дат
                created_at = ""
                updated_at = ""
                if order["created_at"]:
                    try:
                        dt = datetime.fromisoformat(order["created_at"])
                        if dt.tzinfo is not None:
                            dt = dt.replace(tzinfo=None)
                        created_at = dt.strftime("%d.%m.%Y %H:%M")
                    except Exception:
                        created_at = str(order["created_at"])[:16]
                if order["updated_at"]:
                    try:
                        dt = datetime.fromisoformat(order["updated_at"])
                        if dt.tzinfo is not None:
                            dt = dt.replace(tzinfo=None)
                        updated_at = dt.strftime("%d.%m.%Y %H:%M")
                    except Exception:
                        updated_at = str(order["updated_at"])[:16]

                data = [
                    "",
                    order["id"],
                    f"{status_emoji} {order['status']}",
                    order["equipment_type"],
                    order["client_name"],
                    order["client_address"][:30] + "..."
                    if len(order["client_address"]) > 30
                    else order["client_address"],
                    order["client_phone"],
                    created_at,
                    updated_at,
                    float(order["total_amount"] or 0),
                    float(order["materials_cost"] or 0),
                    float(order["master_profit"] or 0),
                    float(order["company_profit"] or 0),
                    "; ".join(notes) if notes else "-",
                    order["refuse_reason"] or "",
                ]

                for col_idx, value in enumerate(data, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.border = thin_border

                    if col_idx == 2:  # ID
                        cell.alignment = center_alignment
                        cell.font = ExcelStyles.SIMPLE_BOLD_FONT
                    elif col_idx == 3:  # Статус
                        cell.alignment = center_alignment
                        if order["status"] == "IN_PROGRESS":
                            cell.fill = ExcelStyles.HIGHLIGHT_FILL
                        elif order["status"] == "CLOSED":
                            cell.fill = ExcelStyles.SUCCESS_FILL
                        elif order["status"] == "REFUSED":
                            cell.fill = ExcelStyles.ERROR_FILL
                    elif col_idx in [4, 5, 6, 7, 8, 14]:
                        cell.alignment = left_alignment
                    else:
                        cell.alignment = right_alignment
                        if col_idx >= 10 and col_idx <= 13:
                            cell.number_format = "#,##0.00 ₽"

                row += 1

            # Итоги по мастеру
            cursor = await connection.execute(
                """
                SELECT
                    SUM(CASE WHEN status = 'CLOSED' THEN total_amount ELSE 0 END) as sum_total,
                    SUM(CASE WHEN status = 'CLOSED' THEN materials_cost ELSE 0 END) as sum_materials,
                    SUM(CASE WHEN status = 'CLOSED' THEN master_profit ELSE 0 END) as sum_master,
                    SUM(CASE WHEN status = 'CLOSED' THEN company_profit ELSE 0 END) as sum_company
                FROM orders WHERE assigned_master_id = ? AND deleted_at IS NULL
                """,
                (master_id,),
            )
            totals_row = await cursor.fetchone()
            totals: Mapping[str, Any] = (
                dict(totals_row)
                if totals_row is not None
                else {
                    "sum_total": 0,
                    "sum_materials": 0,
                    "sum_master": 0,
                    "sum_company": 0,
                }
            )

            cell_total = ws[f"A{row}"]
            cell_total.value = f"Итого по {master_name}:"
            cell_total.font = ExcelStyles.BOLD_ITALIC_FONT
            ws.merge_cells(f"A{row}:I{row}")

            for col, val in [
                (f"J{row}", totals["sum_total"]),
                (f"K{row}", totals["sum_materials"]),
                (f"L{row}", totals["sum_master"]),
                (f"M{row}", totals["sum_company"]),
            ]:
                cell = ws[col]
                cell.value = float(val or 0)
                cell.font = ExcelStyles.SIMPLE_BOLD_FONT
                cell.number_format = "#,##0.00 ₽"

            row += 2

        # Ширина столбцов листа "Заявки по мастерам"
        widths = {
            "A": 25,
            "B": 6,
            "C": 15,
            "D": 20,
            "E": 20,
            "F": 30,
            "G": 15,
            "H": 18,
            "I": 18,
            "J": 15,
            "K": 15,
            "L": 18,
            "M": 18,
            "N": 35,
            "O": 40,
        }
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

    async def _add_individual_master_sheets(
        self,
        wb,
        master_reports,
        thin_border,
        header_font,
        header_fill,
        center_alignment,
        left_alignment,
        right_alignment,
    ):
        """Добавляет отдельные листы для каждого мастера"""
        connection = self._get_connection()
        for master_report in master_reports:
            master_id = master_report.master_id
            master_name = master_report.master_name

            if not master_id:
                continue

            # Создаем безопасное имя листа (максимум 31 символ)
            safe_sheet_name = (
                master_name[:31]
                .replace("/", "_")
                .replace("\\", "_")
                .replace("*", "_")
                .replace("?", "_")
                .replace("[", "_")
                .replace("]", "_")
                .replace(":", "_")
            )
            ws = wb.create_sheet(title=safe_sheet_name)

            # Заголовок
            row = 1
            # A1: "ЗАКАЗЫ МАСТЕРА:"
            cell_a1 = ws.cell(row=row, column=1)
            cell_a1.value = "ЗАКАЗЫ МАСТЕРА:"
            cell_a1.font = header_font
            cell_a1.fill = header_fill
            cell_a1.alignment = center_alignment

            # B1: имя мастера
            cell_b1 = ws.cell(row=row, column=2)
            cell_b1.value = master_name
            cell_b1.font = header_font
            cell_b1.fill = header_fill
            cell_b1.alignment = center_alignment

            # Растягиваем заголовок на остальные столбцы
            for col in range(3, 14):  # C1:M1
                ws.cell(row=row, column=col).fill = header_fill

            ws.row_dimensions[row].height = ExcelStyles.HEADER_ROW_HEIGHT

            row += 1

            # Заголовки колонок
            headers = [
                "ID",
                "Статус",
                "Тип техники",
                "Клиент",
                "Адрес",
                "Телефон",
                "Создана",
                "Обновлена",
                "Сумма",
                "Материалы",
                "Прибыль мастера",
                "Сдача в кассу",
                "Примечания",
            ]

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = ExcelStyles.TABLE_HEADER_FONT
                cell.fill = ExcelStyles.TABLE_HEADER_FILL
                cell.alignment = center_alignment
                cell.border = thin_border

            row += 1

            # Получаем все заявки мастера
            cursor = await connection.execute(
                """
                SELECT
                    o.id, o.status, o.equipment_type, o.client_name,
                    o.client_address, o.client_phone, o.created_at, o.updated_at,
                    o.total_amount, o.materials_cost, o.master_profit, o.company_profit,
                    o.notes, o.scheduled_time, o.out_of_city, o.has_review
                FROM orders o
                WHERE o.assigned_master_id = ?
                    AND o.status IN ('ASSIGNED', 'ACCEPTED', 'IN_PROGRESS', 'COMPLETED', 'CLOSED', 'REFUSED')
                    AND o.deleted_at IS NULL
                ORDER BY
                    CASE o.status
                        WHEN 'IN_PROGRESS' THEN 1
                        WHEN 'ACCEPTED' THEN 2
                        WHEN 'ASSIGNED' THEN 3
                        WHEN 'COMPLETED' THEN 4
                        WHEN 'CLOSED' THEN 5
                        WHEN 'REFUSED' THEN 6
                        ELSE 7
                    END,
                    o.created_at DESC
                """,
                (master_id,),
            )

            orders = await cursor.fetchall()

            if not orders:
                # Если заказов нет, добавляем сообщение
                cell = ws[f"A{row}"]
                cell.value = "Заказов не найдено"
                cell.font = ExcelStyles.SIMPLE_ITALIC_FONT
                cell.alignment = center_alignment
                ws.merge_cells(f"A{row}:M{row}")
                row += 1
            else:
                # Заявки мастера
                for order in orders:
                    status_emoji = {
                        "ASSIGNED": "🆕",
                        "ACCEPTED": "✅",
                        "IN_PROGRESS": "⚙️",
                        "COMPLETED": "✔️",
                        "CLOSED": "🔒",
                        "REFUSED": "❌",
                    }.get(order["status"], "❓")

                    notes = []
                    if order["out_of_city"]:
                        notes.append("Выезд")
                    if order["has_review"]:
                        notes.append("Отзыв")
                    if order["scheduled_time"] and order["scheduled_time"] != "None":
                        notes.append(f"Время: {order['scheduled_time']}")
                    if order["notes"]:
                        notes.append(order["notes"][:50])  # Ограничиваем длину

                    # Безопасное форматирование дат
                    created_at = ""
                    updated_at = ""
                    if order["created_at"]:
                        try:
                            dt = datetime.fromisoformat(order["created_at"])
                            if dt.tzinfo is not None:
                                dt = dt.replace(tzinfo=None)
                            created_at = dt.strftime("%d.%m.%Y %H:%M")
                        except Exception:
                            created_at = str(order["created_at"])[:16]
                    if order["updated_at"]:
                        try:
                            dt = datetime.fromisoformat(order["updated_at"])
                            if dt.tzinfo is not None:
                                dt = dt.replace(tzinfo=None)
                            updated_at = dt.strftime("%d.%m.%Y %H:%M")
                        except Exception:
                            updated_at = str(order["updated_at"])[:16]

                    data = [
                        order["id"],
                        f"{status_emoji} {order['status']}",
                        order["equipment_type"],
                        order["client_name"],
                        order["client_address"][:30] + "..."
                        if len(order["client_address"]) > 30
                        else order["client_address"],
                        order["client_phone"],
                        created_at,
                        updated_at,
                        float(order["total_amount"] or 0),
                        float(order["materials_cost"] or 0),
                        float(order["master_profit"] or 0),
                        float(order["company_profit"] or 0),
                        "; ".join(notes) if notes else "-",
                    ]

                    for col_idx, value in enumerate(data, start=1):
                        cell = ws.cell(row=row, column=col_idx, value=value)
                        cell.border = thin_border

                        if col_idx == 1:  # ID
                            cell.alignment = center_alignment
                            cell.font = ExcelStyles.SIMPLE_BOLD_FONT
                        elif col_idx == 2:  # Статус
                            cell.alignment = center_alignment
                            if order["status"] == "IN_PROGRESS":
                                cell.fill = ExcelStyles.HIGHLIGHT_FILL
                            elif order["status"] == "CLOSED":
                                cell.fill = ExcelStyles.SUCCESS_FILL
                            elif order["status"] == "REFUSED":
                                cell.fill = ExcelStyles.ERROR_FILL
                        elif col_idx in [3, 4, 5, 6, 7, 13]:  # Текстовые поля
                            cell.alignment = left_alignment
                        else:  # Числовые поля
                            cell.alignment = right_alignment
                            if col_idx >= 9 and col_idx <= 12:
                                cell.number_format = "#,##0.00 ₽"

                    row += 1

                # Итоги по мастеру
                totals_cursor = await connection.execute(
                    """
                    SELECT
                        COUNT(*) as total_orders,
                        SUM(CASE WHEN status = 'CLOSED' THEN 1 ELSE 0 END) as closed_orders,
                        SUM(CASE WHEN status = 'CLOSED' THEN total_amount ELSE 0 END) as sum_total,
                        SUM(CASE WHEN status = 'CLOSED' THEN materials_cost ELSE 0 END) as sum_materials,
                        SUM(CASE WHEN status = 'CLOSED' THEN master_profit ELSE 0 END) as sum_master,
                        SUM(CASE WHEN status = 'CLOSED' THEN company_profit ELSE 0 END) as sum_company
                    FROM orders WHERE assigned_master_id = ? AND deleted_at IS NULL
                    """,
                    (master_id,),
                )
                totals_row = await totals_cursor.fetchone()
                totals: Mapping[str, Any] = (
                    dict(totals_row)
                    if totals_row is not None
                    else {
                        "total_orders": 0,
                        "closed_orders": 0,
                        "sum_total": 0,
                        "sum_materials": 0,
                        "sum_master": 0,
                        "sum_company": 0,
                    }
                )

                # Добавляем пустую строку
                row += 1

                # Итоги
                summary_data = [
                    "ИТОГО:",
                    f"Всего: {totals['total_orders']}",
                    f"Закрыто: {totals['closed_orders']}",
                    "",
                    "",
                    "",
                    "",
                    "",
                    float(totals["sum_total"] or 0),
                    float(totals["sum_materials"] or 0),
                    float(totals["sum_master"] or 0),
                    float(totals["sum_company"] or 0),
                    "",
                ]

                for col_idx, value in enumerate(summary_data, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.font = ExcelStyles.SIMPLE_BOLD_FONT

                    if col_idx == 1:  # "ИТОГО:"
                        cell.alignment = left_alignment
                        cell.fill = ExcelStyles.TABLE_HEADER_FILL
                    elif col_idx in (2, 3):  # "Всего:" or "Закрыто:"
                        cell.alignment = center_alignment
                        cell.fill = ExcelStyles.TABLE_HEADER_FILL
                    elif col_idx >= 9 and col_idx <= 12:  # Числовые поля
                        cell.alignment = right_alignment
                        cell.number_format = "#,##0.00 ₽"
                        cell.fill = ExcelStyles.TABLE_HEADER_FILL

            # Ширина столбцов для листа мастера
            widths: dict[str, int] = {
                "A": 8,  # ID
                "B": 15,  # Статус
                "C": 20,  # Тип техники
                "D": 20,  # Клиент
                "E": 30,  # Адрес
                "F": 15,  # Телефон
                "G": 16,  # Создана
                "H": 16,  # Обновлена
                "I": 15,  # Сумма
                "J": 15,  # Материалы
                "K": 18,  # Прибыль мастера
                "L": 18,  # Сдача в кассу
                "M": 35,  # Примечания
            }
            for col_letter, width in widths.items():
                ws.column_dimensions[col_letter].width = width

    async def export_closed_orders_to_excel(self, period_days: int = 30) -> str | None:
        """
        Экспорт закрытых заказов в Excel (обновляет существующий файл)

        Args:
            period_days: За сколько дней показывать заказы

        Returns:
            Путь к файлу или None
        """
        await self.db.connect()

        try:
            # Валидация периода
            if period_days <= 0:
                logger.error(f"Неверный период для закрытых заказов: {period_days} дней")
                return None
            if period_days > 365:
                logger.warning(f"Слишком большой период для закрытых заказов: {period_days} дней")

            # Имя файла
            connection = self._get_connection()
            reports_dir = Path("reports")
            reports_dir.mkdir(exist_ok=True)
            filepath = reports_dir / "closed_orders.xlsx"

            # Создаем новый workbook (перезаписываем файл)
            wb = Workbook()
            ws = wb.active
            ws.title = "Закрытые заказы"

            # Стили из ExcelStyles
            header_font = ExcelStyles.HEADER_FONT
            header_fill = ExcelStyles.HEADER_FILL
            center_alignment = ExcelStyles.CENTER_ALIGNMENT
            left_alignment = ExcelStyles.LEFT_ALIGNMENT
            right_alignment = ExcelStyles.RIGHT_ALIGNMENT
            thin_border = ExcelStyles.THIN_BORDER

            # Заголовок
            row = 1
            ws.merge_cells(f"A{row}:K{row}")
            cell = ws[f"A{row}"]
            cell.value = f"ЗАКРЫТЫЕ ЗАКАЗЫ (за {period_days} дней)"
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            ws.row_dimensions[row].height = ExcelStyles.HEADER_ROW_HEIGHT

            row += 1
            ws.merge_cells(f"A{row}:K{row}")
            cell = ws[f"A{row}"]
            cell.value = f"Обновлено: {get_now().strftime('%d.%m.%Y %H:%M')}"
            cell.font = ExcelStyles.BOLD_FONT
            cell.alignment = center_alignment

            row += 2

            # Заголовки колонок
            headers = [
                "ID",
                "Техника",
                "Клиент",
                "Мастер",
                "Создано",
                "Закрыто",
                "Сумма",
                "Материалы",
                "Прибыль мастера",
                "Прибыль компании",
                "Доп. инфо",
            ]

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = ExcelStyles.TABLE_HEADER_FONT
                cell.fill = ExcelStyles.TABLE_HEADER_FILL
                cell.alignment = center_alignment
                cell.border = thin_border

            row += 1

            # Получаем закрытые заказы
            from datetime import timedelta

            start_date = get_now() - timedelta(days=period_days)

            cursor = await connection.execute(
                """
                SELECT
                    o.id, o.equipment_type, o.client_name, o.created_at, o.updated_at,
                    o.total_amount, o.materials_cost, o.master_profit, o.company_profit,
                    o.out_of_city, o.has_review,
                    u.first_name || ' ' || COALESCE(u.last_name, '') as master_name
                FROM orders o
                LEFT JOIN masters m ON o.assigned_master_id = m.id
                LEFT JOIN users u ON m.telegram_id = u.telegram_id
                WHERE o.status = 'CLOSED'
                    AND o.updated_at >= ?
                    AND o.deleted_at IS NULL
                ORDER BY o.updated_at DESC
                """,
                (start_date.isoformat(),),
            )

            orders = await cursor.fetchall()

            if not orders:
                ws[f"A{row}"] = "Нет закрытых заказов за этот период"
                ws[f"A{row}"].font = ExcelStyles.SIMPLE_ITALIC_FONT
                ws.merge_cells(f"A{row}:K{row}")
            else:
                # Выводим заказы
                for order in orders:
                    additional_info = []
                    if order["out_of_city"]:
                        additional_info.append("Выезд")
                    if order["has_review"]:
                        additional_info.append("Отзыв")

                    # Безопасное форматирование дат
                    created_at = ""
                    updated_at = ""
                    if order["created_at"]:
                        try:
                            dt = datetime.fromisoformat(order["created_at"])
                            if dt.tzinfo is not None:
                                dt = dt.replace(tzinfo=None)
                            created_at = dt.strftime("%d.%m.%Y %H:%M")
                        except Exception:
                            created_at = str(order["created_at"])[:16]
                    if order["updated_at"]:
                        try:
                            dt = datetime.fromisoformat(order["updated_at"])
                            if dt.tzinfo is not None:
                                dt = dt.replace(tzinfo=None)
                            updated_at = dt.strftime("%d.%m.%Y %H:%M")
                        except Exception:
                            updated_at = str(order["updated_at"])[:16]

                    data = [
                        order["id"],
                        order["equipment_type"],
                        order["client_name"],
                        order["master_name"] or "Не назначен",
                        created_at,
                        updated_at,
                        float(order["total_amount"] or 0),
                        float(order["materials_cost"] or 0),
                        float(order["master_profit"] or 0),
                        float(order["company_profit"] or 0),
                        ", ".join(additional_info) if additional_info else "-",
                    ]

                    for col_idx, value in enumerate(data, start=1):
                        cell = ws.cell(row=row, column=col_idx, value=value)
                        cell.border = thin_border

                        if col_idx == 1:
                            cell.alignment = center_alignment
                            cell.font = ExcelStyles.SIMPLE_BOLD_FONT
                        elif col_idx in [2, 3, 4, 5, 6, 11]:
                            cell.alignment = left_alignment
                        else:
                            cell.alignment = right_alignment
                            if col_idx >= 7 and col_idx <= 10:
                                cell.number_format = "#,##0.00 ₽"

                    row += 1

                # Итоги
                row += 1
                ws[f"A{row}"] = "ИТОГО:"
                ws[f"A{row}"].font = ExcelStyles.BOLD_FONT

                total_sum = sum(float(o["total_amount"] or 0) for o in orders)
                total_materials = sum(float(o["materials_cost"] or 0) for o in orders)
                total_master_profit = sum(float(o["master_profit"] or 0) for o in orders)
                total_company_profit = sum(float(o["company_profit"] or 0) for o in orders)

                for col, val in [
                    (f"G{row}", total_sum),
                    (f"H{row}", total_materials),
                    (f"I{row}", total_master_profit),
                    (f"J{row}", total_company_profit),
                ]:
                    cell = ws[col]
                    cell.value = val
                    cell.font = ExcelStyles.BOLD_FONT
                    cell.number_format = "#,##0.00 ₽"
                    cell.fill = ExcelStyles.HIGHLIGHT_FILL

            # Ширина столбцов
            widths: dict[str, int] = {
                "A": 12,  # ID - соответствует документации (6-12)
                "B": 25,
                "C": 20,
                "D": 20,
                "E": 18,
                "F": 18,
                "G": 15,
                "H": 15,
                "I": 18,
                "J": 18,
                "K": 22,
            }
            for col_letter, width in widths.items():
                ws.column_dimensions[col_letter].width = width

            # Сохраняем файл
            wb.save(filepath)
            logger.info(f"Closed orders Excel saved: {filepath}")

            return str(filepath)

        except Exception as e:
            logger.error(f"Error exporting closed orders: {e}")
            return None

        finally:
            await self.db.disconnect()

    async def export_masters_statistics_to_excel(self) -> str | None:
        """
        Экспорт статистики по мастерам в Excel (обновляет существующий файл)

        Returns:
            Путь к файлу или None
        """
        await self.db.connect()

        try:
            # Имя файла
            reports_dir = Path("reports")
            reports_dir.mkdir(exist_ok=True)
            filepath = reports_dir / "masters_statistics.xlsx"

            # Создаем новый workbook (перезаписываем файл)
            wb = Workbook()
            ws = wb.active
            ws.title = "Статистика мастеров"

            # Стили из ExcelStyles
            header_font = ExcelStyles.HEADER_FONT
            header_fill = ExcelStyles.HEADER_FILL
            center_alignment = ExcelStyles.CENTER_ALIGNMENT
            left_alignment = ExcelStyles.LEFT_ALIGNMENT
            right_alignment = ExcelStyles.RIGHT_ALIGNMENT
            thin_border = ExcelStyles.THIN_BORDER

            # Заголовок
            row = 1
            ws.merge_cells(f"A{row}:N{row}")
            cell = ws[f"A{row}"]
            cell.value = "СТАТИСТИКА ПО МАСТЕРАМ"
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            ws.row_dimensions[row].height = ExcelStyles.HEADER_ROW_HEIGHT

            row += 1
            ws.merge_cells(f"A{row}:N{row}")
            cell = ws[f"A{row}"]
            cell.value = f"Обновлено: {get_now().strftime('%d.%m.%Y %H:%M')}"
            cell.font = ExcelStyles.BOLD_FONT
            cell.alignment = center_alignment

            row += 2

            # Заголовки колонок
            headers = [
                "ID",
                "Мастер",
                "Всего заявок",
                "Завершено",
                "В работе",
                "Отказано",
                "Прочие статусы",
                "Общая сумма",
                "Материалы",
                "Чистая прибыль",
                "Прибыль компании",
                "Сдача в кассу",
                "Средний чек",
                "Выездов",
                "Отзывов",
            ]

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = ExcelStyles.TABLE_HEADER_FONT
                cell.fill = ExcelStyles.TABLE_HEADER_FILL
                cell.alignment = center_alignment
                cell.border = thin_border

            row += 1

            # Получаем всех мастеров
            connection = self._get_connection()
            masters_cursor = await connection.execute(
                """
                SELECT
                    m.id,
                    u.first_name || ' ' || COALESCE(u.last_name, '') as full_name
                FROM masters m
                LEFT JOIN users u ON m.telegram_id = u.telegram_id
                WHERE m.is_approved = 1 AND m.deleted_at IS NULL
                ORDER BY u.first_name
                """
            )
            masters = await masters_cursor.fetchall()

            if not masters:
                ws[f"A{row}"] = "Нет утвержденных мастеров"
                ws[f"A{row}"].font = ExcelStyles.SIMPLE_ITALIC_FONT
                ws.merge_cells(f"A{row}:N{row}")
            else:
                # Данные по каждому мастеру
                for master in masters:
                    master_id = master["id"]
                    master_name = master["full_name"]

                    # Получаем статистику
                    cursor = await connection.execute(
                        """
                        SELECT
                            COUNT(*) as total_orders,
                            SUM(CASE WHEN status = 'CLOSED' THEN 1 ELSE 0 END) as closed,
                            SUM(CASE WHEN status IN ('ASSIGNED', 'IN_PROGRESS', 'ACCEPTED') THEN 1 ELSE 0 END) as in_work,
                            SUM(CASE WHEN status = 'REFUSED' THEN 1 ELSE 0 END) as refused,
                            SUM(CASE WHEN status = 'CLOSED' THEN total_amount ELSE 0 END) as total_sum,
                            SUM(CASE WHEN status = 'CLOSED' THEN materials_cost ELSE 0 END) as materials_sum,
                            SUM(CASE WHEN status = 'CLOSED' THEN master_profit ELSE 0 END) as master_profit_sum,
                            SUM(CASE WHEN status = 'CLOSED' THEN company_profit ELSE 0 END) as company_profit_sum,
                            SUM(CASE WHEN status = 'CLOSED' AND out_of_city = 1 THEN 1 ELSE 0 END) as out_of_city,
                            SUM(CASE WHEN status = 'CLOSED' AND has_review = 1 THEN 1 ELSE 0 END) as reviews,
                            AVG(CASE WHEN status = 'CLOSED' THEN total_amount ELSE NULL END) as avg_check
                        FROM orders
                        WHERE assigned_master_id = ?
                            AND deleted_at IS NULL
                        """,
                        (master_id,),
                    )

                    stats_row = await cursor.fetchone()

                    if not stats_row or not stats_row["total_orders"]:
                        continue

                    # Вычисляем данные
                    total_sum = float(stats_row["total_sum"] or 0)
                    materials = float(stats_row["materials_sum"] or 0)
                    net_profit = total_sum - materials
                    cash_to_company = float(stats_row["company_profit_sum"] or 0)
                    other_statuses = (stats_row["total_orders"] or 0) - (
                        (stats_row["closed"] or 0)
                        + (stats_row["in_work"] or 0)
                        + (stats_row["refused"] or 0)
                    )

                    # Данные по мастеру
                    master_data = [
                        master_id,
                        master_name,
                        stats_row["total_orders"] or 0,
                        stats_row["closed"] or 0,
                        stats_row["in_work"] or 0,
                        stats_row["refused"] or 0,
                        other_statuses,
                        total_sum,
                        materials,
                        net_profit,
                        cash_to_company,
                        cash_to_company,
                        float(stats_row["avg_check"] or 0),
                        stats_row["out_of_city"] or 0,
                        stats_row["reviews"] or 0,
                    ]

                    for col_idx, value in enumerate(master_data, start=1):
                        cell = ws.cell(row=row, column=col_idx, value=value)
                        cell.border = thin_border

                        if col_idx == 1:
                            cell.alignment = center_alignment
                            cell.font = ExcelStyles.SIMPLE_BOLD_FONT
                        elif col_idx == 2:
                            cell.alignment = left_alignment
                            cell.font = ExcelStyles.SIMPLE_BOLD_FONT
                        elif col_idx in [3, 4, 5, 6, 7, 14, 15]:
                            cell.alignment = center_alignment
                        else:
                            cell.alignment = right_alignment
                            if col_idx >= 8 and col_idx <= 13:
                                cell.number_format = "#,##0.00 ₽"

                    row += 1

                # ИТОГО
                row += 1
                ws[f"A{row}"] = "ИТОГО:"
                ws[f"A{row}"].font = ExcelStyles.SUBHEADER_FONT
                ws[f"A{row}"].fill = ExcelStyles.HIGHLIGHT_FILL
                ws.merge_cells(f"A{row}:B{row}")

                # Суммы
                cursor = await connection.execute(
                    """
                    SELECT
                        COUNT(*) as total_orders,
                        SUM(CASE WHEN status = 'CLOSED' THEN 1 ELSE 0 END) as closed,
                        SUM(CASE WHEN status IN ('ASSIGNED', 'IN_PROGRESS', 'ACCEPTED') THEN 1 ELSE 0 END) as in_work,
                        SUM(CASE WHEN status = 'REFUSED' THEN 1 ELSE 0 END) as refused,
                        SUM(CASE WHEN status = 'CLOSED' THEN total_amount ELSE 0 END) as total_sum,
                        SUM(CASE WHEN status = 'CLOSED' THEN materials_cost ELSE 0 END) as materials_sum,
                        SUM(CASE WHEN status = 'CLOSED' THEN company_profit ELSE 0 END) as company_profit_sum,
                        AVG(CASE WHEN status = 'CLOSED' THEN total_amount ELSE NULL END) as avg_check,
                        SUM(CASE WHEN status = 'CLOSED' AND out_of_city = 1 THEN 1 ELSE 0 END) as out_of_city,
                        SUM(CASE WHEN status = 'CLOSED' AND has_review = 1 THEN 1 ELSE 0 END) as reviews
                    FROM orders
                    WHERE assigned_master_id IS NOT NULL
                        AND deleted_at IS NULL
                    """
                )

                totals_row = await cursor.fetchone()
                totals: Mapping[str, Any] = (
                    dict(totals_row)
                    if totals_row is not None
                    else {
                        "total_sum": 0,
                        "materials_sum": 0,
                        "total_orders": 0,
                        "closed": 0,
                        "in_work": 0,
                        "refused": 0,
                        "company_profit_sum": 0,
                        "avg_check": 0,
                        "out_of_city": 0,
                        "reviews": 0,
                    }
                )

                total_sum = float(totals["total_sum"] or 0)
                materials_sum = float(totals["materials_sum"] or 0)
                net_profit_total = total_sum - materials_sum
                others_total = (totals["total_orders"] or 0) - (
                    (totals["closed"] or 0) + (totals["in_work"] or 0) + (totals["refused"] or 0)
                )

                totals_data: list[Any] = [
                    totals["total_orders"],
                    totals["closed"],
                    totals["in_work"],
                    totals["refused"],
                    others_total,
                    total_sum,
                    materials_sum,
                    net_profit_total,
                    float(totals["company_profit_sum"] or 0),
                    float(totals["company_profit_sum"] or 0),
                    float(totals["avg_check"] or 0),
                    totals["out_of_city"],
                    totals["reviews"],
                ]

                for col_idx, value in enumerate(totals_data, start=3):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.font = ExcelStyles.BOLD_FONT
                    cell.fill = ExcelStyles.HIGHLIGHT_FILL
                    cell.border = thin_border

                    if col_idx in [3, 4, 5, 6, 7, 14, 15]:
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = right_alignment
                        if col_idx >= 8 and col_idx <= 13:
                            cell.number_format = "#,##0.00 ₽"

            # Ширина столбцов
            widths: dict[str, int] = {
                "A": 12,  # ID - соответствует документации (6-12)
                "B": 25,
                "C": 12,
                "D": 12,
                "E": 12,
                "F": 12,
                "G": 15,
                "H": 15,
                "I": 15,
                "J": 18,
                "K": 15,
                "L": 15,
                "M": 10,
                "N": 10,
            }
            for col_letter, width in widths.items():
                ws.column_dimensions[col_letter].width = width

            # Сохраняем файл
            wb.save(filepath)
            logger.info(f"Masters statistics Excel saved: {filepath}")

            return str(filepath)

        except Exception as e:
            logger.error(f"Error exporting masters statistics: {e}")
            return None

        finally:
            await self.db.disconnect()

    async def export_master_orders_to_excel(self, master_id: int) -> str | None:
        """
        Экспорт всех заявок одного мастера в Excel

        Args:
            master_id: ID мастера

        Returns:
            Путь к файлу или None
        """
        await self.db.connect()

        try:
            connection = self._get_connection()

            # Получаем информацию о мастере
            cursor = await connection.execute(
                """
                SELECT
                    m.id,
                    u.first_name || ' ' || COALESCE(u.last_name, '') as full_name,
                    m.phone
                FROM masters m
                LEFT JOIN users u ON m.telegram_id = u.telegram_id
                WHERE m.id = ?
                """,
                (master_id,),
            )
            master = await cursor.fetchone()

            if not master:
                logger.error(f"Master {master_id} not found")
                return None

            master_name = master["full_name"]

            # Имя файла
            reports_dir = Path("reports")
            reports_dir.mkdir(exist_ok=True)
            safe_name = "".join(
                c for c in master_name if c.isalnum() or c in (" ", "-", "_")
            ).strip()
            filepath = reports_dir / f"master_{master_id}_{safe_name}.xlsx"

            # Получаем все заявки мастера для разделения
            connection = self._get_connection()
            all_orders_cursor = await connection.execute(
                """
                SELECT
                    id, status, equipment_type, client_name, client_address, client_phone,
                    created_at, updated_at, total_amount, materials_cost,
                    master_profit, company_profit, out_of_city, has_review, refuse_reason
                FROM orders
                WHERE assigned_master_id = ? AND deleted_at IS NULL
                ORDER BY created_at DESC
                """,
                (master_id,),
            )
            all_orders = await all_orders_cursor.fetchall()

            # Разделяем заявки на активные и завершенные
            active_orders = [o for o in all_orders if o["status"] not in ["CLOSED", "REFUSED"]]
            completed_orders = [o for o in all_orders if o["status"] in ["CLOSED", "REFUSED"]]

            # Создаем workbook с двумя листами
            wb = Workbook()
            ws_active = wb.create_sheet("Активные заявки", 0)
            ws_completed = wb.create_sheet("Завершенные заявки", 1)

            # Удаляем стандартный лист
            for sheet_name in wb.sheetnames:
                if sheet_name in ["Sheet", "Sheet1"]:
                    del wb[sheet_name]

            # Сначала заполним лист активных заявок
            ws = ws_active

            # Стили из ExcelStyles
            header_font = ExcelStyles.HEADER_FONT
            header_fill = ExcelStyles.HEADER_FILL
            center_alignment = ExcelStyles.CENTER_ALIGNMENT
            left_alignment = ExcelStyles.LEFT_ALIGNMENT
            right_alignment = ExcelStyles.RIGHT_ALIGNMENT
            thin_border = ExcelStyles.THIN_BORDER

            # Заголовок
            row = 1
            # Заполняем A1:D1 фоном
            for col in range(1, 5):  # A1:D1
                ws.cell(row=row, column=col).fill = header_fill

            # E1: "ОТЧЕТ ПО МАСТЕРУ:"
            cell_e1 = ws.cell(row=row, column=5)
            cell_e1.value = "ОТЧЕТ ПО МАСТЕРУ:"
            cell_e1.font = header_font
            cell_e1.fill = header_fill
            cell_e1.alignment = center_alignment

            # F1: имя мастера
            cell_f1 = ws.cell(row=row, column=6)
            cell_f1.value = master_name
            cell_f1.font = header_font
            cell_f1.fill = header_fill
            cell_f1.alignment = center_alignment

            # Растягиваем заголовок на остальные столбцы
            for col in range(7, 9):  # G1:H1
                ws.cell(row=row, column=col).fill = header_fill

            ws.row_dimensions[row].height = ExcelStyles.HEADER_ROW_HEIGHT

            row += 1
            ws.merge_cells(f"A{row}:H{row}")
            cell = ws[f"A{row}"]
            cell.value = (
                f"Обновлено: {get_now().strftime('%d.%m.%Y %H:%M')} | Телефон: {master['phone']}"
            )
            cell.font = ExcelStyles.BOLD_FONT
            cell.alignment = center_alignment

            row += 2

            # Заголовки колонок таблицы заявок
            headers = [
                "ID",
                "Статус",
                "Тип техники",
                "Клиент",
                "Адрес",
                "Телефон",
                "Создана",
                "Обновлена",
            ]

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = ExcelStyles.TABLE_HEADER_FONT
                cell.fill = ExcelStyles.TABLE_HEADER_FILL
                cell.alignment = center_alignment
                cell.border = thin_border

            row += 1

            # Используем активные заявки для первого листа
            orders = active_orders

            if not orders:
                ws[f"A{row}"] = "У мастера пока нет активных заявок"
                ws[f"A{row}"].font = ExcelStyles.SIMPLE_ITALIC_FONT
                ws.merge_cells(f"A{row}:H{row}")
            else:
                # Выводим заявки
                for order in orders:
                    status_emoji = {
                        "NEW": "🆕",
                        "ASSIGNED": "📋",
                        "ACCEPTED": "✅",
                        "IN_PROGRESS": "⚙️",
                        "COMPLETED": "✔️",
                        "CLOSED": "🔒",
                        "REFUSED": "❌",
                    }.get(order["status"], "❓")

                    # Безопасное форматирование дат
                    created_at = ""
                    updated_at = ""
                    if order["created_at"]:
                        try:
                            dt = datetime.fromisoformat(order["created_at"])
                            if dt.tzinfo is not None:
                                dt = dt.replace(tzinfo=None)
                            created_at = dt.strftime("%d.%m.%Y %H:%M")
                        except Exception:
                            created_at = str(order["created_at"])[:16]
                    if order["updated_at"]:
                        try:
                            dt = datetime.fromisoformat(order["updated_at"])
                            if dt.tzinfo is not None:
                                dt = dt.replace(tzinfo=None)
                            updated_at = dt.strftime("%d.%m.%Y %H:%M")
                        except Exception:
                            updated_at = str(order["updated_at"])[:16]

                    data = [
                        order["id"],
                        f"{status_emoji} {order['status']}",
                        order["equipment_type"],
                        order["client_name"],
                        order["client_address"][:30] + "..."
                        if len(order["client_address"] or "") > 30
                        else (order["client_address"] or ""),
                        order["client_phone"],
                        created_at,
                        updated_at,
                    ]

                    for col_idx, value in enumerate(data, start=1):
                        cell = ws.cell(row=row, column=col_idx, value=value)
                        cell.border = thin_border

                        if col_idx == 1:  # ID
                            cell.alignment = center_alignment
                            cell.font = ExcelStyles.SIMPLE_BOLD_FONT
                        elif col_idx == 2:  # Статус
                            cell.alignment = center_alignment
                            # Цветовое кодирование
                            if order["status"] == "IN_PROGRESS":
                                cell.fill = ExcelStyles.HIGHLIGHT_FILL
                            elif order["status"] == "CLOSED":
                                cell.fill = ExcelStyles.SUCCESS_FILL
                            elif order["status"] == "REFUSED":
                                cell.fill = ExcelStyles.ERROR_FILL
                        else:  # Текстовые поля
                            cell.alignment = left_alignment

                    row += 1

            # Ширина столбцов для активных заявок
            widths = {
                "A": 12,  # ID
                "B": 15,  # Статус
                "C": 20,  # Тип техники
                "D": 20,  # Клиент
                "E": 35,  # Адрес
                "F": 15,  # Телефон
                "G": 18,  # Создана
                "H": 18,  # Обновлена
            }
            for col_letter, width in widths.items():
                ws.column_dimensions[col_letter].width = width

            # ==============================================
            # ВТОРОЙ ЛИСТ - ЗАВЕРШЕННЫЕ ЗАЯВКИ
            # ==============================================
            ws = ws_completed
            row = 1

            # Заголовок
            # Заполняем A1:D1 фоном
            for col in range(1, 5):  # A1:D1
                ws.cell(row=row, column=col).fill = header_fill

            cell_e1 = ws.cell(row=row, column=5)
            cell_e1.value = "ЗАВЕРШЕННЫЕ ЗАЯВКИ:"
            cell_e1.font = header_font
            cell_e1.fill = header_fill
            cell_e1.alignment = center_alignment

            cell_f1 = ws.cell(row=row, column=6)
            cell_f1.value = master_name
            cell_f1.font = header_font
            cell_f1.fill = header_fill
            cell_f1.alignment = center_alignment

            for col in range(7, 16):
                ws.cell(row=row, column=col).fill = header_fill

            ws.row_dimensions[row].height = ExcelStyles.HEADER_ROW_HEIGHT

            row += 1
            ws.merge_cells(f"A{row}:O{row}")
            cell = ws[f"A{row}"]
            cell.value = (
                f"Обновлено: {get_now().strftime('%d.%m.%Y %H:%M')} | Телефон: {master['phone']}"
            )
            cell.font = ExcelStyles.SMALL_BOLD_FONT
            cell.alignment = center_alignment

            row += 2

            # Статистика мастера
            stats_cursor = await connection.execute(
                """
                SELECT
                    COUNT(*) as total_orders,
                    SUM(CASE WHEN status = 'CLOSED' THEN 1 ELSE 0 END) as closed,
                    SUM(CASE WHEN status IN ('ASSIGNED', 'IN_PROGRESS', 'ACCEPTED') THEN 1 ELSE 0 END) as in_work,
                    SUM(CASE WHEN status = 'REFUSED' THEN 1 ELSE 0 END) as refused,
                    SUM(CASE WHEN status = 'CLOSED' THEN total_amount ELSE 0 END) as total_sum,
                    SUM(CASE WHEN status = 'CLOSED' THEN materials_cost ELSE 0 END) as materials_sum,
                    SUM(CASE WHEN status = 'CLOSED' THEN company_profit ELSE 0 END) as company_profit_sum,
                    AVG(CASE WHEN status = 'CLOSED' THEN total_amount ELSE NULL END) as avg_check
                FROM orders
                WHERE assigned_master_id = ? AND deleted_at IS NULL
                """,
                (master_id,),
            )
            stats_row = await stats_cursor.fetchone()
            stats: Mapping[str, Any] = (
                dict(stats_row)
                if stats_row is not None
                else {
                    "total_orders": 0,
                    "closed": 0,
                    "in_work": 0,
                    "refused": 0,
                    "total_sum": 0,
                    "materials_sum": 0,
                    "company_profit_sum": 0,
                    "avg_check": 0,
                }
            )

            # Блок статистики
            ws[f"A{row}"] = "СТАТИСТИКА:"
            ws[f"A{row}"].font = ExcelStyles.BOLD_FONT
            ws.merge_cells(f"A{row}:O{row}")
            row += 1

            stat_data: list[tuple[str, Any]] = [
                ("Всего заявок:", stats["total_orders"] or 0),
                ("Завершено:", stats["closed"] or 0),
                ("В работе:", stats["in_work"] or 0),
                ("Отказано:", stats["refused"] or 0),
                ("Общая сумма:", f"{float(stats['total_sum'] or 0):,.2f} ₽"),
                ("Материалы:", f"{float(stats['materials_sum'] or 0):,.2f} ₽"),
                ("Сдача в кассу:", f"{float(stats['company_profit_sum'] or 0):,.2f} ₽"),
                ("Средний чек:", f"{float(stats['avg_check'] or 0):,.2f} ₽"),
            ]

            for label, value in stat_data:
                ws[f"A{row}"] = label
                ws[f"A{row}"].font = ExcelStyles.SIMPLE_BOLD_FONT
                ws[f"B{row}"] = value
                ws[f"B{row}"].font = ExcelStyles.SIMPLE_BOLD_FONT
                ws[f"B{row}"].alignment = right_alignment
                ws.merge_cells(f"B{row}:C{row}")
                row += 1

            row += 1

            # Заголовки колонок для завершенных заявок
            headers_completed = [
                "ID",
                "Статус",
                "Тип техники",
                "Клиент",
                "Адрес",
                "Телефон",
                "Создана",
                "Обновлена",
                "Сумма",
                "Материалы",
                "Прибыль мастера",
                "Прибыль компании",
                "Выезд",
                "Отзыв",
                "Причина отказа",
            ]

            for col_idx, header in enumerate(headers_completed, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = ExcelStyles.TABLE_HEADER_FONT
                cell.fill = ExcelStyles.TABLE_HEADER_FILL
                cell.alignment = center_alignment
                cell.border = thin_border

            row += 1

            # Данные завершенных заявок
            if not completed_orders:
                ws[f"A{row}"] = "Нет завершенных заявок"
                ws[f"A{row}"].font = ExcelStyles.SIMPLE_ITALIC_FONT
                ws.merge_cells(f"A{row}:O{row}")
            else:
                for order in completed_orders:
                    status_emoji = {
                        "CLOSED": "🔒",
                        "REFUSED": "❌",
                    }.get(order["status"], "❓")

                    # Безопасное форматирование дат
                    created_at = ""
                    updated_at = ""
                    if order["created_at"]:
                        try:
                            dt = datetime.fromisoformat(order["created_at"])
                            if dt.tzinfo is not None:
                                dt = dt.replace(tzinfo=None)
                            created_at = dt.strftime("%d.%м.%Y %H:%M")
                        except Exception:
                            created_at = str(order["created_at"])[:16]
                    if order["updated_at"]:
                        try:
                            dt = datetime.fromisoformat(order["updated_at"])
                            if dt.tzinfo is not None:
                                dt = dt.replace(tzinfo=None)
                            updated_at = dt.strftime("%d.%m.%Y %H:%M")
                        except Exception:
                            updated_at = str(order["updated_at"])[:16]

                    data = [
                        order["id"],
                        f"{status_emoji} {order['status']}",
                        order["equipment_type"],
                        order["client_name"],
                        order["client_address"][:30] + "..."
                        if len(order["client_address"] or "") > 30
                        else (order["client_address"] or ""),
                        order["client_phone"],
                        created_at,
                        updated_at,
                        float(order["total_amount"] or 0),
                        float(order["materials_cost"] or 0),
                        float(order["master_profit"] or 0),
                        float(order["company_profit"] or 0),
                        "Да" if order["out_of_city"] else "",
                        "Да" if order["has_review"] else "",
                        order["refuse_reason"] or "",
                    ]

                    for col_idx, value in enumerate(data, start=1):
                        cell = ws.cell(row=row, column=col_idx, value=value)
                        cell.border = thin_border

                        if col_idx == 1:  # ID
                            cell.alignment = center_alignment
                            cell.font = ExcelStyles.SIMPLE_BOLD_FONT
                        elif col_idx == 2:  # Статус
                            cell.alignment = center_alignment
                            if order["status"] == "CLOSED":
                                cell.fill = ExcelStyles.SUCCESS_FILL
                            elif order["status"] == "REFUSED":
                                cell.fill = ExcelStyles.ERROR_FILL
                        elif col_idx in [3, 4, 5, 6, 7, 8]:  # Текстовые поля
                            cell.alignment = left_alignment
                        elif col_idx == 15:  # Причина отказа
                            cell.alignment = ExcelStyles.WRAP_TOP_ALIGNMENT
                        else:
                            cell.alignment = center_alignment if col_idx >= 13 else right_alignment
                            if col_idx >= 9 and col_idx <= 12:  # Денежные поля
                                cell.number_format = "#,##0.00 ₽"

                    row += 1

                # Итоги для завершенных
                row += 1
                ws[f"A{row}"] = "ИТОГО:"
                ws[f"A{row}"].font = ExcelStyles.BOLD_FONT
                ws.merge_cells(f"A{row}:H{row}")

                # Подсчитываем суммы
                total_sum_completed = sum(
                    float(o["total_amount"] or 0)
                    for o in completed_orders
                    if o["status"] == "CLOSED"
                )
                total_materials_completed = sum(
                    float(o["materials_cost"] or 0)
                    for o in completed_orders
                    if o["status"] == "CLOSED"
                )
                total_master_profit_completed = sum(
                    float(o["master_profit"] or 0)
                    for o in completed_orders
                    if o["status"] == "CLOSED"
                )
                total_company_profit_completed = sum(
                    float(o["company_profit"] or 0)
                    for o in completed_orders
                    if o["status"] == "CLOSED"
                )

                # Добавляем суммы на ту же строку что и "ИТОГО:"
                for cell_ref, val in [
                    (f"I{row}", total_sum_completed),
                    (f"J{row}", total_materials_completed),
                    (f"K{row}", total_master_profit_completed),
                    (f"L{row}", total_company_profit_completed),
                ]:
                    cell = ws[cell_ref]
                    cell.value = val
                    cell.font = ExcelStyles.BOLD_FONT
                    cell.number_format = "#,##0.00 ₽"
                    cell.alignment = right_alignment
                    cell.border = thin_border

                # Статистика по отказам на следующей строке
                refused_count = sum(1 for o in completed_orders if o["status"] == "REFUSED")
                refused_with_reason = sum(
                    1 for o in completed_orders if o["status"] == "REFUSED" and o["refuse_reason"]
                )
                closed_count = sum(1 for o in completed_orders if o["status"] == "CLOSED")

                row += 1
                ws[
                    f"A{row}"
                ] = f"Завершено: {closed_count} | Отказов: {refused_count} (с причиной: {refused_with_reason})"
                ws[f"A{row}"].font = ExcelStyles.ITALIC_FONT
                ws.merge_cells(f"A{row}:H{row}")

            # Ширина столбцов для завершенных заявок
            widths_completed: dict[str, int] = {
                "A": 12,
                "B": 12,
                "C": 20,
                "D": 20,
                "E": 30,
                "F": 15,
                "G": 18,
                "H": 18,
                "I": 15,
                "J": 15,
                "K": 18,
                "L": 18,
                "M": 10,
                "N": 10,
                "O": 45,  # Причина отказа
            }
            for col_letter, width in widths_completed.items():
                ws.column_dimensions[col_letter].width = width

            # Сохраняем файл
            wb.save(filepath)
            logger.info(f"Master orders Excel saved with 2 sheets: {filepath}")

            return str(filepath)

        except Exception as e:
            logger.error(f"Error exporting master orders: {e}")
            return None

        finally:
            await self.db.disconnect()

    async def _add_closed_orders_sheet(
        self,
        wb,
        report,
        thin_border,
        header_font,
        header_fill,
        subheader_fill,
        center_alignment,
        left_alignment,
        right_alignment,
    ):
        """Добавляет лист с детализацией закрытых заказов"""
        ws = wb.create_sheet(title="Закрытые заказы")

        # Заголовок
        row = 1
        ws.merge_cells(f"A{row}:K{row}")
        cell = ws[f"A{row}"]
        cell.value = "ДЕТАЛИЗАЦИЯ ЗАКРЫТЫХ ЗАКАЗОВ"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        ws.row_dimensions[row].height = ExcelStyles.HEADER_ROW_HEIGHT

        row += 1

        # Заголовки колонок
        headers = [
            "ID",
            "Техника",
            "Клиент",
            "Мастер",
            "Создано",
            "Закрыто",
            "Сумма",
            "Материалы",
            "Прибыль мастера",
            "Прибыль компании",
            "Доп. инфо",
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = ExcelStyles.SIMPLE_BOLD_FONT
            cell.fill = subheader_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        row += 1

        # Получаем все закрытые заказы за период
        connection = self._get_connection()
        cursor = await connection.execute(
            """
            SELECT
                o.id, o.equipment_type, o.client_name, o.created_at, o.updated_at,
                o.total_amount, o.materials_cost, o.master_profit, o.company_profit,
                o.out_of_city, o.has_review,
                m.first_name || ' ' || m.last_name as master_name
            FROM orders o
            LEFT JOIN masters m ON o.assigned_master_id = m.id
            WHERE o.status = 'CLOSED'
                AND o.updated_at >= ?
                AND o.updated_at <= ?
                AND o.deleted_at IS NULL
            ORDER BY o.updated_at DESC
            """,
            (report.period_start.isoformat(), report.period_end.isoformat()),
        )

        orders = await cursor.fetchall()

        if not orders:
            # Если нет заказов
            ws[f"A{row}"] = "Нет закрытых заказов за этот период"
            ws[f"A{row}"].font = ExcelStyles.SIMPLE_ITALIC_FONT
            ws.merge_cells(f"A{row}:K{row}")
        else:
            # Выводим заказы
            for order in orders:
                additional_info = []
                if order["out_of_city"]:
                    additional_info.append("Выезд за город")
                if order["has_review"]:
                    additional_info.append("Отзыв")

                data = [
                    order["id"],
                    order["equipment_type"],
                    order["client_name"],
                    order["master_name"] or "Не назначен",
                    order["created_at"][:16] if order["created_at"] else "",
                    order["updated_at"][:16] if order["updated_at"] else "",
                    float(order["total_amount"] or 0),
                    float(order["materials_cost"] or 0),
                    float(order["master_profit"] or 0),
                    float(order["company_profit"] or 0),
                    ", ".join(additional_info) if additional_info else "-",
                ]

                for col_idx, value in enumerate(data, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.border = thin_border

                    if col_idx == 1:  # ID
                        cell.alignment = center_alignment
                        cell.font = ExcelStyles.SIMPLE_BOLD_FONT
                    elif col_idx in [2, 3, 4, 5, 6, 11]:  # Текстовые поля
                        cell.alignment = left_alignment
                    else:
                        cell.alignment = right_alignment
                        if col_idx >= 7 and col_idx <= 10:  # Денежные поля
                            cell.number_format = "#,##0.00 ₽"

                row += 1

            # Итоги
            row += 1
            ws[f"A{row}"] = "ИТОГО:"
            ws[f"A{row}"].font = ExcelStyles.BOLD_FONT

            total_sum = sum(float(o["total_amount"] or 0) for o in orders)
            total_materials = sum(float(o["materials_cost"] or 0) for o in orders)
            total_master_profit = sum(float(o["master_profit"] or 0) for o in orders)
            total_company_profit = sum(float(o["company_profit"] or 0) for o in orders)

            for col, val in [
                (f"G{row}", total_sum),
                (f"H{row}", total_materials),
                (f"I{row}", total_master_profit),
                (f"J{row}", total_company_profit),
            ]:
                cell = ws[col]
                cell.value = val
                cell.font = ExcelStyles.BOLD_FONT
                cell.number_format = "#,##0.00 ₽"
                cell.fill = ExcelStyles.HIGHLIGHT_FILL

        # Ширина столбцов
        widths = {
            "A": 20,  # ID - делаем шире для полного отображения
            "B": 25,
            "C": 20,
            "D": 20,
            "E": 18,
            "F": 18,
            "G": 15,
            "H": 15,
            "I": 18,
            "J": 18,
            "K": 22,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    async def _add_masters_statistics_sheet(
        self,
        wb,
        master_reports,
        thin_border,
        header_font,
        header_fill,
        subheader_fill,
        center_alignment,
        left_alignment,
        right_alignment,
    ):
        """Добавляет лист со статистикой по мастерам"""
        ws = wb.create_sheet(title="Статистика мастеров")

        # Заголовок
        row = 1
        ws.merge_cells(f"A{row}:N{row}")
        cell = ws[f"A{row}"]
        cell.value = "РАСШИРЕННАЯ СТАТИСТИКА ПО МАСТЕРАМ"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        ws.row_dimensions[row].height = ExcelStyles.HEADER_ROW_HEIGHT

        row += 1

        # Заголовки колонок
        headers = [
            "ID",
            "Мастер",
            "Заявок всего",
            "Завершено",
            "В работе",
            "Отказано",
            "Общая сумма",
            "Материалы",
            "Чистая прибыль",
            "Прибыль компании",
            "Сдача в кассу",
            "Средний чек",
            "Выездов",
            "Отзывов",
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = ExcelStyles.SIMPLE_BOLD_FONT
            cell.fill = subheader_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        row += 1

        # Данные по каждому мастеру
        for master_report in master_reports:
            master_id = master_report.master_id
            master_name = master_report.master_name

            if not master_id:
                continue

            # Получаем расширенную статистику по мастеру через ORM
            from app.core.constants import OrderStatus

            orders = await self.db.get_orders_by_master(master_id, exclude_closed=False)

            # Подсчитываем статистику
            total_orders = len(orders)
            closed = len([o for o in orders if o.status == OrderStatus.CLOSED])
            in_work = len(
                [
                    o
                    for o in orders
                    if o.status in [OrderStatus.ASSIGNED, OrderStatus.ACCEPTED, OrderStatus.ONSITE]
                ]
            )
            refused = len([o for o in orders if o.status == OrderStatus.REFUSED])

            total_sum = sum(o.total_amount or 0 for o in orders if o.status == OrderStatus.CLOSED)
            materials_sum = sum(
                o.materials_cost or 0 for o in orders if o.status == OrderStatus.CLOSED
            )
            company_profit_sum = sum(
                o.company_profit or 0 for o in orders if o.status == OrderStatus.CLOSED
            )

            out_of_city_count = sum(
                1 for o in orders if o.status == OrderStatus.CLOSED and o.out_of_city is True
            )
            reviews_count = sum(
                1 for o in orders if o.status == OrderStatus.CLOSED and o.has_review is True
            )

            # Средний чек
            closed_amounts = [
                o.total_amount for o in orders if o.status == OrderStatus.CLOSED and o.total_amount
            ]
            avg_check = sum(closed_amounts) / len(closed_amounts) if closed_amounts else 0

            # Вычисляем данные
            materials = float(materials_sum)
            net_profit = total_sum - materials
            cash_to_company = float(company_profit_sum)

            # Данные по мастеру
            master_data = [
                master_id,
                master_name,
                total_orders,
                closed,
                in_work,
                refused,
                total_sum,
                materials,
                net_profit,
                cash_to_company,
                cash_to_company,  # Сдача в кассу = прибыль компании
                float(avg_check),
                out_of_city_count,
                reviews_count,
            ]

            for col_idx, value in enumerate(master_data, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = thin_border

                # Форматирование
                if col_idx == 1:  # ID
                    cell.alignment = center_alignment
                    cell.font = ExcelStyles.SIMPLE_BOLD_FONT
                elif col_idx == 2:  # Имя
                    cell.alignment = left_alignment
                    cell.font = ExcelStyles.SIMPLE_BOLD_FONT
                elif col_idx in [3, 4, 5, 6, 13, 14]:  # Счетчики
                    cell.alignment = center_alignment
                else:  # Денежные поля
                    cell.alignment = right_alignment
                    if col_idx >= 7 and col_idx <= 12:
                        cell.number_format = "#,##0.00 ₽"

            row += 1

        # Добавляем новую колонку "Отказов с причиной"
        ws.cell(row=2, column=15, value="Отказов с причиной")
        ws.cell(row=2, column=15).font = ExcelStyles.SIMPLE_BOLD_FONT
        ws.cell(row=2, column=15).fill = subheader_fill
        ws.cell(row=2, column=15).alignment = center_alignment
        ws.cell(row=2, column=15).border = thin_border

        # Пересчитываем для добавления колонки "Отказов с причиной"
        row_idx = 3
        for master_report in master_reports:
            master_id = master_report.master_id
            if not master_id:
                continue

            from app.core.constants import OrderStatus

            orders = await self.db.get_orders_by_master(master_id, exclude_closed=False)

            # Подсчитываем отказы с причиной
            refused_with_reason = len(
                [o for o in orders if o.status == OrderStatus.REFUSED and o.refuse_reason]
            )

            cell = ws.cell(row=row_idx, column=15, value=refused_with_reason)
            cell.border = thin_border
            cell.alignment = center_alignment

            row_idx += 1

        # ИТОГО по всем мастерам
        row += 1
        ws[f"A{row}"] = "ИТОГО:"
        ws[f"A{row}"].font = ExcelStyles.SUBHEADER_FONT
        ws[f"A{row}"].fill = ExcelStyles.HIGHLIGHT_FILL
        ws.merge_cells(f"A{row}:B{row}")

        # Суммы по всем мастерам
        connection = self._get_connection()
        cursor = await connection.execute(
            """
            SELECT
                COUNT(*) as total_orders,
                SUM(CASE WHEN status = 'CLOSED' THEN 1 ELSE 0 END) as closed,
                SUM(CASE WHEN status IN ('ASSIGNED', 'IN_PROGRESS', 'ACCEPTED') THEN 1 ELSE 0 END) as in_work,
                SUM(CASE WHEN status = 'REFUSED' THEN 1 ELSE 0 END) as refused,
                SUM(CASE WHEN status = 'CLOSED' THEN total_amount ELSE 0 END) as total_sum,
                SUM(CASE WHEN status = 'CLOSED' THEN materials_cost ELSE 0 END) as materials_sum,
                SUM(CASE WHEN status = 'CLOSED' THEN company_profit ELSE 0 END) as company_profit_sum,
                AVG(CASE WHEN status = 'CLOSED' THEN total_amount ELSE NULL END) as avg_check,
                SUM(CASE WHEN status = 'CLOSED' AND out_of_city = 1 THEN 1 ELSE 0 END) as out_of_city,
                SUM(CASE WHEN status = 'CLOSED' AND has_review = 1 THEN 1 ELSE 0 END) as reviews
            FROM orders
            WHERE assigned_master_id IS NOT NULL
                AND deleted_at IS NULL
            """
        )

        totals_row = await cursor.fetchone()
        totals: Mapping[str, Any] = (
            dict(totals_row)
            if totals_row is not None
            else {
                "total_orders": 0,
                "closed": 0,
                "in_work": 0,
                "refused": 0,
                "total_sum": 0,
                "materials_sum": 0,
                "company_profit_sum": 0,
                "avg_check": 0,
                "out_of_city": 0,
                "reviews": 0,
            }
        )

        total_sum = float(totals["total_sum"] or 0)
        materials_sum = float(totals["materials_sum"] or 0)
        net_profit_total = total_sum - materials_sum

        totals_data = [
            totals["total_orders"],
            totals["closed"],
            totals["in_work"],
            totals["refused"],
            total_sum,
            materials_sum,
            net_profit_total,
            float(totals["company_profit_sum"] or 0),
            float(totals["company_profit_sum"] or 0),
            float(totals["avg_check"] or 0),
            totals["out_of_city"],
            totals["reviews"],
        ]

        for col_idx, value in enumerate(totals_data, start=3):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = ExcelStyles.BOLD_FONT
            cell.fill = ExcelStyles.HIGHLIGHT_FILL
            cell.border = thin_border

            if col_idx in [3, 4, 5, 6, 13, 14]:  # Счетчики
                cell.alignment = center_alignment
            else:
                cell.alignment = right_alignment
                if col_idx >= 7 and col_idx <= 12:
                    cell.number_format = "#,##0.00 ₽"

        # Ширина столбцов
        widths = {
            "A": 20,  # ID - делаем шире для полного отображения
            "B": 25,
            "C": 12,
            "D": 12,
            "E": 12,
            "F": 12,
            "G": 15,
            "H": 15,
            "I": 15,
            "J": 18,
            "K": 15,
            "L": 15,
            "M": 10,
            "N": 10,
            "O": 18,  # Новая колонка
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    async def _add_refusals_details_sheet(
        self,
        wb,
        master_reports,
        thin_border,
        header_font,
        header_fill,
        subheader_fill,
        center_alignment,
        left_alignment,
    ):
        """Добавляет лист с детальной информацией об отказах и причинах"""
        ws = wb.create_sheet(title="Причины отказов")

        # Заголовок
        row = 1
        ws.merge_cells(f"A{row}:F{row}")
        cell = ws[f"A{row}"]
        cell.value = "ДЕТАЛЬНАЯ ИНФОРМАЦИЯ ОБ ОТКАЗАХ"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        ws.row_dimensions[row].height = ExcelStyles.HEADER_ROW_HEIGHT

        row += 1

        # Заголовки колонок
        headers = [
            "ID заявки",
            "Мастер",
            "Клиент",
            "Адрес",
            "Дата отказа",
            "Причина отказа",
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = ExcelStyles.SIMPLE_BOLD_FONT
            cell.fill = subheader_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        row += 1

        # Получаем все отказы с причинами
        from app.core.constants import OrderStatus

        for master_report in master_reports:
            master_id = master_report.master_id
            master_name = master_report.master_name

            if not master_id:
                continue

            # Получаем отказанные заявки мастера
            orders = await self.db.get_orders_by_master(master_id, exclude_closed=False)
            refused_orders = [
                o for o in orders if o.status == OrderStatus.REFUSED and o.refuse_reason
            ]

            for order in refused_orders:
                order_data = [
                    order.id,
                    master_name,
                    order.client_name,
                    order.client_address,
                    order.updated_at.strftime("%d.%m.%Y %H:%M") if order.updated_at else "",
                    order.refuse_reason or "Не указана",
                ]

                for col_idx, value in enumerate(order_data, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.border = thin_border

                    # Форматирование
                    if col_idx == 1:  # ID
                        cell.alignment = center_alignment
                    elif col_idx == 6:  # Причина
                        cell.alignment = left_alignment
                    else:
                        cell.alignment = left_alignment

                row += 1

        # Если нет отказов с причинами
        if row == 3:
            ws.cell(row=row, column=1, value="Нет отказов с указанными причинами")
            ws.merge_cells(f"A{row}:F{row}")
            ws.cell(row=row, column=1).alignment = center_alignment

        # Ширина столбцов
        widths = {
            "A": 12,  # ID заявки
            "B": 25,  # Мастер
            "C": 25,  # Клиент
            "D": 35,  # Адрес
            "E": 18,  # Дата отказа
            "F": 50,  # Причина отказа
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
