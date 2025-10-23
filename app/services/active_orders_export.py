"""
Сервис для экспорта активных (незакрытых) заявок в Excel
"""

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.config import OrderStatus
from app.database.db import Database
from app.utils.helpers import get_now


logger = logging.getLogger(__name__)


class ActiveOrdersExportService:
    """Сервис для экспорта активных заявок в Excel"""

    def __init__(self):
        self.db = Database()

    async def export_active_orders_to_excel(self) -> str | None:
        """
        Экспорт всех активных (незакрытых) заявок в Excel

        Returns:
            Путь к созданному файлу или None
        """
        await self.db.connect()

        try:
            # Получаем все активные заявки (не CLOSED и не REFUSED)
            all_orders = await self.db.get_all_orders()
            active_orders = [
                order
                for order in all_orders
                if order.status not in [OrderStatus.CLOSED, OrderStatus.REFUSED]
            ]

            if not active_orders:
                logger.info("No active orders found")
                return None

            # Создаем Excel файл
            wb = Workbook()
            
            # Удаляем стандартный лист
            wb.remove(wb.active)
            
            # Создаем сводный лист
            summary_sheet = wb.create_sheet("Сводка")
            await self._create_summary_sheet(summary_sheet, active_orders)
            
            # Получаем всех активных и одобренных мастеров
            masters = await self.db.get_all_masters(only_approved=True, only_active=True)
            
            # Создаем листы для каждого мастера (только если есть активные заказы)
            for master in masters:
                master_orders = [o for o in active_orders if o.assigned_master_id == master.id]
                if master_orders:  # Создаем лист только если есть активные заказы
                    master_sheet = wb.create_sheet(master.get_display_name())
                    await self._create_master_sheet(master_sheet, master, master_orders)


            # Создаем директорию для отчетов
            reports_dir = Path("reports")
            reports_dir.mkdir(exist_ok=True)

            # Генерируем имя файла
            timestamp = get_now().strftime("%Y%m%d_%H%M%S")
            filename = f"active_orders_{timestamp}.xlsx"
            filepath = reports_dir / filename

            # Сохраняем файл
            wb.save(filepath)
            logger.info(f"Active orders Excel report saved: {filepath}")

            return str(filepath)

        except Exception as e:
            logger.error(f"Error exporting active orders to Excel: {e}")
            return None

        finally:
            await self.db.disconnect()

    async def _create_summary_sheet(self, ws, active_orders: list):
        """Создание сводного листа с общей информацией"""
        # Стили
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        data_font = Font(size=10)
        
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Заголовок
        ws.merge_cells("A1:K1")
        cell = ws["A1"]
        cell.value = f"АКТИВНЫЕ ЗАЯВКИ - {get_now().strftime('%d.%m.%Y %H:%M')}"
        cell.font = Font(bold=True, size=14)
        cell.fill = header_fill
        cell.alignment = center_alignment
        ws.row_dimensions[1].height = 25

        # Общая информация
        ws.merge_cells("A2:K2")
        cell = ws["A2"]
        cell.value = f"Всего активных заявок: {len(active_orders)}"
        cell.font = Font(bold=True, size=11)
        cell.alignment = center_alignment
        ws.row_dimensions[2].height = 20

        # Заголовки столбцов
        headers = [
            "№ Заявки",
            "Статус",
            "Оборудование",
            "Описание",
            "Клиент",
            "Адрес",
            "Телефон",
            "Мастер",
            "Диспетчер",
            "Время прибытия",
            "Создана",
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # Сортируем заявки по статусу и дате создания
        status_priority = {
            OrderStatus.NEW: 1,
            OrderStatus.ASSIGNED: 2,
            OrderStatus.ACCEPTED: 3,
            OrderStatus.ONSITE: 4,
            OrderStatus.DR: 5,
        }

        sorted_orders = sorted(
            active_orders,
            key=lambda x: (status_priority.get(x.status, 99), x.created_at or datetime.min),
        )

        # Данные по заявкам
        row = 5
        for order in sorted_orders:
            status_name = OrderStatus.get_status_name(order.status)
            status_emoji = OrderStatus.get_status_emoji(order.status)

            # Форматируем дату создания
            created_str = ""
            if order.created_at:
                if isinstance(order.created_at, str):
                    try:
                        created_dt = datetime.fromisoformat(
                            order.created_at.replace("Z", "+00:00")
                        )
                        created_str = created_dt.strftime("%d.%m.%Y %H:%M")
                    except:
                        created_str = order.created_at
                else:
                    created_str = order.created_at.strftime("%d.%m.%Y %H:%M")

            data = [
                order.id,
                f"{status_emoji} {status_name}",
                order.equipment_type or "",
                order.description or "",
                order.client_name or "",
                order.client_address or "",
                order.client_phone or "",
                order.master_name or "Не назначен",
                order.dispatcher_name or "",
                order.scheduled_time if order.scheduled_time else "",
                created_str,
            ]

            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = left_alignment if col_idx not in [1, 2] else center_alignment

            row += 1

        # Устанавливаем ширину столбцов
        column_widths = {
            "A": 10,  # № Заявки
            "B": 18,  # Статус
            "C": 20,  # Оборудование
            "D": 30,  # Описание
            "E": 20,  # Клиент
            "F": 30,  # Адрес
            "G": 15,  # Телефон
            "H": 20,  # Мастер
            "I": 20,  # Диспетчер
            "J": 20,  # Время прибытия
            "K": 18,  # Создана
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    async def _create_master_sheet(self, ws, master, orders: list):
        """Создание листа для конкретного мастера"""
        # Стили
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        data_font = Font(size=10)
        
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Заголовок
        # A1: "Активные заказы мастера:"
        cell_a1 = ws.cell(row=1, column=1)
        cell_a1.value = "Активные заказы мастера:"
        cell_a1.font = header_font
        cell_a1.fill = header_fill
        cell_a1.alignment = center_alignment
        
        # B1: имя мастера
        cell_b1 = ws.cell(row=1, column=2)
        cell_b1.value = master.get_display_name()
        cell_b1.font = header_font
        cell_b1.fill = header_fill
        cell_b1.alignment = center_alignment
        
        ws.row_dimensions[1].height = 25
        # Растягиваем заголовок на остальные столбцы
        for col in range(3, 8):  # C1:G1
            ws.cell(row=1, column=col).fill = header_fill

        # Заголовки столбцов (упрощенная структура для мастера)
        headers = ["№ Заказа", "Статус", "Оборудование", "Клиент", "Адрес", "Время прибытия", "Создана"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # Данные по заказам
        row = 4
        for order in orders:
            status_name = OrderStatus.get_status_name(order.status)
            status_emoji = OrderStatus.get_status_emoji(order.status)

            # Форматируем дату создания
            created_str = ""
            if order.created_at:
                if isinstance(order.created_at, str):
                    try:
                        created_dt = datetime.fromisoformat(
                            order.created_at.replace("Z", "+00:00")
                        )
                        created_str = created_dt.strftime("%d.%m.%Y %H:%M")
                    except:
                        created_str = order.created_at
                else:
                    created_str = order.created_at.strftime("%d.%m.%Y %H:%M")

            data = [
                order.id,
                f"{status_emoji} {status_name}",
                order.equipment_type or "",
                order.client_name or "",
                order.client_address or "",
                order.scheduled_time if order.scheduled_time else "",
                created_str,
            ]
            
            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = left_alignment if col_idx not in [1, 2] else center_alignment
            
            row += 1

        # Устанавливаем ширину столбцов
        column_widths = {"A": 12, "B": 18, "C": 20, "D": 20, "E": 30, "F": 20, "G": 18}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
