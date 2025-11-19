"""
Сервис для генерации отчетов по мастерам с детализацией
"""

import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.config import OrderStatus
from app.database import DatabaseType, get_database
from app.utils.helpers import MOSCOW_TZ, get_now


logger = logging.getLogger(__name__)


class MasterReportsService:
    """Сервис для генерации детализированных отчетов по мастерам"""

    def __init__(self) -> None:
        self.db: DatabaseType = get_database()

    async def generate_daily_master_report(self, target_date: datetime) -> str | None:
        """
        Генерация ежедневного отчета по мастерам

        Args:
            target_date: Дата для отчета

        Returns:
            Путь к созданному файлу или None
        """
        return await self._generate_master_report(target_date, "daily")

    async def generate_weekly_master_report(self, week_start: datetime) -> str | None:
        """
        Генерация еженедельного отчета по мастерам

        Args:
            week_start: Начало недели (понедельник)

        Returns:
            Путь к созданному файлу или None
        """
        return await self._generate_master_report(week_start, "weekly")

    async def generate_monthly_master_report(self, month_start: datetime) -> str | None:
        """
        Генерация ежемесячного отчета по мастерам

        Args:
            month_start: Начало месяца

        Returns:
            Путь к созданному файлу или None
        """
        return await self._generate_master_report(month_start, "monthly")

    async def _generate_master_report(self, start_date: datetime, report_type: str) -> str | None:
        """
        Генерация отчета по мастерам

        Args:
            start_date: Начальная дата
            report_type: Тип отчета (daily, weekly, monthly)

        Returns:
            Путь к созданному файлу или None
        """
        await self.db.connect()

        try:
            # Определяем период
            if report_type == "daily":
                end_date = start_date + timedelta(days=1)
                period_name = "Ежедневный отчет"
            elif report_type == "weekly":
                end_date = start_date + timedelta(days=7)
                period_name = "Еженедельный отчет"
            else:  # monthly
                if start_date.month == 12:
                    end_date = start_date.replace(year=start_date.year + 1, month=1)
                else:
                    end_date = start_date.replace(month=start_date.month + 1)
                period_name = "Ежемесячный отчет"

            # Получаем закрытые заказы за период
            orders = await self._get_closed_orders_in_period(start_date, end_date)

            if not orders:
                logger.info(f"No closed orders found for {report_type} report")
                return None

            # Создаем Excel файл
            wb = Workbook()

            # Удаляем стандартный лист
            wb.remove(wb.active)

            # Создаем сводный лист
            summary_sheet = wb.create_sheet("Сводка")
            await self._create_summary_sheet(
                summary_sheet, orders, period_name, start_date, end_date, report_type
            )

            # Получаем всех активных и одобренных мастеров
            masters = await self.db.get_all_masters(only_approved=True, only_active=True)

            # Создаем листы для каждого мастера (только если есть заказы)
            for master in masters:
                master_orders = [o for o in orders if o.assigned_master_id == master.id]
                if master_orders:  # Создаем лист только если есть заказы
                    master_sheet = wb.create_sheet(master.get_display_name())
                    await self._create_master_sheet(
                        master_sheet, master, master_orders, report_type
                    )

            # Создаем директорию для отчетов
            reports_dir = Path("reports")
            reports_dir.mkdir(exist_ok=True)

            # Генерируем имя файла
            timestamp = get_now().strftime("%Y%m%d_%H%M%S")
            filename = f"master_report_{report_type}_{timestamp}.xlsx"
            filepath = reports_dir / filename

            # Сохраняем файл
            wb.save(filepath)
            logger.info(f"Master {report_type} report saved: {filepath}")

            return str(filepath)

        except Exception as e:
            logger.error(f"Error generating master {report_type} report: {e}")
            return None

        finally:
            await self.db.disconnect()

    async def _get_closed_orders_in_period(self, start_date: datetime, end_date: datetime) -> list:
        """Получение закрытых заказов за период"""
        # Убеждаемся, что даты имеют часовой пояс
        if start_date.tzinfo is None:
            start_date = start_date.replace(tzinfo=MOSCOW_TZ)
        if end_date.tzinfo is None:
            end_date = end_date.replace(tzinfo=MOSCOW_TZ)

        all_orders = await self.db.get_all_orders()

        closed_orders = []
        for order in all_orders:
            if order.status == OrderStatus.CLOSED and order.updated_at:
                # Убеждаемся, что order.updated_at имеет часовой пояс
                order_updated_at = order.updated_at
                if order_updated_at.tzinfo is None:
                    order_updated_at = order_updated_at.replace(tzinfo=MOSCOW_TZ)

                if start_date <= order_updated_at < end_date:
                    closed_orders.append(order)

        return closed_orders

    async def _create_summary_sheet(
        self,
        ws,
        orders: list,
        period_name: str,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
        report_type: str = "daily",
    ):
        """Создание сводного листа"""
        # Стили
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        data_font = Font(size=11)
        data_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        total_font = Font(bold=True, size=12)
        total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Заголовок
        cell = ws.cell(row=1, column=1)
        cell.value = period_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        ws.row_dimensions[1].height = 30
        # Определяем количество столбцов в зависимости от типа отчета
        is_monthly = report_type == "monthly"
        num_cols = (
            6 if is_monthly else 5
        )  # 6 столбцов для ежемесячной (с материалами), 5 для остальных

        # Растягиваем заголовок на несколько столбцов
        for col in range(2, num_cols + 1):
            ws.cell(row=1, column=col).fill = header_fill

        # Заголовки столбцов
        if is_monthly:
            headers = [
                "Мастер",
                "Кол-во заявок",
                "Сумма заявок",
                "Материалы",
                "Сумма к сдаче",
                "Средний чек",
            ]
        else:
            headers = ["Мастер", "Кол-во заявок", "Сумма заявок", "Сумма к сдаче", "Средний чек"]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # Группируем заказы по мастерам
        master_stats: dict[int | None, dict[str, Any]] = {}
        for order in orders:
            master_id = order.assigned_master_id
            if master_id not in master_stats:
                master_stats[master_id] = {
                    "name": order.master_name or "Неизвестен",
                    "orders": [],
                    "total_amount": 0,
                    "materials_cost": 0,
                    "handover_amount": 0,
                }

            master_stats[master_id]["orders"].append(order)
            # Сумма заказов включает расходный материал
            order_amount = (order.total_amount or 0) + (order.materials_cost or 0)
            master_stats[master_id]["total_amount"] += order_amount
            master_stats[master_id]["materials_cost"] += order.materials_cost or 0
            master_stats[master_id]["handover_amount"] += order.company_profit or 0

        # Сортируем мастеров по сумме заказов (по убыванию)
        sorted_masters = sorted(
            master_stats.items(), key=lambda x: x[1]["total_amount"], reverse=True
        )

        # Заполняем данные
        row = 4
        total_orders = 0
        total_amount = 0
        total_materials = 0
        total_handover = 0

        for _master_id, stats in sorted_masters:
            order_count = len(stats["orders"])
            avg_check = stats["total_amount"] / order_count if order_count > 0 else 0

            if is_monthly:
                data = [
                    stats["name"],
                    order_count,
                    stats["total_amount"],
                    stats["materials_cost"],
                    stats["handover_amount"],
                    avg_check,
                ]
            else:
                data = [
                    stats["name"],
                    order_count,
                    stats["total_amount"],
                    stats["handover_amount"],
                    avg_check,
                ]

            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = data_font
                cell.fill = data_fill
                cell.border = thin_border
                cell.alignment = left_alignment if col_idx == 1 else center_alignment

                # Форматируем числа
                if is_monthly:
                    if col_idx in [3, 4, 5, 6]:  # Суммы (включая материалы)
                        cell.number_format = "#,##0.00"
                elif col_idx in [3, 4, 5]:  # Суммы
                    cell.number_format = "#,##0.00"

            total_orders += order_count
            total_amount += stats["total_amount"]
            total_materials += stats["materials_cost"]
            total_handover += stats["handover_amount"]
            row += 1

        # Итоговая строка
        total_avg = total_amount / total_orders if total_orders > 0 else 0

        # Период - только даты без подписей
        start_date_str = start_date.strftime("%d.%m.%Y") if start_date else ""
        end_date_str = end_date.strftime("%d.%m.%Y") if end_date else ""
        ws.cell(row=row + 1, column=1, value=start_date_str).font = data_font
        ws.cell(row=row + 2, column=1, value=end_date_str).font = data_font

        ws.cell(row=row + 1, column=3, value="Общее").font = total_font
        ws.cell(row=row + 1, column=3).fill = total_fill

        if is_monthly:
            ws.cell(row=row + 1, column=4, value="")
            ws.cell(row=row + 1, column=5, value="")
            ws.cell(row=row + 1, column=6, value="")
            total_data = [total_orders, total_amount, total_materials, total_handover, total_avg]
        else:
            ws.cell(row=row + 1, column=4, value="")
            ws.cell(row=row + 1, column=5, value="")
            total_data = [total_orders, total_amount, total_handover, total_avg]

        for col_idx, value in enumerate(total_data, start=2):
            cell = ws.cell(row=row + 2, column=col_idx, value=value)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.number_format = "#,##0.00"

        # Статистика по типам техники
        row += 4  # Пустая строка для разделения
        equipment_stats: dict[str, int] = {}
        for order in orders:
            equipment_type = order.equipment_type or "Не указано"
            equipment_stats[equipment_type] = equipment_stats.get(equipment_type, 0) + 1

        logger.info(f"Equipment stats for {len(orders)} orders: {equipment_stats}")

        # Показываем статистику всегда, даже если все заказы без типа техники
        if orders:  # Показываем статистику, если есть заказы
            # Заголовок секции
            subheader_font = Font(bold=True, size=12, color="FFFFFF")
            subheader_fill = PatternFill(
                start_color="70AD47", end_color="70AD47", fill_type="solid"
            )

            cell = ws.cell(row=row, column=1, value="По типам техники:")
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.alignment = left_alignment
            # Объединяем ячейки в зависимости от количества столбцов
            if is_monthly:
                ws.merge_cells(f"A{row}:F{row}")
            else:
                ws.merge_cells(f"A{row}:E{row}")
            row += 1

            # Заголовки столбцов
            equipment_headers = ["Тип техники", "Количество", "Процент", "", ""]
            for col_idx, header in enumerate(equipment_headers, start=1):
                if header:  # Пропускаем пустые столбцы
                    cell = ws.cell(row=row, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = thin_border
            row += 1

            # Сортируем по количеству (по убыванию)
            sorted_equipment = sorted(equipment_stats.items(), key=lambda x: x[1], reverse=True)

            for equipment_type, count in sorted_equipment:
                percentage = (count / len(orders) * 100) if len(orders) > 0 else 0

                data = [
                    equipment_type,
                    count,
                    percentage,
                    "",  # Пустой столбец
                    "",  # Пустой столбец
                ]

                for col_idx, value in enumerate(data, start=1):
                    if value != "":  # Пропускаем пустые столбцы
                        cell = ws.cell(row=row, column=col_idx, value=value)
                        cell.font = data_font
                        cell.fill = data_fill
                        cell.border = thin_border
                        cell.alignment = left_alignment if col_idx == 1 else center_alignment

                        # Форматируем процент
                        if col_idx == 3:
                            cell.number_format = "0.0"

                row += 1

        # Устанавливаем ширину столбцов (увеличиваем для лучшего отображения)
        column_widths: dict[str, int]
        if is_monthly:
            column_widths = {"A": 25, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18}
        else:
            column_widths = {"A": 25, "B": 18, "C": 18, "D": 18, "E": 18}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

    async def _create_master_sheet(self, ws, master, orders: list, report_type: str = "daily"):
        """Создание листа для конкретного мастера"""
        # Стили
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        data_font = Font(size=10)

        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Заголовок
        # A1: "Заказы мастера:"
        cell_a1 = ws.cell(row=1, column=1)
        cell_a1.value = "Заказы мастера:"
        cell_a1.font = header_font
        cell_a1.fill = header_fill
        cell_a1.alignment = center_alignment

        # B1: имя мастера
        cell_b1 = ws.cell(row=1, column=2)
        cell_b1.value = master.get_display_name()
        cell_b1.font = header_font
        cell_b1.fill = header_fill
        cell_b1.alignment = center_alignment

        ws.row_dimensions[1].height = 25
        # Определяем заголовки в зависимости от типа отчета
        if report_type == "monthly":
            # Для ежемесячных отчетов добавляем дату выполнения и материалы
            headers = [
                "№ Заказа",
                "Сумма",
                "Материалы",
                "Тип техники",
                "Сумма к сдаче",
                "Дата выполнения",
                "Выезд",
                "Отзыв",
            ]
            # Растягиваем заголовок на остальные столбцы
            for col in range(3, 9):  # C1:H1
                ws.cell(row=1, column=col).fill = header_fill
        elif report_type == "weekly":
            # Для еженедельных отчетов добавляем дату выполнения
            headers = [
                "№ Заказа",
                "Сумма",
                "Тип техники",
                "Сумма к сдаче",
                "Дата выполнения",
                "Выезд",
                "Отзыв",
            ]
            # Растягиваем заголовок на остальные столбцы
            for col in range(3, 8):  # C1:G1
                ws.cell(row=1, column=col).fill = header_fill
        else:
            # Для ежедневных отчетов без даты выполнения
            headers = ["№ Заказа", "Сумма", "Тип техники", "Сумма к сдаче", "Выезд", "Отзыв"]
            # Растягиваем заголовок на остальные столбцы
            for col in range(3, 7):  # C1:F1
                ws.cell(row=1, column=col).fill = header_fill
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # Данные по заказам
        row = 4
        for order in orders:
            if report_type == "monthly":
                # Для ежемесячных отчетов добавляем дату выполнения и материалы
                completion_date = ""
                if order.updated_at:
                    completion_date = order.updated_at.strftime("%d.%m.%Y %H:%M")

                # Сумма заказов включает расходный материал (как в сводке)
                total_order_amount = (order.total_amount or 0) + (order.materials_cost or 0)

                data = [
                    order.id,
                    total_order_amount,
                    order.materials_cost or 0,
                    order.equipment_type or "",
                    order.company_profit or 0,
                    completion_date,
                    "✅" if order.out_of_city else "❌",
                    "✅" if order.has_review else "❌",
                ]
            elif report_type == "weekly":
                # Для еженедельных отчетов добавляем дату выполнения
                completion_date = ""
                if order.updated_at:
                    completion_date = order.updated_at.strftime("%d.%m.%Y %H:%M")

                # Сумма заказов включает расходный материал (как в сводке)
                total_order_amount = (order.total_amount or 0) + (order.materials_cost or 0)

                data = [
                    order.id,
                    total_order_amount,
                    order.equipment_type or "",
                    order.company_profit or 0,
                    completion_date,
                    "✅" if order.out_of_city else "❌",
                    "✅" if order.has_review else "❌",
                ]
            else:
                # Для ежедневных отчетов без даты выполнения
                data = [
                    order.id,
                    order.total_amount or 0,
                    order.equipment_type or "",
                    order.company_profit or 0,
                    "✅" if order.out_of_city else "❌",
                    "✅" if order.has_review else "❌",
                ]

            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                # Для ежемесячных: колонка 1 (№ Заказа) и 4 (Тип техники) - выравнивание по левому краю
                # Для еженедельных и ежедневных: колонка 1 (№ Заказа) и 3 (Тип техники) - выравнивание по левому краю
                if report_type == "monthly":
                    cell.alignment = left_alignment if col_idx in [1, 4] else center_alignment
                else:
                    cell.alignment = left_alignment if col_idx in [1, 3] else center_alignment

                # Форматируем числа
                if report_type == "monthly":
                    if col_idx in [2, 3, 5]:  # Сумма, Материалы, Сумма к сдаче
                        cell.number_format = "#,##0.00"
                elif col_idx in [2, 4]:  # Суммы
                    cell.number_format = "#,##0.00"

            row += 1

        # Устанавливаем ширину столбцов (увеличиваем для лучшего отображения)
        column_widths: dict[str, int]
        if report_type == "monthly":
            column_widths = {"A": 16, "B": 18, "C": 18, "D": 25, "E": 18, "F": 20, "G": 12, "H": 12}
        elif report_type == "weekly":
            column_widths = {"A": 16, "B": 18, "C": 25, "D": 18, "E": 20, "F": 12, "G": 12}
        else:
            column_widths = {"A": 16, "B": 18, "C": 25, "D": 18, "E": 12, "F": 12}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
