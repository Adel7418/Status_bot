"""
Сервис для управления таблицей активных заказов в реальном времени
"""

import logging
import os
from pathlib import Path

from app.database import Database
from app.services.active_orders_export import ActiveOrdersExportService


logger = logging.getLogger(__name__)


class RealtimeActiveOrdersService:
    """Сервис для управления таблицей активных заказов в реальном времени"""

    def __init__(self):
        self.db = Database()
        self.export_service = ActiveOrdersExportService()
        self.current_table_path = None

    async def init(self):
        """Инициализация сервиса"""
        # ORM база данных не требует init_db
        if hasattr(self.db, "init_db"):
            await self.db.init_db()

        # Проверяем, есть ли текущая таблица активных заказов
        await self._ensure_current_table_exists()

    async def _ensure_current_table_exists(self):
        """Убеждаемся, что текущая таблица активных заказов существует"""
        table_filename = "active_orders_current.xlsx"
        reports_dir = Path("reports")
        reports_dir.mkdir(exist_ok=True)
        table_path = reports_dir / table_filename

        if not os.path.exists(table_path):
            # Создаем новую таблицу
            logger.info("Создание новой таблицы активных заказов")
            await self._create_active_orders_table()
        else:
            logger.info(f"Таблица активных заказов уже существует: {table_path}")

        self.current_table_path = str(table_path)

    async def _create_active_orders_table(self):
        """Создает новую таблицу активных заказов"""
        # Создаем отчет
        report_path = await self.export_service.export_active_orders_to_excel()

        # Переименовываем файл в стандартное имя
        table_filename = "active_orders_current.xlsx"
        reports_dir = Path("reports")
        reports_dir.mkdir(exist_ok=True)
        new_path = reports_dir / table_filename

        if report_path and os.path.exists(report_path):
            # Если файл уже существует, удаляем его
            if os.path.exists(new_path):
                os.remove(new_path)
            os.rename(report_path, new_path)
            logger.info(f"Таблица активных заказов создана: {new_path}")
        else:
            # Нет активных заявок или произошла тихая ошибка экспорта.
            # Создаём пустую таблицу по умолчанию, чтобы downstream-логика не падала.
            try:
                from openpyxl import Workbook

                wb = Workbook()
                ws = wb.active
                ws.title = "Сводка"
                ws["A1"] = "АКТИВНЫЕ ЗАЯВКИ"
                ws["A2"] = "Всего активных заявок: 0"

                # Если файл уже существует, удаляем его
                if os.path.exists(new_path):
                    os.remove(new_path)
                wb.save(new_path)
                logger.info(
                    f"Создана пустая таблица активных заказов (нет активных заявок): {new_path}"
                )
            except Exception as e:
                logger.error(
                    f"Не удалось создать таблицу активных заказов (fallback) {new_path}: {e}. Источник: {report_path}"
                )

    async def update_table(self):
        """Обновляет текущую таблицу активных заказов"""
        try:
            if not self.current_table_path:
                logger.warning("Текущая таблица не инициализирована, пропускаем обновление.")
                return

            logger.info("Обновление таблицы активных заказов")
            await self._update_current_table()

        except Exception as e:
            logger.error(f"Ошибка при обновлении таблицы активных заказов: {e}")

    async def _update_current_table(self):
        """Обновляет текущую таблицу активных заказов"""
        if not self.current_table_path:
            logger.error("Невозможно обновить таблицу: путь не установлен.")
            return

        # Создаем новую версию таблицы
        report_path = await self.export_service.export_active_orders_to_excel()

        if report_path and os.path.exists(report_path):
            # Заменяем текущую таблицу
            if os.path.exists(self.current_table_path):
                os.remove(self.current_table_path)
            os.rename(report_path, self.current_table_path)
            logger.info(f"Таблица активных заказов обновлена: {self.current_table_path}")
        else:
            logger.error(f"Не удалось обновить таблицу: {report_path}")

    async def get_current_table_path(self) -> str | None:
        """Возвращает путь к текущей таблице активных заказов"""
        if self.current_table_path and os.path.exists(self.current_table_path):
            return self.current_table_path
        return None


# Глобальный экземпляр сервиса
realtime_active_orders_service = RealtimeActiveOrdersService()
