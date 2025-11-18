"""
Сервис для управления ежедневными таблицами в реальном времени
"""

import logging
import os
from datetime import date, datetime, timedelta
from pathlib import Path

from app.database import DatabaseType, get_database
from app.services.master_reports_detailed import MasterReportsService
from app.utils.helpers import MOSCOW_TZ, get_now


logger = logging.getLogger(__name__)


class RealtimeDailyTableService:
    """Сервис для управления ежедневными таблицами в реальном времени"""

    def __init__(self) -> None:
        self.db: DatabaseType = get_database()
        self.reports_service = MasterReportsService()
        self.current_table_path: str | None = None
        self.current_date: date | None = None

    async def init(self):
        """Инициализация сервиса"""
        # ORM база данных не требует init_db
        if hasattr(self.db, "init_db"):
            await self.db.init_db()

        # Проверяем, есть ли текущая таблица за сегодня
        today = get_now().date()
        await self._ensure_current_table_exists(today)

    async def _ensure_current_table_exists(self, date: date):
        """Убеждаемся, что текущая таблица за указанную дату существует"""
        today_str = date.strftime("%Y%m%d")
        table_filename = f"daily_table_{today_str}.xlsx"
        reports_dir = Path("reports")
        reports_dir.mkdir(exist_ok=True)
        table_path = reports_dir / table_filename

        if not os.path.exists(table_path):
            # Создаем новую таблицу
            logger.info(f"Создание новой ежедневной таблицы за {date.strftime('%d.%m.%Y')}")
            await self._create_daily_table(date)

        self.current_table_path = str(table_path)
        self.current_date = date

    async def _create_daily_table(self, date: date):
        """Создает новую ежедневную таблицу"""
        # Преобразуем date в datetime с timezone
        start_datetime = datetime.combine(date, datetime.min.time()).replace(tzinfo=MOSCOW_TZ)
        start_datetime + timedelta(days=1)

        # Создаем отчет
        report_path = await self.reports_service.generate_daily_master_report(start_datetime)

        # Переименовываем файл в стандартное имя
        today_str = date.strftime("%Y%m%d")
        new_filename = f"daily_table_{today_str}.xlsx"
        reports_dir = Path("reports")
        reports_dir.mkdir(exist_ok=True)
        new_path = reports_dir / new_filename

        if report_path and os.path.exists(report_path):
            os.rename(report_path, new_path)
            logger.info(f"Ежедневная таблица создана: {new_path}")
        else:
            logger.info(f"Создание пустой ежедневной таблицы: {new_path}")
            # Создаем пустую таблицу с базовой структурой
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Сводка"
            ws.cell(row=1, column=1, value="Ежедневный отчет")
            ws.cell(row=2, column=1, value=f"Дата: {date.strftime('%d.%m.%Y')}")
            ws.cell(row=3, column=1, value="Нет закрытых заказов за этот день")
            wb.save(new_path)
            logger.info(f"Пустая ежедневная таблица создана: {new_path}")

    async def update_table_on_order_completion(self, order_id: int):
        """Обновляет таблицу при закрытии заказа"""
        try:
            if not self.current_table_path or not self.current_date:
                logger.warning("Текущая таблица не инициализирована")
                return

            # Проверяем, что заказ был закрыт сегодня
            order = await self.db.get_order_by_id(order_id)
            if not order or order.status != "completed":
                return

            if order.updated_at:
                updated_at = order.updated_at
                if updated_at.tzinfo is None:
                    updated_at = updated_at.replace(tzinfo=MOSCOW_TZ)

                if updated_at.date() != self.current_date:
                    logger.info(f"Заказ {order_id} закрыт не сегодня, пропускаем обновление")
                    return

            # Обновляем таблицу
            logger.info(f"Обновление ежедневной таблицы после закрытия заказа {order_id}")
            await self._update_current_table()

        except Exception as e:
            logger.error(f"Ошибка при обновлении таблицы после закрытия заказа {order_id}: {e}")

    async def _update_current_table(self):
        """Обновляет текущую таблицу"""
        if not self.current_date:
            return

        path = self.current_table_path
        if not path:
            return

        # Создаем новую версию таблицы
        start_datetime = datetime.combine(self.current_date, datetime.min.time()).replace(
            tzinfo=MOSCOW_TZ
        )
        start_datetime + timedelta(days=1)

        # Генерируем обновленный отчет
        report_path = await self.reports_service.generate_daily_master_report(start_datetime)

        if report_path and os.path.exists(report_path):
            # Заменяем текущую таблицу
            if os.path.exists(path):
                os.remove(path)
            os.rename(report_path, path)
            logger.info(f"Ежедневная таблица обновлена: {path}")
        else:
            logger.error(f"Не удалось обновить таблицу: {report_path}")

    async def save_and_create_new_table(self):
        """Сохраняет текущую таблицу и создает новую пустую"""
        try:
            if (
                self.current_table_path
                and self.current_date
                and os.path.exists(self.current_table_path)
            ):
                # Переименовываем текущую таблицу с финальным именем
                today_str = self.current_date.strftime("%Y%m%d")
                final_filename = f"daily_table_final_{today_str}.xlsx"
                reports_dir = Path("reports")
                reports_dir.mkdir(exist_ok=True)
                final_path = reports_dir / final_filename

                os.rename(self.current_table_path, final_path)
                logger.info(f"Ежедневная таблица сохранена: {final_path}")

            # Создаем новую таблицу за завтра
            tomorrow = get_now().date() + timedelta(days=1)
            await self._create_daily_table(tomorrow)

            logger.info("Создана новая пустая ежедневная таблица")

        except Exception as e:
            logger.error(f"Ошибка при сохранении и создании новой таблицы: {e}")

    async def get_current_table_path(self) -> str | None:
        """Возвращает путь к текущей таблице"""
        return self.current_table_path


# Глобальный экземпляр сервиса
realtime_table_service = RealtimeDailyTableService()
