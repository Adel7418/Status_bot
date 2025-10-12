"""
Сервис генерации отчетов
"""

import logging
from datetime import UTC, datetime, timedelta
from io import BytesIO

from aiogram.types import BufferedInputFile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.config import OrderStatus
from app.database import Database
from app.utils import format_date, format_datetime


logger = logging.getLogger(__name__)


class ReportsService:
    """Сервис для генерации отчетов"""

    def __init__(self, db: Database):
        """
        Инициализация

        Args:
            db: Экземпляр базы данных
        """
        self.db = db

    async def generate_masters_report(self) -> str:
        """
        Генерация отчета по мастерам

        Returns:
            Текст отчета
        """
        masters = await self.db.get_all_masters(only_approved=True)

        text = "📊 <b>Отчет по мастерам</b>\n\n"

        if not masters:
            return text + "Мастеров в системе нет."

        text += f"<b>Всего мастеров:</b> {len(masters)}\n\n"

        for master in masters:
            orders = await self.db.get_orders_by_master(master.id, exclude_closed=False)
            total = len(orders)
            completed = len([o for o in orders if o.status == OrderStatus.CLOSED])
            active = len(
                [o for o in orders if o.status not in [OrderStatus.CLOSED, OrderStatus.REFUSED]]
            )

            status = "🟢" if master.is_active else "🔴"

            text += (
                f"{status} <b>{master.get_display_name()}</b>\n"
                f"   🔧 {master.specialization}\n"
                f"   📊 Заявок: {total} (завершено: {completed}, активных: {active})\n\n"
            )

        return text

    async def generate_statuses_report(self) -> str:
        """
        Генерация отчета по статусам заявок

        Returns:
            Текст отчета
        """
        stats = await self.db.get_statistics()
        orders_by_status = stats.get("orders_by_status", {})
        total = stats.get("total_orders", 0)

        text = (
            "📊 <b>Отчет по статусам заявок</b>\n\n"
            f"<b>Всего заявок:</b> {total}\n\n"
            "<b>По статусам:</b>\n"
        )

        for status in OrderStatus.all_statuses():
            count = orders_by_status.get(status, 0)
            emoji = OrderStatus.get_status_emoji(status)
            name = OrderStatus.get_status_name(status)
            percentage = (count / total * 100) if total > 0 else 0

            text += f"{emoji} {name}: {count} ({percentage:.1f}%)\n"

        return text

    async def generate_equipment_report(self) -> str:
        """
        Генерация отчета по типам техники

        Returns:
            Текст отчета
        """
        # Получаем все заявки
        orders = await self.db.get_all_orders()

        # Подсчитываем по типам
        by_equipment = {}
        for order in orders:
            by_equipment[order.equipment_type] = by_equipment.get(order.equipment_type, 0) + 1

        text = (
            "📊 <b>Отчет по типам техники</b>\n\n"
            f"<b>Всего заявок:</b> {len(orders)}\n\n"
            "<b>По типам техники:</b>\n"
        )

        # Сортируем по количеству
        sorted_equipment = sorted(by_equipment.items(), key=lambda x: x[1], reverse=True)

        for equipment, count in sorted_equipment:
            percentage = (count / len(orders) * 100) if len(orders) > 0 else 0
            text += f"🔧 {equipment}: {count} ({percentage:.1f}%)\n"

        return text

    async def generate_period_report(self, start_date: datetime, end_date: datetime) -> str:
        """
        Генерация отчета за период

        Args:
            start_date: Начало периода
            end_date: Конец периода

        Returns:
            Текст отчета
        """
        # Получаем все заявки
        all_orders = await self.db.get_all_orders()

        # Фильтруем по периоду
        orders = [o for o in all_orders if o.created_at and start_date <= o.created_at <= end_date]

        text = (
            f"📊 <b>Отчет за период</b>\n"
            f"📅 {format_date(start_date)} - {format_date(end_date)}\n\n"
            f"<b>Всего заявок:</b> {len(orders)}\n\n"
        )

        if not orders:
            return text + "Заявок за этот период нет."

        # Статистика по статусам
        by_status = {}
        for order in orders:
            by_status[order.status] = by_status.get(order.status, 0) + 1

        text += "<b>По статусам:</b>\n"
        for status in OrderStatus.all_statuses():
            if status in by_status:
                emoji = OrderStatus.get_status_emoji(status)
                name = OrderStatus.get_status_name(status)
                text += f"{emoji} {name}: {by_status[status]}\n"

        # Статистика по мастерам
        by_master = {}
        for order in orders:
            if order.assigned_master_id:
                by_master[order.master_name] = by_master.get(order.master_name, 0) + 1

        if by_master:
            text += "\n<b>По мастерам:</b>\n"
            for master, count in sorted(by_master.items(), key=lambda x: x[1], reverse=True):
                text += f"👨‍🔧 {master}: {count}\n"

        return text

    async def generate_excel_report(
        self,
        report_type: str = "all",
        start_date: datetime | None = None,
        end_date: datetime | None = None,
    ) -> BufferedInputFile:
        """
        Генерация отчета в формате Excel

        Args:
            report_type: Тип отчета (all, masters, statuses, equipment)
            start_date: Начало периода (для периода)
            end_date: Конец периода (для периода)

        Returns:
            Файл Excel
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет"

        # Стили
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        if report_type == "masters":
            await self._fill_masters_sheet(ws, header_font, header_fill, header_alignment)
        elif report_type == "statuses":
            await self._fill_statuses_sheet(ws, header_font, header_fill, header_alignment)
        elif report_type == "equipment":
            await self._fill_equipment_sheet(ws, header_font, header_fill, header_alignment)
        else:  # all
            await self._fill_all_orders_sheet(
                ws, header_font, header_fill, header_alignment, start_date, end_date
            )

        # Сохраняем в BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        filename = f"report_{report_type}_{datetime.now(UTC).strftime('%Y%m%d_%H%M%S')}.xlsx"

        return BufferedInputFile(excel_file.read(), filename=filename)

    async def _fill_masters_sheet(self, ws, header_font, header_fill, header_alignment):
        """Заполнение листа отчета по мастерам"""
        # Заголовки
        headers = [
            "ID",
            "Имя",
            "Телефон",
            "Специализация",
            "Статус",
            "Всего заявок",
            "Завершено",
            "Активных",
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Данные
        masters = await self.db.get_all_masters(only_approved=True)

        for master in masters:
            orders = await self.db.get_orders_by_master(master.id, exclude_closed=False)
            total = len(orders)
            completed = len([o for o in orders if o.status == OrderStatus.CLOSED])
            active = len(
                [o for o in orders if o.status not in [OrderStatus.CLOSED, OrderStatus.REFUSED]]
            )

            ws.append(
                [
                    master.id,
                    master.get_display_name(),
                    master.phone,
                    master.specialization,
                    "Активен" if master.is_active else "Неактивен",
                    total,
                    completed,
                    active,
                ]
            )

        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

    async def _fill_statuses_sheet(self, ws, header_font, header_fill, header_alignment):
        """Заполнение листа отчета по статусам"""
        headers = ["Статус", "Количество", "Процент"]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        stats = await self.db.get_statistics()
        orders_by_status = stats.get("orders_by_status", {})
        total = stats.get("total_orders", 0)

        for status in OrderStatus.all_statuses():
            count = orders_by_status.get(status, 0)
            percentage = (count / total * 100) if total > 0 else 0

            ws.append([OrderStatus.get_status_name(status), count, f"{percentage:.1f}%"])

        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

    async def _fill_equipment_sheet(self, ws, header_font, header_fill, header_alignment):
        """Заполнение листа отчета по типам техники"""
        headers = ["Тип техники", "Количество", "Процент"]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        orders = await self.db.get_all_orders()

        by_equipment = {}
        for order in orders:
            by_equipment[order.equipment_type] = by_equipment.get(order.equipment_type, 0) + 1

        total = len(orders)
        sorted_equipment = sorted(by_equipment.items(), key=lambda x: x[1], reverse=True)

        for equipment, count in sorted_equipment:
            percentage = (count / total * 100) if total > 0 else 0
            ws.append([equipment, count, f"{percentage:.1f}%"])

        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

    async def _fill_all_orders_sheet(
        self, ws, header_font, header_fill, header_alignment, start_date, end_date
    ):
        """Заполнение листа со всеми заявками"""
        headers = [
            "ID",
            "Дата создания",
            "Тип техники",
            "Описание",
            "Клиент",
            "Телефон",
            "Адрес",
            "Статус",
            "Мастер",
            "Диспетчер",
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        all_orders = await self.db.get_all_orders()

        # Фильтруем по периоду если указан
        if start_date and end_date:
            orders = [
                o for o in all_orders if o.created_at and start_date <= o.created_at <= end_date
            ]
        else:
            orders = all_orders

        for order in orders:
            ws.append(
                [
                    order.id,
                    format_datetime(order.created_at) if order.created_at else "",
                    order.equipment_type,
                    order.description,
                    order.client_name,
                    order.client_phone,
                    order.client_address,
                    OrderStatus.get_status_name(order.status),
                    order.master_name or "",
                    order.dispatcher_name or "",
                ]
            )

        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

    @staticmethod
    def get_period_dates(period: str) -> tuple[datetime, datetime]:
        """
        Получение дат для периода

        Args:
            period: Период (today, yesterday, week, month, year)

        Returns:
            Кортеж (start_date, end_date)
        """
        now = datetime.now()

        if period == "today":
            start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end = now.replace(hour=23, minute=59, second=59, microsecond=999999)
        elif period == "yesterday":
            yesterday = now - timedelta(days=1)
            start = yesterday.replace(hour=0, minute=0, second=0, microsecond=0)
            end = yesterday.replace(hour=23, minute=59, second=59, microsecond=999999)
        elif period == "week":
            start = now - timedelta(days=7)
            end = now
        elif period == "month":
            start = now - timedelta(days=30)
            end = now
        elif period == "year":
            start = now - timedelta(days=365)
            end = now
        else:
            start = now - timedelta(days=30)
            end = now

        return start, end
