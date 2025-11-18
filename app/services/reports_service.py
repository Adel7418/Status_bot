"""
Ð¡ÐµÑ€Ð²Ð¸Ñ Ð´Ð»Ñ Ð³ÐµÐ½ÐµÑ€Ð°Ñ†Ð¸Ð¸ Ð¸ Ð¾Ñ‚Ð¿Ñ€Ð°Ð²ÐºÐ¸ Ð¾Ñ‚Ñ‡ÐµÑ‚Ð¾Ð²
"""

import logging
from collections.abc import Mapping
from datetime import datetime, timedelta
from pathlib import Path
from typing import TYPE_CHECKING, Any

from app.database import DatabaseType, get_database
from app.repositories.order_repository_extended import OrderRepositoryExtended
from app.utils.helpers import get_now


if TYPE_CHECKING:
    from app.database.db import Database as LegacyDatabase


logger = logging.getLogger(__name__)


class ReportsService:
    """Ð¡ÐµÑ€Ð²Ð¸Ñ Ð´Ð»Ñ Ð³ÐµÐ½ÐµÑ€Ð°Ñ†Ð¸Ð¸ Ð¾Ñ‚Ñ‡ÐµÑ‚Ð¾Ð²"""

    def __init__(self) -> None:
        self.db: DatabaseType = get_database()
        self._order_repo_extended: OrderRepositoryExtended | None = None

    def _get_legacy_db(self) -> "LegacyDatabase":
        """
        ÐŸÐ¾Ð»ÑƒÑ‡Ð¸Ñ‚ÑŒ legacy-Ñ€ÐµÐ°Ð»Ð¸Ð·Ð°Ñ†Ð¸ÑŽ Ð‘Ð”.

        ÐžÑ‚Ñ‡ÐµÑ‚Ñ‹ Ð¸ÑÐ¿Ð¾Ð»ÑŒÐ·ÑƒÑŽÑ‚ Ð¿Ñ€ÑÐ¼Ð¾Ð¹ Ð´Ð¾ÑÑ‚ÑƒÐ¿ Ðº SQLite, Ð¿Ð¾ÑÑ‚Ð¾Ð¼Ñƒ
        Ð¿Ð¾Ð´Ð´ÐµÑ€Ð¶Ð¸Ð²Ð°ÐµÑ‚ÑÑ Ñ‚Ð¾Ð»ÑŒÐºÐ¾ legacy Database, Ð° Ð½Ðµ ORMDatabase.
        """
        from app.database.db import Database as LegacyDatabaseRuntime

        if not isinstance(self.db, LegacyDatabaseRuntime):
            raise RuntimeError("ReportsService Ð¿Ð¾Ð´Ð´ÐµÑ€Ð¶Ð¸Ð²Ð°ÐµÑ‚ Ñ‚Ð¾Ð»ÑŒÐºÐ¾ legacy Database (SQLite)")

        return self.db

    def _get_connection(self):
        """
        Ð¢Ð¸Ð¿Ð¾Ð±ÐµÐ·Ð¾Ð¿Ð°ÑÐ½Ð¾ Ð¿Ð¾Ð»ÑƒÑ‡Ð¸Ñ‚ÑŒ ÑÐ¾ÐµÐ´Ð¸Ð½ÐµÐ½Ð¸Ðµ Ñ Ð‘Ð”.

        ÐŸÑ€ÐµÐ´Ð¿Ð¾Ð»Ð°Ð³Ð°ÐµÑ‚ÑÑ, Ñ‡Ñ‚Ð¾ Ð´Ð¾ Ð²Ñ‹Ð·Ð¾Ð²Ð° ÑƒÐ¶Ðµ Ð²Ñ‹Ð¿Ð¾Ð»Ð½ÐµÐ½ self.db.connect().
        """
        legacy_db = self._get_legacy_db()
        return legacy_db.get_connection()

    async def _get_extended_repo(self) -> OrderRepositoryExtended:
        """ÐŸÐ¾Ð»ÑƒÑ‡Ð¸Ñ‚ÑŒ Ñ€Ð°ÑÑˆÐ¸Ñ€ÐµÐ½Ð½Ñ‹Ð¹ Ñ€ÐµÐ¿Ð¾Ð·Ð¸Ñ‚Ð¾Ñ€Ð¸Ð¹"""
        if self._order_repo_extended is None:
            connection = self._get_connection()
            self._order_repo_extended = OrderRepositoryExtended(connection)
        return self._order_repo_extended

    async def generate_daily_report(self) -> dict[str, Any]:
        """Ð“ÐµÐ½ÐµÑ€Ð¸Ñ€ÑƒÐµÑ‚ ÐµÐ¶ÐµÐ´Ð½ÐµÐ²Ð½Ñ‹Ð¹ Ð¾Ñ‚Ñ‡ÐµÑ‚"""
        await self.db.connect()

        try:
            today = get_now().date()
            yesterday = today - timedelta(days=1)

            # Ð¡Ñ‚Ð°Ñ‚Ð¸ÑÑ‚Ð¸ÐºÐ° Ð·Ð° Ð²Ñ‡ÐµÑ€Ð°
            orders_stats = await self._get_orders_stats(yesterday, today)
            masters_stats = await self._get_masters_stats(yesterday, today)
            closed_orders = await self._get_closed_orders_list(yesterday, today)

            return {
                "type": "daily",
                "period": f"{yesterday.strftime('%d.%m.%Y')}",
                "date_generated": get_now().isoformat(),
                "start_date": yesterday,
                "end_date": today,
                "orders": orders_stats,
                "masters": masters_stats,
                "summary": await self._get_summary_stats(yesterday, today),
                "closed_orders": closed_orders,
            }

        finally:
            await self.db.disconnect()

    async def generate_weekly_report(self) -> dict[str, Any]:
        """Ð“ÐµÐ½ÐµÑ€Ð¸Ñ€ÑƒÐµÑ‚ ÐµÐ¶ÐµÐ½ÐµÐ´ÐµÐ»ÑŒÐ½Ñ‹Ð¹ Ð¾Ñ‚Ñ‡ÐµÑ‚"""
        await self.db.connect()

        try:
            today = get_now().date()
            current_week_start = today - timedelta(
                days=today.weekday()
            )  # Ð¿Ð¾Ð½ÐµÐ´ÐµÐ»ÑŒÐ½Ð¸Ðº Ñ‚ÐµÐºÑƒÑ‰ÐµÐ¹ Ð½ÐµÐ´ÐµÐ»Ð¸
            prev_week_end = current_week_start - timedelta(days=1)  # Ð²Ð¾ÑÐºÑ€ÐµÑÐµÐ½ÑŒÐµ Ð¿Ñ€Ð¾ÑˆÐ»Ð¾Ð¹ Ð½ÐµÐ´ÐµÐ»Ð¸
            prev_week_start = prev_week_end - timedelta(days=6)  # Ð¿Ð¾Ð½ÐµÐ´ÐµÐ»ÑŒÐ½Ð¸Ðº Ð¿Ñ€Ð¾ÑˆÐ»Ð¾Ð¹ Ð½ÐµÐ´ÐµÐ»Ð¸

            orders_stats = await self._get_orders_stats(
                prev_week_start, prev_week_end + timedelta(days=1)
            )
            masters_stats = await self._get_masters_stats(
                prev_week_start, prev_week_end + timedelta(days=1)
            )
            closed_orders = await self._get_closed_orders_list(
                prev_week_start, prev_week_end + timedelta(days=1)
            )

            return {
                "type": "weekly",
                "period": f"{prev_week_start.strftime('%d.%m.%Y')} - {prev_week_end.strftime('%d.%m.%Y')}",
                "date_generated": get_now().isoformat(),
                "orders": orders_stats,
                "masters": masters_stats,
                "summary": await self._get_summary_stats(
                    prev_week_start, prev_week_end + timedelta(days=1)
                ),
                "closed_orders": closed_orders,
            }

        finally:
            await self.db.disconnect()

    async def generate_monthly_report(self) -> dict[str, Any]:
        """Ð“ÐµÐ½ÐµÑ€Ð¸Ñ€ÑƒÐµÑ‚ ÐµÐ¶ÐµÐ¼ÐµÑÑÑ‡Ð½Ñ‹Ð¹ Ð¾Ñ‚Ñ‡ÐµÑ‚"""
        await self.db.connect()

        try:
            today = get_now().date()
            month_start = today.replace(day=1)
            if today.month == 12:
                month_end = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                month_end = today.replace(month=today.month + 1, day=1) - timedelta(days=1)

            orders_stats = await self._get_orders_stats(month_start, month_end + timedelta(days=1))
            masters_stats = await self._get_masters_stats(
                month_start, month_end + timedelta(days=1)
            )
            closed_orders = await self._get_closed_orders_list(
                month_start, month_end + timedelta(days=1)
            )

            return {
                "type": "monthly",
                "period": f"{month_start.strftime('%d.%m.%Y')} - {month_end.strftime('%d.%m.%Y')}",
                "date_generated": get_now().isoformat(),
                "orders": orders_stats,
                "masters": masters_stats,
                "summary": await self._get_summary_stats(
                    month_start, month_end + timedelta(days=1)
                ),
                "closed_orders": closed_orders,
            }

        finally:
            await self.db.disconnect()

    async def _get_orders_stats(self, start_date, end_date) -> dict[str, Any]:
        """ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÑ‚ ÑÑ‚Ð°Ñ‚Ð¸ÑÑ‚Ð¸ÐºÑƒ Ð¿Ð¾ Ð·Ð°ÐºÐ°Ð·Ð°Ð¼ Ð·Ð° Ð¿ÐµÑ€Ð¸Ð¾Ð´"""
        connection = self._get_connection()
        cursor = await connection.execute(
            """
            SELECT
                COUNT(*) as total_orders,
                SUM(CASE WHEN status = 'NEW' THEN 1 ELSE 0 END) as new_orders,
                SUM(CASE WHEN status = 'ASSIGNED' THEN 1 ELSE 0 END) as assigned_orders,
                SUM(CASE WHEN status = 'ACCEPTED' THEN 1 ELSE 0 END) as accepted_orders,
                SUM(CASE WHEN status = 'IN_PROGRESS' THEN 1 ELSE 0 END) as in_progress_orders,
                SUM(CASE WHEN status = 'CLOSED' THEN 1 ELSE 0 END) as closed_orders,
                SUM(CASE WHEN status = 'CANCELLED' THEN 1 ELSE 0 END) as cancelled_orders,
                SUM(CASE WHEN out_of_city = 1 THEN 1 ELSE 0 END) as out_of_city_orders,
                SUM(CASE WHEN has_review = 1 THEN 1 ELSE 0 END) as review_orders,
                SUM(CASE WHEN status = 'CLOSED' THEN total_amount ELSE 0 END) as total_amount,
                SUM(CASE WHEN status = 'CLOSED' THEN materials_cost ELSE 0 END) as total_materials_cost,
                SUM(CASE WHEN status = 'CLOSED' THEN master_profit ELSE 0 END) as total_master_profit,
                SUM(CASE WHEN status = 'CLOSED' THEN company_profit ELSE 0 END) as total_company_profit
            FROM orders
            WHERE DATE(created_at) >= ? AND DATE(created_at) <= ?
        """,
            (start_date, end_date),
        )

        row_raw = await cursor.fetchone()
        row: Mapping[str, Any] = dict(row_raw) if row_raw is not None else {}

        return {
            "total_orders": row.get("total_orders", 0) or 0,
            "new_orders": row.get("new_orders", 0) or 0,
            "assigned_orders": row.get("assigned_orders", 0) or 0,
            "accepted_orders": row.get("accepted_orders", 0) or 0,
            "in_progress_orders": row.get("in_progress_orders", 0) or 0,
            "closed_orders": row.get("closed_orders", 0) or 0,
            "cancelled_orders": row.get("cancelled_orders", 0) or 0,
            "out_of_city_orders": row.get("out_of_city_orders", 0) or 0,
            "review_orders": row.get("review_orders", 0) or 0,
            "total_amount": float(row.get("total_amount", 0) or 0),
            "total_materials_cost": float(row.get("total_materials_cost", 0) or 0),
            "total_master_profit": float(row.get("total_master_profit", 0) or 0),
            "total_company_profit": float(row.get("total_company_profit", 0) or 0),
        }

    async def _get_masters_stats(self, start_date, end_date) -> list[dict[str, Any]]:
        """ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÑ‚ ÑÑ‚Ð°Ñ‚Ð¸ÑÑ‚Ð¸ÐºÑƒ Ð¿Ð¾ Ð¼Ð°ÑÑ‚ÐµÑ€Ð°Ð¼ Ð·Ð° Ð¿ÐµÑ€Ð¸Ð¾Ð´"""
        from datetime import datetime

        # ÐŸÑ€ÐµÐ¾Ð±Ñ€Ð°Ð·ÑƒÐµÐ¼ Ð´Ð°Ñ‚Ñ‹ Ð² datetime Ð´Ð»Ñ ÐºÐ¾Ñ€Ñ€ÐµÐºÑ‚Ð½Ð¾Ð³Ð¾ ÑÑ€Ð°Ð²Ð½ÐµÐ½Ð¸Ñ
        if isinstance(start_date, str):
            start_date = datetime.fromisoformat(start_date)
        if isinstance(end_date, str):
            end_date = datetime.fromisoformat(end_date)

        # ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÐ¼ Ð²ÑÐµÑ… Ð°ÐºÑ‚Ð¸Ð²Ð½Ñ‹Ñ… Ð¼Ð°ÑÑ‚ÐµÑ€Ð¾Ð² Ð¸ Ð·Ð°ÐºÐ°Ð·Ñ‹ Ð·Ð° Ð¿ÐµÑ€Ð¸Ð¾Ð´
        from app.config import OrderStatus

        masters = await self.db.get_all_masters(only_active=True, only_approved=True)
        orders_in_period = await self.db.get_orders_by_period(start_date, end_date)

        result: list[dict[str, Any]] = []
        for master in masters:
            # Ð¤Ð¸Ð»ÑŒÑ‚Ñ€ÑƒÐµÐ¼ Ð·Ð°ÑÐ²ÐºÐ¸ Ð¼Ð°ÑÑ‚ÐµÑ€Ð°
            orders = [o for o in orders_in_period if o.assigned_master_id == master.id]

            # ÐŸÐ¾Ð´ÑÑ‡Ð¸Ñ‚Ñ‹Ð²Ð°ÐµÐ¼ ÑÑ‚Ð°Ñ‚Ð¸ÑÑ‚Ð¸ÐºÑƒ
            orders_count = len(orders)
            closed_orders = len([o for o in orders if o.status == OrderStatus.CLOSED])
            refused_orders = len([o for o in orders if o.status == OrderStatus.REFUSED])
            out_of_city_count = sum(1 for o in orders if o.out_of_city is True)
            reviews_count = sum(1 for o in orders if o.has_review is True)
            total_profit = sum(
                o.master_profit or 0 for o in orders if o.status == OrderStatus.CLOSED
            )

            # Ð¡Ñ€ÐµÐ´Ð½Ð¸Ð¹ Ñ‡ÐµÐº Ð¿Ð¾ Ð·Ð°ÐºÑ€Ñ‹Ñ‚Ñ‹Ð¼ Ð·Ð°ÑÐ²ÐºÐ°Ð¼
            closed_amounts = [
                o.total_amount for o in orders if o.status == OrderStatus.CLOSED and o.total_amount
            ]
            avg_order_amount = sum(closed_amounts) / len(closed_amounts) if closed_amounts else 0

            result.append(
                {
                    "id": master.id,
                    "name": master.get_display_name(),
                    "orders_count": orders_count,
                    "closed_orders": closed_orders,
                    "refused_orders": refused_orders,
                    "out_of_city_count": out_of_city_count,
                    "reviews_count": reviews_count,
                    "total_profit": float(total_profit),
                    "avg_order_amount": float(avg_order_amount),
                }
            )

        # Ð¡Ð¾Ñ€Ñ‚Ð¸Ñ€ÑƒÐµÐ¼ Ð¿Ð¾ Ð¿Ñ€Ð¸Ð±Ñ‹Ð»Ð¸
        result.sort(key=lambda x: float(x.get("total_profit") or 0.0), reverse=True)

        return result

    async def _get_accepted_orders_details(self, start_date, end_date) -> list[dict[str, Any]]:
        """ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÑ‚ Ð´ÐµÑ‚Ð°Ð»ÑŒÐ½ÑƒÑŽ Ð¸Ð½Ñ„Ð¾Ñ€Ð¼Ð°Ñ†Ð¸ÑŽ Ð¾ Ð¿Ñ€Ð¸Ð½ÑÑ‚Ñ‹Ñ… Ð·Ð°ÐºÐ°Ð·Ð°Ñ… Ð·Ð° Ð¿ÐµÑ€Ð¸Ð¾Ð´"""
        connection = self._get_connection()
        cursor = await connection.execute(
            """
            SELECT
                o.id,
                o.equipment_type,
                o.client_name,
                o.client_address,
                o.client_phone,
                o.created_at,
                o.scheduled_time,
                o.notes,
                u.first_name || ' ' || COALESCE(u.last_name, '') as master_name
            FROM orders o
            LEFT JOIN masters m ON o.assigned_master_id = m.id
            LEFT JOIN users u ON m.telegram_id = u.telegram_id
            WHERE o.status = 'ACCEPTED'
                AND DATE(o.created_at) >= ? AND DATE(o.created_at) <= ?
            ORDER BY o.created_at DESC
        """,
            (start_date, end_date),
        )

        rows = await cursor.fetchall()

        orders_list = []
        for row in rows:
            orders_list.append(
                {
                    "id": row["id"],
                    "equipment_type": row["equipment_type"],
                    "client_name": row["client_name"],
                    "client_address": row["client_address"],
                    "client_phone": row["client_phone"],
                    "created_at": row["created_at"],
                    "scheduled_time": row["scheduled_time"],
                    "notes": row["notes"],
                    "master_name": row["master_name"] or "ÐÐµ Ð½Ð°Ð·Ð½Ð°Ñ‡ÐµÐ½",
                }
            )

        return orders_list

    async def _get_summary_stats(self, start_date, end_date) -> dict[str, Any]:
        """ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÑ‚ Ð¾Ð±Ñ‰ÑƒÑŽ ÑÑ‚Ð°Ñ‚Ð¸ÑÑ‚Ð¸ÐºÑƒ Ð·Ð° Ð¿ÐµÑ€Ð¸Ð¾Ð´"""
        connection = self._get_connection()
        cursor = await connection.execute(
            """
            SELECT
                COUNT(DISTINCT assigned_master_id) as active_masters,
                AVG(CASE WHEN status = 'CLOSED' THEN total_amount ELSE NULL END) as avg_order_amount,
                MAX(CASE WHEN status = 'CLOSED' THEN total_amount ELSE NULL END) as max_order_amount,
                MIN(CASE WHEN status = 'CLOSED' THEN total_amount ELSE NULL END) as min_order_amount
            FROM orders
            WHERE DATE(created_at) >= ? AND DATE(created_at) <= ?
        """,
            (start_date, end_date),
        )

        row_raw = await cursor.fetchone()
        row: Mapping[str, Any] = dict(row_raw) if row_raw is not None else {}

        return {
            "active_masters": row.get("active_masters", 0) or 0,
            "avg_order_amount": float(row.get("avg_order_amount", 0) or 0),
            "max_order_amount": float(row.get("max_order_amount", 0) or 0),
            "min_order_amount": float(row.get("min_order_amount", 0) or 0),
        }

    async def _get_closed_orders_list(self, start_date, end_date) -> list[dict[str, Any]]:
        """ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÑ‚ ÑÐ¿Ð¸ÑÐ¾Ðº Ð·Ð°ÐºÑ€Ñ‹Ñ‚Ñ‹Ñ… Ð·Ð°ÐºÐ°Ð·Ð¾Ð² Ð·Ð° Ð¿ÐµÑ€Ð¸Ð¾Ð´ Ñ Ð¸ÑÑ‚Ð¾Ñ€Ð¸ÐµÐ¹"""
        connection = self._get_connection()
        cursor = await connection.execute(
            """
            SELECT
                o.id,
                o.equipment_type,
                o.client_name,
                o.client_address,
                o.total_amount,
                o.materials_cost,
                o.master_profit,
                o.company_profit,
                o.out_of_city,
                o.has_review,
                o.created_at,
                o.updated_at,
                u.first_name || ' ' || COALESCE(u.last_name, '') as master_name
            FROM orders o
            LEFT JOIN masters m ON o.assigned_master_id = m.id
            LEFT JOIN users u ON m.telegram_id = u.telegram_id
            WHERE o.status = 'CLOSED'
                AND DATE(o.updated_at) >= ?
                AND DATE(o.updated_at) <= ?
            ORDER BY o.updated_at DESC
        """,
            (start_date, end_date),
        )

        rows = await cursor.fetchall()

        # ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÐ¼ Ñ€Ð°ÑÑˆÐ¸Ñ€ÐµÐ½Ð½Ñ‹Ð¹ Ñ€ÐµÐ¿Ð¾Ð·Ð¸Ñ‚Ð¾Ñ€Ð¸Ð¹
        order_repo = await self._get_extended_repo()

        orders_list = []
        for row in rows:
            order_id = row["id"]

            # ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÐ¼ Ð¸ÑÑ‚Ð¾Ñ€Ð¸ÑŽ ÑÑ‚Ð°Ñ‚ÑƒÑÐ¾Ð² Ð´Ð»Ñ ÐºÐ°Ð¶Ð´Ð¾Ð¹ Ð·Ð°ÑÐ²ÐºÐ¸
            status_history = await order_repo.get_status_history(order_id)

            # Ð¡Ñ‡Ð¸Ñ‚Ð°ÐµÐ¼ ÑÑ‚Ð°Ñ‚Ð¸ÑÑ‚Ð¸ÐºÑƒ Ð¿Ð¾ Ð¸ÑÑ‚Ð¾Ñ€Ð¸Ð¸
            history_stats = {
                "total_changes": len(status_history),
                "days_to_complete": 0,
                "status_changes": [],
            }

            if status_history:
                # Ð‘ÐµÑ€ÐµÐ¼ Ð¿ÐµÑ€Ð²Ñ‹Ðµ 3 Ð¸Ð·Ð¼ÐµÐ½ÐµÐ½Ð¸Ñ Ð´Ð»Ñ Ð¾Ñ‚Ñ‡ÐµÑ‚Ð°
                history_stats["status_changes"] = [
                    {
                        "from": h["old_status"],
                        "to": h["new_status"],
                        "date": h["changed_at"],
                        "changed_by": h.get("username", "Ð¡Ð¸ÑÑ‚ÐµÐ¼Ð°"),
                    }
                    for h in status_history[:3]
                ]

                # Ð¡Ñ‡Ð¸Ñ‚Ð°ÐµÐ¼ Ð´Ð½Ð¸ Ð¾Ñ‚ ÑÐ¾Ð·Ð´Ð°Ð½Ð¸Ñ Ð´Ð¾ Ð·Ð°ÐºÑ€Ñ‹Ñ‚Ð¸Ñ
                from datetime import datetime

                if row["created_at"] and row["updated_at"]:
                    created = datetime.fromisoformat(row["created_at"])
                    updated = datetime.fromisoformat(row["updated_at"])
                    # Ð£Ð´Ð°Ð»ÑÐµÐ¼ timezone Ð¸Ð½Ñ„Ð¾Ñ€Ð¼Ð°Ñ†Ð¸ÑŽ Ð´Ð»Ñ ÐºÐ¾Ñ€Ñ€ÐµÐºÑ‚Ð½Ð¾Ð³Ð¾ Ð²Ñ‹Ñ‡Ð¸Ñ‚Ð°Ð½Ð¸Ñ
                    if created.tzinfo is not None:
                        created = created.replace(tzinfo=None)
                    if updated.tzinfo is not None:
                        updated = updated.replace(tzinfo=None)
                    history_stats["days_to_complete"] = (updated - created).days

            orders_list.append(
                {
                    "id": order_id,
                    "equipment_type": row["equipment_type"],
                    "client_name": row["client_name"],
                    "client_address": row["client_address"],
                    "master_name": row["master_name"] or "ÐÐµ Ð½Ð°Ð·Ð½Ð°Ñ‡ÐµÐ½",
                    "total_amount": float(row["total_amount"] or 0),
                    "materials_cost": float(row["materials_cost"] or 0),
                    "master_profit": float(row["master_profit"] or 0),
                    "company_profit": float(row["company_profit"] or 0),
                    "out_of_city": bool(row["out_of_city"]),
                    "has_review": bool(row["has_review"]),
                    "created_at": row["created_at"],
                    "closed_at": row["updated_at"],
                    "history": history_stats,  # âœ¨ ÐÐžÐ’ÐžÐ•: Ð˜ÑÑ‚Ð¾Ñ€Ð¸Ñ Ð¸Ð·Ð¼ÐµÐ½ÐµÐ½Ð¸Ð¹
                }
            )

        return orders_list

    async def _get_order_history_summary(self, order_id: int) -> str:
        """
        ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÑ‚ ÐºÑ€Ð°Ñ‚ÐºÑƒÑŽ Ð¸ÑÑ‚Ð¾Ñ€Ð¸ÑŽ Ð·Ð°ÑÐ²ÐºÐ¸ Ð´Ð»Ñ Ð²ÐºÐ»ÑŽÑ‡ÐµÐ½Ð¸Ñ Ð² Ð¾Ñ‚Ñ‡ÐµÑ‚

        Args:
            order_id: ID Ð·Ð°ÑÐ²ÐºÐ¸

        Returns:
            Ð¢ÐµÐºÑÑ‚Ð¾Ð²Ð¾Ðµ Ð¾Ð¿Ð¸ÑÐ°Ð½Ð¸Ðµ Ð¸ÑÑ‚Ð¾Ñ€Ð¸Ð¸
        """
        order_repo = await self._get_extended_repo()

        try:
            status_history = await order_repo.get_status_history(order_id)

            if not status_history:
                return "Ð˜ÑÑ‚Ð¾Ñ€Ð¸Ñ Ð¸Ð·Ð¼ÐµÐ½ÐµÐ½Ð¸Ð¹ Ð¾Ñ‚ÑÑƒÑ‚ÑÑ‚Ð²ÑƒÐµÑ‚"

            # Ð¤Ð¾Ñ€Ð¼Ð¸Ñ€ÑƒÐµÐ¼ ÐºÑ€Ð°Ñ‚ÐºÑƒÑŽ Ð¸ÑÑ‚Ð¾Ñ€Ð¸ÑŽ
            summary = f"Ð˜Ð·Ð¼ÐµÐ½ÐµÐ½Ð¸Ð¹: {len(status_history)}\n"

            for i, h in enumerate(status_history[:3], 1):  # ÐŸÐµÑ€Ð²Ñ‹Ðµ 3 Ð¸Ð·Ð¼ÐµÐ½ÐµÐ½Ð¸Ñ
                summary += f"  {i}. {h['changed_at'][:16]}: {h['old_status']} â†’ {h['new_status']}\n"

            if len(status_history) > 3:
                summary += f"  ... Ð¸ ÐµÑ‰Ðµ {len(status_history) - 3} Ð¸Ð·Ð¼ÐµÐ½ÐµÐ½Ð¸Ð¹\n"

            return summary
        except Exception as e:
            logger.error(f"ÐžÑˆÐ¸Ð±ÐºÐ° Ð¿Ð¾Ð»ÑƒÑ‡ÐµÐ½Ð¸Ñ Ð¸ÑÑ‚Ð¾Ñ€Ð¸Ð¸ Ð´Ð»Ñ Ð·Ð°ÑÐ²ÐºÐ¸ #{order_id}: {e}")
            return "ÐžÑˆÐ¸Ð±ÐºÐ° Ð¿Ð¾Ð»ÑƒÑ‡ÐµÐ½Ð¸Ñ Ð¸ÑÑ‚Ð¾Ñ€Ð¸Ð¸"

    async def format_report_to_text(self, report: dict[str, Any]) -> str:
        """Ð¤Ð¾Ñ€Ð¼Ð°Ñ‚Ð¸Ñ€ÑƒÐµÑ‚ Ð¾Ñ‚Ñ‡ÐµÑ‚ Ð² Ñ‚ÐµÐºÑÑ‚Ð¾Ð²Ñ‹Ð¹ Ð²Ð¸Ð´"""
        report_type = report["type"]
        period = report["period"]
        orders = report["orders"]
        masters = report["masters"]
        summary = report["summary"]

        if report_type == "daily":
            title = "ðŸ“Š Ð•Ð–Ð•Ð”ÐÐ•Ð’ÐÐ«Ð™ ÐžÐ¢Ð§Ð•Ð¢"
            icon = "ðŸ“…"
        elif report_type == "weekly":
            title = "ðŸ“Š Ð•Ð–Ð•ÐÐ•Ð”Ð•Ð›Ð¬ÐÐ«Ð™ ÐžÐ¢Ð§Ð•Ð¢"
            icon = "ðŸ“†"
        else:
            title = "ðŸ“Š Ð•Ð–Ð•ÐœÐ•Ð¡Ð¯Ð§ÐÐ«Ð™ ÐžÐ¢Ð§Ð•Ð¢"
            icon = "ðŸ—“ï¸"

        text = f"{title}\n"
        text += f"{icon} ÐŸÐµÑ€Ð¸Ð¾Ð´: {period}\n"
        text += f"ðŸ“… Ð¡Ð³ÐµÐ½ÐµÑ€Ð¸Ñ€Ð¾Ð²Ð°Ð½: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n\n"

        # Ð¡Ñ‚Ð°Ñ‚Ð¸ÑÑ‚Ð¸ÐºÐ° Ð¿Ð¾ Ð·Ð°ÐºÐ°Ð·Ð°Ð¼
        text += "ðŸ“‹ Ð¡Ð¢ÐÐ¢Ð˜Ð¡Ð¢Ð˜ÐšÐ ÐŸÐž Ð—ÐÐšÐÐ—ÐÐœ:\n"
        text += f"â€¢ Ð’ÑÐµÐ³Ð¾ Ð·Ð°ÐºÐ°Ð·Ð¾Ð²: {orders['total_orders']}\n"
        text += f"â€¢ ÐÐ¾Ð²Ñ‹Ñ…: {orders['new_orders']}\n"
        text += f"â€¢ ÐÐ°Ð·Ð½Ð°Ñ‡ÐµÐ½Ð½Ñ‹Ñ…: {orders['assigned_orders']}\n"
        text += f"â€¢ ÐŸÑ€Ð¸Ð½ÑÑ‚Ñ‹Ñ…: {orders['accepted_orders']}\n"
        text += f"â€¢ Ð’ Ñ€Ð°Ð±Ð¾Ñ‚Ðµ: {orders['in_progress_orders']}\n"
        text += f"â€¢ Ð—Ð°Ð²ÐµÑ€ÑˆÐµÐ½Ð½Ñ‹Ñ…: {orders['closed_orders']}\n"
        text += f"â€¢ ÐžÑ‚Ð¼ÐµÐ½ÐµÐ½Ð½Ñ‹Ñ…: {orders['cancelled_orders']}\n"
        text += f"â€¢ Ð¡ Ð²Ñ‹ÐµÐ·Ð´Ð¾Ð¼ Ð·Ð° Ð³Ð¾Ñ€Ð¾Ð´: {orders['out_of_city_orders']}\n"
        text += f"â€¢ Ð¡ Ð¾Ñ‚Ð·Ñ‹Ð²Ð°Ð¼Ð¸: {orders['review_orders']}\n\n"

        # Ð¡Ñ‚Ð°Ñ‚Ð¸ÑÑ‚Ð¸ÐºÐ° Ð¿Ð¾ Ñ‚Ð¸Ð¿Ð°Ð¼ Ñ‚ÐµÑ…Ð½Ð¸ÐºÐ¸
        await self.db.connect()
        try:
            # ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÐ¼ Ð·Ð°ÐºÐ°Ð·Ñ‹ Ð·Ð° Ð¿ÐµÑ€Ð¸Ð¾Ð´
            start_date = report.get("start_date")
            end_date = report.get("end_date")
            if start_date and end_date:
                connection = self._get_connection()
                cursor = await connection.execute(
                    """
                    SELECT equipment_type, COUNT(*) as count
                    FROM orders
                    WHERE DATE(created_at) >= ? AND DATE(created_at) <= ?
                    GROUP BY equipment_type
                    ORDER BY count DESC
                    """,
                    (start_date, end_date),
                )
                rows = await cursor.fetchall()
                if rows:
                    text += "ðŸ”§ <b>ÐŸÐ¾ Ñ‚Ð¸Ð¿Ð°Ð¼ Ñ‚ÐµÑ…Ð½Ð¸ÐºÐ¸:</b>\n"
                    total_orders_for_percent = sum(row["count"] for row in rows)
                    for row in rows:
                        equipment_type = row["equipment_type"] or "ÐÐµ ÑƒÐºÐ°Ð·Ð°Ð½Ð¾"
                        count = row["count"]
                        percentage = (
                            (count / total_orders_for_percent * 100)
                            if total_orders_for_percent > 0
                            else 0
                        )
                        text += f"â€¢ {equipment_type}: {count} ({percentage:.1f}%)\n"
                    text += "\n"
        finally:
            await self.db.disconnect()

        # Ð”ÐµÑ‚Ð°Ð»ÑŒÐ½Ð°Ñ Ð¸Ð½Ñ„Ð¾Ñ€Ð¼Ð°Ñ†Ð¸Ñ Ð¾ Ð¿Ñ€Ð¸Ð½ÑÑ‚Ñ‹Ñ… Ð·Ð°ÐºÐ°Ð·Ð°Ñ…
        if orders["accepted_orders"] > 0:
            text += "âœ… ÐŸÐ Ð˜ÐÐ¯Ð¢Ð«Ð• Ð—ÐÐšÐÐ—Ð« (Ð´ÐµÑ‚Ð°Ð»Ð¸ Ð² Excel):\n"
            # ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÐ¼ Ð´ÐµÑ‚Ð°Ð»ÑŒÐ½ÑƒÑŽ Ð¸Ð½Ñ„Ð¾Ñ€Ð¼Ð°Ñ†Ð¸ÑŽ Ð¾ Ð¿Ñ€Ð¸Ð½ÑÑ‚Ñ‹Ñ… Ð·Ð°ÐºÐ°Ð·Ð°Ñ…
            accepted_orders_details = await self._get_accepted_orders_details(
                report.get("start_date", ""), report.get("end_date", "")
            )
            for order in accepted_orders_details[:10]:  # ÐŸÐ¾ÐºÐ°Ð·Ñ‹Ð²Ð°ÐµÐ¼ Ð¿ÐµÑ€Ð²Ñ‹Ðµ 10
                text += f"â€¢ #{order['id']} - {order['equipment_type']} | {order['client_name']} | {order['master_name']}\n"
            if len(accepted_orders_details) > 10:
                text += f"... Ð¸ ÐµÑ‰Ðµ {len(accepted_orders_details) - 10} Ð·Ð°ÐºÐ°Ð·Ð¾Ð²\n"
            text += "\n"

        # Ð¤Ð¸Ð½Ð°Ð½ÑÐ¾Ð²Ð°Ñ Ð¸Ð½Ñ„Ð¾Ñ€Ð¼Ð°Ñ†Ð¸Ñ
        if orders["closed_orders"] > 0:
            text += "ðŸ’° Ð¤Ð˜ÐÐÐÐ¡ÐžÐ’ÐÐ¯ Ð˜ÐÐ¤ÐžÐ ÐœÐÐ¦Ð˜Ð¯:\n"
            text += f"â€¢ ÐžÐ±Ñ‰Ð°Ñ ÑÑƒÐ¼Ð¼Ð°: {orders['total_amount']:.2f} â‚½\n"
            text += f"â€¢ Ð Ð°ÑÑ…Ð¾Ð´Ñ‹ Ð½Ð° Ð¼Ð°Ñ‚ÐµÑ€Ð¸Ð°Ð»Ñ‹: {orders['total_materials_cost']:.2f} â‚½\n"
            text += f"â€¢ ÐŸÑ€Ð¸Ð±Ñ‹Ð»ÑŒ Ð¼Ð°ÑÑ‚ÐµÑ€Ð¾Ð²: {orders['total_master_profit']:.2f} â‚½\n"
            text += f"â€¢ ÐŸÑ€Ð¸Ð±Ñ‹Ð»ÑŒ ÐºÐ¾Ð¼Ð¿Ð°Ð½Ð¸Ð¸: {orders['total_company_profit']:.2f} â‚½\n"
            text += f"â€¢ Ð¡Ñ€ÐµÐ´Ð½Ð¸Ð¹ Ñ‡ÐµÐº: {summary['avg_order_amount']:.2f} â‚½\n"
            text += f"â€¢ ÐœÐ°ÐºÑÐ¸Ð¼Ð°Ð»ÑŒÐ½Ñ‹Ð¹ Ñ‡ÐµÐº: {summary['max_order_amount']:.2f} â‚½\n\n"

        # Ð¡Ñ‚Ð°Ñ‚Ð¸ÑÑ‚Ð¸ÐºÐ° Ð¿Ð¾ Ð¼Ð°ÑÑ‚ÐµÑ€Ð°Ð¼
        if masters:
            text += "ðŸ‘¨â€ðŸ”§ Ð¢ÐžÐŸ ÐœÐÐ¡Ð¢Ð•Ð ÐžÐ’:\n"
            for i, master in enumerate(masters[:5], 1):  # Ð¢Ð¾Ð¿ 5
                text += f"{i}. {master['name']}\n"
                text += (
                    f"   Ð—Ð°ÐºÐ°Ð·Ð¾Ð²: {master['orders_count']} | Ð—Ð°Ð²ÐµÑ€ÑˆÐµÐ½Ð¾: {master['closed_orders']}\n"
                )
                text += f"   Ð’Ñ‹ÐµÐ·Ð´Ð¾Ð²: {master['out_of_city_count']} | ÐžÑ‚Ð·Ñ‹Ð²Ð¾Ð²: {master['reviews_count']}\n"
                text += f"   ÐŸÑ€Ð¸Ð±Ñ‹Ð»ÑŒ: {master['total_profit']:.2f} â‚½\n\n"

        text += f"ðŸ‘¥ ÐÐºÑ‚Ð¸Ð²Ð½Ñ‹Ñ… Ð¼Ð°ÑÑ‚ÐµÑ€Ð¾Ð²: {summary['active_masters']}\n\n"

        # Ð˜Ð½Ñ„Ð¾Ñ€Ð¼Ð°Ñ†Ð¸Ñ Ð¾ Ð´ÐµÑ‚Ð°Ð»Ð¸Ð·Ð°Ñ†Ð¸Ð¸
        closed_orders = report.get("closed_orders", [])
        if closed_orders:
            text += f"ðŸ“‹ Ð”ÐµÑ‚Ð°Ð»ÑŒÐ½Ð°Ñ Ð¸Ð½Ñ„Ð¾Ñ€Ð¼Ð°Ñ†Ð¸Ñ Ð¿Ð¾ {len(closed_orders)} Ð·Ð°ÐºÑ€Ñ‹Ñ‚Ñ‹Ð¼ Ð·Ð°ÐºÐ°Ð·Ð°Ð¼ Ð´Ð¾ÑÑ‚ÑƒÐ¿Ð½Ð° Ð² Excel Ñ„Ð°Ð¹Ð»Ðµ.\n\n"

            # âœ¨ ÐÐžÐ’ÐžÐ•: Ð”Ð¾Ð±Ð°Ð²Ð»ÑÐµÐ¼ ÑÑ‚Ð°Ñ‚Ð¸ÑÑ‚Ð¸ÐºÑƒ Ð¿Ð¾ Ð¸ÑÑ‚Ð¾Ñ€Ð¸Ð¸ Ð¸Ð·Ð¼ÐµÐ½ÐµÐ½Ð¸Ð¹
            total_changes = sum(
                order.get("history", {}).get("total_changes", 0) for order in closed_orders
            )
            avg_days = (
                sum(order.get("history", {}).get("days_to_complete", 0) for order in closed_orders)
                / len(closed_orders)
                if closed_orders
                else 0
            )

            text += "ðŸ“Š Ð˜Ð¡Ð¢ÐžÐ Ð˜Ð¯ Ð˜Ð—ÐœÐ•ÐÐ•ÐÐ˜Ð™:\n"
            text += f"â€¢ Ð’ÑÐµÐ³Ð¾ Ð¸Ð·Ð¼ÐµÐ½ÐµÐ½Ð¸Ð¹ ÑÑ‚Ð°Ñ‚ÑƒÑÐ¾Ð²: {total_changes}\n"
            text += f"â€¢ Ð¡Ñ€ÐµÐ´Ð½ÐµÐµ Ð²Ñ€ÐµÐ¼Ñ Ð²Ñ‹Ð¿Ð¾Ð»Ð½ÐµÐ½Ð¸Ñ: {avg_days:.1f} Ð´Ð½ÐµÐ¹\n\n"

            # ÐŸÐ¾ÐºÐ°Ð·Ñ‹Ð²Ð°ÐµÐ¼ Ð¿Ñ€Ð¸Ð¼ÐµÑ€Ñ‹ Ð±Ñ‹ÑÑ‚Ñ€Ñ‹Ñ… Ð¸ Ð¼ÐµÐ´Ð»ÐµÐ½Ð½Ñ‹Ñ… Ð·Ð°ÑÐ²Ð¾Ðº
            orders_with_days = [
                (o["id"], o.get("history", {}).get("days_to_complete", 0)) for o in closed_orders
            ]
            orders_with_days.sort(key=lambda x: x[1])

            if orders_with_days:
                fastest = orders_with_days[0]
                slowest = orders_with_days[-1]
                text += f"â€¢ Ð¡Ð°Ð¼Ð°Ñ Ð±Ñ‹ÑÑ‚Ñ€Ð°Ñ: #{fastest[0]} ({fastest[1]} Ð´Ð½.)\n"
                text += f"â€¢ Ð¡Ð°Ð¼Ð°Ñ Ð´Ð¾Ð»Ð³Ð°Ñ: #{slowest[0]} ({slowest[1]} Ð´Ð½.)\n"

        return text

    async def save_report_to_file(self, report: dict[str, Any], filename: str | None = None) -> str:
        """Ð¡Ð¾Ñ…Ñ€Ð°Ð½ÑÐµÑ‚ Ð¾Ñ‚Ñ‡ÐµÑ‚ Ð² Ñ‚ÐµÐºÑÑ‚Ð¾Ð²Ñ‹Ð¹ Ñ„Ð°Ð¹Ð»"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"report_{report['type']}_{timestamp}.txt"

        reports_dir = Path("reports")
        reports_dir.mkdir(exist_ok=True)

        file_path = reports_dir / filename

        text = await self.format_report_to_text(report)
        with open(file_path, "w", encoding="utf-8") as f:  # noqa: ASYNC101
            f.write(text)

        logger.info(f"ÐžÑ‚Ñ‡ÐµÑ‚ ÑÐ¾Ñ…Ñ€Ð°Ð½ÐµÐ½ Ð² Ñ„Ð°Ð¹Ð»: {file_path}")
        return str(file_path)

    async def save_report_to_excel(
        self, report: dict[str, Any], filename: str | None = None
    ) -> str:
        """Ð¡Ð¾Ñ…Ñ€Ð°Ð½ÑÐµÑ‚ Ð¾Ñ‚Ñ‡ÐµÑ‚ Ð² Excel Ñ„Ð°Ð¹Ð» Ñ Ð´ÐµÑ‚Ð°Ð»Ð¸Ð·Ð°Ñ†Ð¸ÐµÐ¹"""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        if not filename:
            # Ð¤Ð¸ÐºÑÐ¸Ñ€Ð¾Ð²Ð°Ð½Ð½Ð¾Ðµ Ð¸Ð¼Ñ Ñ„Ð°Ð¹Ð»Ð° (Ð¾Ð±Ð½Ð¾Ð²Ð»ÑÐµÑ‚ÑÑ ÐºÐ°Ð¶Ð´Ñ‹Ð¹ Ñ€Ð°Ð·)
            filename = f"report_{report['type']}.xlsx"

        reports_dir = Path("reports")
        reports_dir.mkdir(exist_ok=True)

        file_path = reports_dir / filename

        wb = Workbook()

        # Ð¡Ñ‚Ð¸Ð»Ð¸
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        subheader_font = Font(bold=True, size=12)
        subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        table_header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Ð›Ð¸ÑÑ‚ 1: ÐžÐ±Ñ‰Ð°Ñ ÑÑ‚Ð°Ñ‚Ð¸ÑÑ‚Ð¸ÐºÐ°
        ws1 = wb.active
        ws1.title = "Ð¡Ñ‚Ð°Ñ‚Ð¸ÑÑ‚Ð¸ÐºÐ°"

        row = 1
        ws1.merge_cells(f"A{row}:F{row}")
        cell = ws1[f"A{row}"]
        cell.value = f"{report['type'].upper()} ÐžÐ¢Ð§Ð•Ð¢"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        ws1.row_dimensions[row].height = 25

        row += 1
        ws1.merge_cells(f"A{row}:F{row}")
        cell = ws1[f"A{row}"]
        cell.value = f"ÐŸÐµÑ€Ð¸Ð¾Ð´: {report['period']}"
        cell.font = Font(bold=True, size=11)
        cell.alignment = center_alignment

        row += 2

        # Ð¡Ñ‚Ð°Ñ‚Ð¸ÑÑ‚Ð¸ÐºÐ° Ð¿Ð¾ Ð·Ð°ÐºÐ°Ð·Ð°Ð¼
        orders = report["orders"]
        ws1[f"A{row}"] = "Ð¡Ð¢ÐÐ¢Ð˜Ð¡Ð¢Ð˜ÐšÐ ÐŸÐž Ð—ÐÐšÐÐ—ÐÐœ"
        ws1[f"A{row}"].font = subheader_font
        ws1[f"A{row}"].fill = subheader_fill
        ws1.merge_cells(f"A{row}:F{row}")

        row += 1
        stats_data = [
            ["ÐŸÐ¾ÐºÐ°Ð·Ð°Ñ‚ÐµÐ»ÑŒ", "Ð—Ð½Ð°Ñ‡ÐµÐ½Ð¸Ðµ"],
            ["Ð’ÑÐµÐ³Ð¾ Ð·Ð°ÐºÐ°Ð·Ð¾Ð²", orders["total_orders"]],
            ["ÐÐ¾Ð²Ñ‹Ñ…", orders["new_orders"]],
            ["ÐÐ°Ð·Ð½Ð°Ñ‡ÐµÐ½Ð½Ñ‹Ñ…", orders["assigned_orders"]],
            ["ÐŸÑ€Ð¸Ð½ÑÑ‚Ñ‹Ñ…", orders["accepted_orders"]],
            ["Ð’ Ñ€Ð°Ð±Ð¾Ñ‚Ðµ", orders["in_progress_orders"]],
            ["Ð—Ð°Ð²ÐµÑ€ÑˆÐµÐ½Ð½Ñ‹Ñ…", orders["closed_orders"]],
            ["ÐžÑ‚Ð¼ÐµÐ½ÐµÐ½Ð½Ñ‹Ñ…", orders["cancelled_orders"]],
            ["Ð¡ Ð²Ñ‹ÐµÐ·Ð´Ð¾Ð¼ Ð·Ð° Ð³Ð¾Ñ€Ð¾Ð´", orders["out_of_city_orders"]],
            ["Ð¡ Ð¾Ñ‚Ð·Ñ‹Ð²Ð°Ð¼Ð¸", orders["review_orders"]],
        ]

        for row_data in stats_data:
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws1.cell(row=row, column=col_idx, value=value)
                cell.border = thin_border
                if row_data == stats_data[0]:
                    cell.font = Font(bold=True)
                    cell.fill = table_header_fill
                cell.alignment = left_alignment if col_idx == 1 else right_alignment
            row += 1

        row += 1

        # Ð¤Ð¸Ð½Ð°Ð½ÑÐ¾Ð²Ð°Ñ Ð¸Ð½Ñ„Ð¾Ñ€Ð¼Ð°Ñ†Ð¸Ñ
        if orders["closed_orders"] > 0:
            ws1[f"A{row}"] = "Ð¤Ð˜ÐÐÐÐ¡ÐžÐ’ÐÐ¯ Ð˜ÐÐ¤ÐžÐ ÐœÐÐ¦Ð˜Ð¯"
            ws1[f"A{row}"].font = subheader_font
            ws1[f"A{row}"].fill = subheader_fill
            ws1.merge_cells(f"A{row}:F{row}")

            row += 1
            financial_data = [
                ["ÐŸÐ¾ÐºÐ°Ð·Ð°Ñ‚ÐµÐ»ÑŒ", "Ð—Ð½Ð°Ñ‡ÐµÐ½Ð¸Ðµ"],
                ["ÐžÐ±Ñ‰Ð°Ñ ÑÑƒÐ¼Ð¼Ð°", f"{orders['total_amount']:.2f} â‚½"],
                ["Ð Ð°ÑÑ…Ð¾Ð´Ñ‹ Ð½Ð° Ð¼Ð°Ñ‚ÐµÑ€Ð¸Ð°Ð»Ñ‹", f"{orders['total_materials_cost']:.2f} â‚½"],
                ["ÐŸÑ€Ð¸Ð±Ñ‹Ð»ÑŒ Ð¼Ð°ÑÑ‚ÐµÑ€Ð¾Ð²", f"{orders['total_master_profit']:.2f} â‚½"],
                ["ÐŸÑ€Ð¸Ð±Ñ‹Ð»ÑŒ ÐºÐ¾Ð¼Ð¿Ð°Ð½Ð¸Ð¸", f"{orders['total_company_profit']:.2f} â‚½"],
                ["Ð¡Ñ€ÐµÐ´Ð½Ð¸Ð¹ Ñ‡ÐµÐº", f"{report['summary']['avg_order_amount']:.2f} â‚½"],
                ["ÐœÐ°ÐºÑÐ¸Ð¼Ð°Ð»ÑŒÐ½Ñ‹Ð¹ Ñ‡ÐµÐº", f"{report['summary']['max_order_amount']:.2f} â‚½"],
            ]

            for row_data in financial_data:
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws1.cell(row=row, column=col_idx, value=value)
                    cell.border = thin_border
                    if row_data == financial_data[0]:
                        cell.font = Font(bold=True)
                        cell.fill = table_header_fill
                    cell.alignment = left_alignment if col_idx == 1 else right_alignment
                row += 1

        # Ð¨Ð¸Ñ€Ð¸Ð½Ð° ÑÑ‚Ð¾Ð»Ð±Ñ†Ð¾Ð²
        ws1.column_dimensions["A"].width = 30
        ws1.column_dimensions["B"].width = 20

        # Ð›Ð¸ÑÑ‚ 2: Ð”ÐµÑ‚Ð°Ð»Ð¸Ð·Ð°Ñ†Ð¸Ñ Ð¿Ð¾ Ð·Ð°ÐºÐ°Ð·Ð°Ð¼
        closed_orders = report.get("closed_orders", [])
        if closed_orders:
            ws2 = wb.create_sheet(title="Ð—Ð°ÐºÐ°Ð·Ñ‹")

            row = 1
            ws2.merge_cells(f"A{row}:K{row}")
            cell = ws2[f"A{row}"]
            cell.value = "Ð”Ð•Ð¢ÐÐ›Ð˜Ð—ÐÐ¦Ð˜Ð¯ ÐŸÐž Ð—ÐÐšÐÐ—ÐÐœ"
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            ws2.row_dimensions[row].height = 25

            row += 1
            headers = [
                "ID",
                "Ð¢ÐµÑ…Ð½Ð¸ÐºÐ°",
                "ÐšÐ»Ð¸ÐµÐ½Ñ‚",
                "ÐœÐ°ÑÑ‚ÐµÑ€",
                "Ð¡Ð¾Ð·Ð´Ð°Ð½Ð¾",
                "Ð—Ð°ÐºÑ€Ñ‹Ñ‚Ð¾",
                "Ð¡ÑƒÐ¼Ð¼Ð°",
                "ÐœÐ°Ñ‚ÐµÑ€Ð¸Ð°Ð»Ñ‹",
                "ÐŸÑ€Ð¸Ð±Ñ‹Ð»ÑŒ Ð¼Ð°ÑÑ‚ÐµÑ€Ð°",
                "ÐŸÑ€Ð¸Ð±Ñ‹Ð»ÑŒ ÐºÐ¾Ð¼Ð¿Ð°Ð½Ð¸Ð¸",
                "Ð”Ð¾Ð¿. Ð¸Ð½Ñ„Ð¾",
            ]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws2.cell(row=row, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = table_header_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            row += 1

        # Ð”Ð°Ð½Ð½Ñ‹Ðµ Ð¿Ð¾ Ð·Ð°ÐºÐ°Ð·Ð°Ð¼
        for order in closed_orders:
            additional_info = []
            if order["out_of_city"]:
                additional_info.append("Ð’Ñ‹ÐµÐ·Ð´ Ð·Ð° Ð³Ð¾Ñ€Ð¾Ð´")
            if order["has_review"]:
                additional_info.append("ÐžÑ‚Ð·Ñ‹Ð²")

            # Ð‘ÐµÐ·Ð¾Ð¿Ð°ÑÐ½Ð¾Ðµ Ñ„Ð¾Ñ€Ð¼Ð°Ñ‚Ð¸Ñ€Ð¾Ð²Ð°Ð½Ð¸Ðµ ÑÑ‚Ñ€Ð¾Ðº
            equipment_type = (
                str(order["equipment_type"] or "").encode("utf-8", errors="ignore").decode("utf-8")
            )
            client_name = (
                str(order["client_name"] or "").encode("utf-8", errors="ignore").decode("utf-8")
            )
            master_name = (
                str(order["master_name"] or "ÐÐµ Ð½Ð°Ð·Ð½Ð°Ñ‡ÐµÐ½")
                .encode("utf-8", errors="ignore")
                .decode("utf-8")
            )

            data = [
                order["id"],
                equipment_type,
                client_name,
                master_name,
                order.get("created_at"),
                order.get("closed_at"),
                order["total_amount"],
                order["materials_cost"],
                order["master_profit"],
                order["company_profit"],
                ", ".join(additional_info) if additional_info else "-",
            ]

            for col_idx, value in enumerate(data, start=1):
                cell = ws2.cell(row=row, column=col_idx, value=value)
                cell.border = thin_border
                if col_idx == 1:
                    cell.alignment = center_alignment
                elif col_idx in [2, 3, 4, 5, 6, 11]:
                    cell.alignment = left_alignment
                else:
                    cell.alignment = right_alignment
                    if col_idx >= 7 and col_idx <= 10:
                        cell.number_format = "#,##0.00 â‚½"
            row += 1

            # Ð˜Ñ‚Ð¾Ð³Ð¸
            row += 1
            ws2[f"A{row}"] = "Ð˜Ð¢ÐžÐ“Ðž:"
            ws2[f"A{row}"].font = Font(bold=True)
            ws2[f"G{row}"] = sum(o["total_amount"] for o in closed_orders)
            ws2[f"G{row}"].font = Font(bold=True)
            ws2[f"G{row}"].number_format = "#,##0.00 â‚½"
            ws2[f"H{row}"] = sum(o["materials_cost"] for o in closed_orders)
            ws2[f"H{row}"].font = Font(bold=True)
            ws2[f"H{row}"].number_format = "#,##0.00 â‚½"
            ws2[f"I{row}"] = sum(o["master_profit"] for o in closed_orders)
            ws2[f"I{row}"].font = Font(bold=True)
            ws2[f"I{row}"].number_format = "#,##0.00 â‚½"
            ws2[f"J{row}"] = sum(o["company_profit"] for o in closed_orders)
            ws2[f"J{row}"].font = Font(bold=True)
            ws2[f"J{row}"].number_format = "#,##0.00 â‚½"

            # Ð¨Ð¸Ñ€Ð¸Ð½Ð° ÑÑ‚Ð¾Ð»Ð±Ñ†Ð¾Ð²
            ws2.column_dimensions["A"].width = 8
            ws2.column_dimensions["B"].width = 25
            ws2.column_dimensions["C"].width = 20
            ws2.column_dimensions["D"].width = 20
            ws2.column_dimensions["E"].width = 18
            ws2.column_dimensions["F"].width = 18
            ws2.column_dimensions["G"].width = 15
            ws2.column_dimensions["H"].width = 15
            ws2.column_dimensions["I"].width = 18
            ws2.column_dimensions["J"].width = 18
            ws2.column_dimensions["K"].width = 22

        # âœ¨ ÐÐžÐ’Ð«Ð™ Ð›Ð˜Ð¡Ð¢ 3: Ð”ÐµÑ‚Ð°Ð»Ð¸Ð·Ð°Ñ†Ð¸Ñ Ð¿Ð¾ Ð¼Ð°ÑÑ‚ÐµÑ€Ð°Ð¼
        masters = report.get("masters", [])
        if masters:
            ws3 = wb.create_sheet(title="ÐœÐ°ÑÑ‚ÐµÑ€Ð°")

            # Ð—Ð°Ð³Ð¾Ð»Ð¾Ð²Ð¾Ðº
            row = 1
            ws3.merge_cells(f"A{row}:K{row}")
            cell = ws3[f"A{row}"]
            cell.value = "Ð”Ð•Ð¢ÐÐ›Ð˜Ð—ÐÐ¦Ð˜Ð¯ ÐŸÐž ÐœÐÐ¡Ð¢Ð•Ð ÐÐœ"
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            ws3.row_dimensions[row].height = 25

            # Ð—Ð°Ð³Ð¾Ð»Ð¾Ð²ÐºÐ¸ ÐºÐ¾Ð»Ð¾Ð½Ð¾Ðº
            row += 1
            master_headers = [
                "ID",
                "ÐœÐ°ÑÑ‚ÐµÑ€",
                "Ð—Ð°ÑÐ²Ð¾Ðº Ð²ÑÐµÐ³Ð¾",
                "Ð—Ð°Ð²ÐµÑ€ÑˆÐµÐ½Ð¾",
                "Ð’ Ñ€Ð°Ð±Ð¾Ñ‚Ðµ",
                "ÐžÑ‚ÐºÐ°Ð·Ð°Ð½Ð¾",
                "ÐžÐ±Ñ‰Ð°Ñ ÑÑƒÐ¼Ð¼Ð°",
                "ÐœÐ°Ñ‚ÐµÑ€Ð¸Ð°Ð»Ñ‹",
                "Ð§Ð¸ÑÑ‚Ð°Ñ Ð¿Ñ€Ð¸Ð±Ñ‹Ð»ÑŒ",
                "ÐŸÑ€Ð¸Ð±Ñ‹Ð»ÑŒ ÐºÐ¾Ð¼Ð¿Ð°Ð½Ð¸Ð¸",
                "Ð¡Ð´Ð°Ñ‡Ð° Ð² ÐºÐ°ÑÑÑƒ",
                "Ð¡Ñ€ÐµÐ´Ð½Ð¸Ð¹ Ñ‡ÐµÐº",
                "Ð’Ñ‹ÐµÐ·Ð´Ð¾Ð²",
                "ÐžÑ‚Ð·Ñ‹Ð²Ð¾Ð²",
            ]

            for col_idx, header in enumerate(master_headers, start=1):
                cell = ws3.cell(row=row, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = table_header_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            row += 1

            # ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÐ¼ Ð´ÐµÑ‚Ð°Ð»ÑŒÐ½ÑƒÑŽ Ð¸Ð½Ñ„Ð¾Ñ€Ð¼Ð°Ñ†Ð¸ÑŽ Ð¿Ð¾ ÐºÐ°Ð¶Ð´Ð¾Ð¼Ñƒ Ð¼Ð°ÑÑ‚ÐµÑ€Ñƒ
            connection = self._get_connection()
            for master in masters:
                master_id = master["id"]

                # ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÐ¼ Ð´ÐµÑ‚Ð°Ð»ÑŒÐ½ÑƒÑŽ ÑÑ‚Ð°Ñ‚Ð¸ÑÑ‚Ð¸ÐºÑƒ Ð¼Ð°ÑÑ‚ÐµÑ€Ð°
                cursor = await connection.execute(
                    """
                    SELECT
                        COUNT(*) as total_orders,
                        SUM(CASE WHEN status = 'CLOSED' THEN 1 ELSE 0 END) as closed,
                        SUM(CASE WHEN status IN ('ASSIGNED', 'IN_PROGRESS', 'ACCEPTED') THEN 1 ELSE 0 END) as in_work,
                        SUM(CASE WHEN status = 'REFUSED' THEN 1 ELSE 0 END) as refused,
                        SUM(CASE WHEN status = 'CLOSED' THEN total_amount ELSE 0 END) as total_sum,
                        SUM(CASE WHEN status = 'CLOSED' THEN materials_cost ELSE 0 END) as materials_sum,
                        SUM(CASE WHEN status = 'CLOSED' THEN master_profit ELSE 0 END) as master_profit_sum,
                        SUM(CASE WHEN status = 'CLOSED' THEN company_profit ELSE 0 END) as company_profit_sum,
                        SUM(CASE WHEN status = 'CLOSED' AND out_of_city = 1 THEN 1 ELSE 0 END) as out_of_city,
                        SUM(CASE WHEN status = 'CLOSED' AND has_review = 1 THEN 1 ELSE 0 END) as reviews,
                        AVG(CASE WHEN status = 'CLOSED' THEN total_amount ELSE NULL END) as avg_check
                    FROM orders
                    WHERE assigned_master_id = ?
                    """,
                    (master_id,),
                )

                stats_row_raw = await cursor.fetchone()
                if not stats_row_raw:
                    continue

                stats_row: Mapping[str, Any] = dict(stats_row_raw)

                # Ð’Ñ‹Ñ‡Ð¸ÑÐ»ÑÐµÐ¼ Ñ‡Ð¸ÑÑ‚ÑƒÑŽ Ð¿Ñ€Ð¸Ð±Ñ‹Ð»ÑŒ (Ð¾Ð±Ñ‰Ð°Ñ ÑÑƒÐ¼Ð¼Ð° - Ð¼Ð°Ñ‚ÐµÑ€Ð¸Ð°Ð»Ñ‹)
                total_sum = float(stats_row["total_sum"] or 0)
                materials = float(stats_row["materials_sum"] or 0)
                net_profit = total_sum - materials

                # Ð¡Ð´Ð°Ñ‡Ð° Ð² ÐºÐ°ÑÑÑƒ = Ð¿Ñ€Ð¸Ð±Ñ‹Ð»ÑŒ ÐºÐ¾Ð¼Ð¿Ð°Ð½Ð¸Ð¸
                cash_to_company = float(stats_row["company_profit_sum"] or 0)

                # Ð”Ð°Ð½Ð½Ñ‹Ðµ Ð¿Ð¾ Ð¼Ð°ÑÑ‚ÐµÑ€Ñƒ
                master_name = (
                    str(master["name"] or "ÐÐµÐ¸Ð·Ð²ÐµÑÑ‚ÐµÐ½")
                    .encode("utf-8", errors="ignore")
                    .decode("utf-8")
                )
                master_data = [
                    master_id,
                    master_name,
                    stats_row["total_orders"] or 0,
                    stats_row["closed"] or 0,
                    stats_row["in_work"] or 0,
                    stats_row["refused"] or 0,
                    total_sum,
                    materials,
                    net_profit,
                    cash_to_company,
                    cash_to_company,  # Ð¡Ð´Ð°Ñ‡Ð° Ð² ÐºÐ°ÑÑÑƒ = Ð¿Ñ€Ð¸Ð±Ñ‹Ð»ÑŒ ÐºÐ¾Ð¼Ð¿Ð°Ð½Ð¸Ð¸
                    float(stats_row["avg_check"] or 0),
                    stats_row["out_of_city"] or 0,
                    stats_row["reviews"] or 0,
                ]

                for col_idx, value in enumerate(master_data, start=1):
                    cell = ws3.cell(row=row, column=col_idx, value=value)
                    cell.border = thin_border

                    # Ð¤Ð¾Ñ€Ð¼Ð°Ñ‚Ð¸Ñ€Ð¾Ð²Ð°Ð½Ð¸Ðµ
                    if col_idx == 1:  # ID
                        cell.alignment = center_alignment
                    elif col_idx in [2]:  # Ð˜Ð¼Ñ
                        cell.alignment = left_alignment
                    elif col_idx in [3, 4, 5, 6, 13, 14]:  # Ð¡Ñ‡ÐµÑ‚Ñ‡Ð¸ÐºÐ¸
                        cell.alignment = center_alignment
                    else:  # Ð”ÐµÐ½ÑŒÐ³Ð¸
                        cell.alignment = right_alignment
                        if col_idx >= 7 and col_idx <= 12:
                            cell.number_format = "#,##0.00 â‚½"

                row += 1

            # Ð˜Ð¢ÐžÐ“Ðž Ð¿Ð¾ Ð²ÑÐµÐ¼ Ð¼Ð°ÑÑ‚ÐµÑ€Ð°Ð¼
            row += 1
            ws3[f"A{row}"] = "Ð˜Ð¢ÐžÐ“Ðž:"
            ws3[f"A{row}"].font = Font(bold=True, size=12)
            ws3.merge_cells(f"A{row}:B{row}")

            # Ð¡Ñ‡Ð¸Ñ‚Ð°ÐµÐ¼ Ð¸Ñ‚Ð¾Ð³Ð¸
            total_orders = sum(m["orders_count"] for m in masters)
            total_closed = sum(m["closed_orders"] for m in masters)

            # Ð˜Ñ‚Ð¾Ð³Ð¸ Ð¿Ð¾ Ñ„Ð¸Ð½Ð°Ð½ÑÐ°Ð¼ Ð¸Ð· Ð·Ð°ÐºÑ€Ñ‹Ñ‚Ñ‹Ñ… Ð·Ð°ÑÐ²Ð¾Ðº
            total_sum_all = sum(o["total_amount"] for o in closed_orders)
            total_materials_all = sum(o["materials_cost"] for o in closed_orders)
            total_net_profit = total_sum_all - total_materials_all
            total_company = sum(o["company_profit"] for o in closed_orders)

            ws3[f"C{row}"] = total_orders
            ws3[f"C{row}"].font = Font(bold=True)
            ws3[f"C{row}"].alignment = center_alignment

            ws3[f"D{row}"] = total_closed
            ws3[f"D{row}"].font = Font(bold=True)
            ws3[f"D{row}"].alignment = center_alignment

            ws3[f"G{row}"] = total_sum_all
            ws3[f"G{row}"].font = Font(bold=True)
            ws3[f"G{row}"].number_format = "#,##0.00 â‚½"

            ws3[f"H{row}"] = total_materials_all
            ws3[f"H{row}"].font = Font(bold=True)
            ws3[f"H{row}"].number_format = "#,##0.00 â‚½"

            ws3[f"I{row}"] = total_net_profit
            ws3[f"I{row}"].font = Font(bold=True)
            ws3[f"I{row}"].number_format = "#,##0.00 â‚½"

            ws3[f"J{row}"] = total_company
            ws3[f"J{row}"].font = Font(bold=True)
            ws3[f"J{row}"].number_format = "#,##0.00 â‚½"

            ws3[f"K{row}"] = total_company
            ws3[f"K{row}"].font = Font(bold=True)
            ws3[f"K{row}"].number_format = "#,##0.00 â‚½"

            # Ð¨Ð¸Ñ€Ð¸Ð½Ð° ÑÑ‚Ð¾Ð»Ð±Ñ†Ð¾Ð²
            ws3.column_dimensions["A"].width = 6
            ws3.column_dimensions["B"].width = 25
            ws3.column_dimensions["C"].width = 12
            ws3.column_dimensions["D"].width = 12
            ws3.column_dimensions["E"].width = 12
            ws3.column_dimensions["F"].width = 12
            ws3.column_dimensions["G"].width = 15
            ws3.column_dimensions["H"].width = 15
            ws3.column_dimensions["I"].width = 15
            ws3.column_dimensions["J"].width = 18
            ws3.column_dimensions["K"].width = 15
            ws3.column_dimensions["L"].width = 15
            ws3.column_dimensions["M"].width = 10
            ws3.column_dimensions["N"].width = 10

        # âœ¨ ÐÐžÐ’Ð«Ð™ Ð›Ð˜Ð¡Ð¢ 4: Ð’ÑÐµ Ð·Ð°ÑÐ²ÐºÐ¸ Ð¿Ð¾ Ð¼Ð°ÑÑ‚ÐµÑ€Ð°Ð¼ (Ð°ÐºÑ‚Ð¸Ð²Ð½Ñ‹Ðµ Ð¸ Ð·Ð°ÐºÑ€Ñ‹Ñ‚Ñ‹Ðµ)
        if masters:
            ws4 = wb.create_sheet(title="Ð—Ð°ÑÐ²ÐºÐ¸ Ð¿Ð¾ Ð¼Ð°ÑÑ‚ÐµÑ€Ð°Ð¼")

            # Ð—Ð°Ð³Ð¾Ð»Ð¾Ð²Ð¾Ðº
            row = 1
            ws4.merge_cells(f"A{row}:N{row}")
            cell = ws4[f"A{row}"]
            cell.value = "Ð”Ð•Ð¢ÐÐ›Ð˜Ð—ÐÐ¦Ð˜Ð¯ Ð—ÐÐ¯Ð’ÐžÐš ÐŸÐž ÐœÐÐ¡Ð¢Ð•Ð ÐÐœ"
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            ws4.row_dimensions[row].height = 25

            row += 1

            # Ð—Ð°Ð³Ð¾Ð»Ð¾Ð²ÐºÐ¸ ÐºÐ¾Ð»Ð¾Ð½Ð¾Ðº
            order_headers = [
                "ÐœÐ°ÑÑ‚ÐµÑ€",
                "ID",
                "Ð¡Ñ‚Ð°Ñ‚ÑƒÑ",
                "Ð¢Ð¸Ð¿ Ñ‚ÐµÑ…Ð½Ð¸ÐºÐ¸",
                "ÐšÐ»Ð¸ÐµÐ½Ñ‚",
                "ÐÐ´Ñ€ÐµÑ",
                "Ð¢ÐµÐ»ÐµÑ„Ð¾Ð½",
                "Ð¡Ð¾Ð·Ð´Ð°Ð½Ð°",
                "ÐžÐ±Ð½Ð¾Ð²Ð»ÐµÐ½Ð°",
                "Ð¡ÑƒÐ¼Ð¼Ð°",
                "ÐœÐ°Ñ‚ÐµÑ€Ð¸Ð°Ð»Ñ‹",
                "ÐŸÑ€Ð¸Ð±Ñ‹Ð»ÑŒ Ð¼Ð°ÑÑ‚ÐµÑ€Ð°",
                "Ð¡Ð´Ð°Ñ‡Ð° Ð² ÐºÐ°ÑÑÑƒ",
                "ÐŸÑ€Ð¸Ð¼ÐµÑ‡Ð°Ð½Ð¸Ñ",
            ]

            for col_idx, header in enumerate(order_headers, start=1):
                cell = ws4.cell(row=row, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = subheader_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            row += 1

            # ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÐ¼ Ð²ÑÐµ Ð·Ð°ÑÐ²ÐºÐ¸ Ð´Ð»Ñ ÐºÐ°Ð¶Ð´Ð¾Ð³Ð¾ Ð¼Ð°ÑÑ‚ÐµÑ€Ð°
            connection = self._get_connection()
            for master in masters:
                master_id = master["id"]
                master_name = (
                    str(master["name"] or "ÐÐµÐ¸Ð·Ð²ÐµÑÑ‚ÐµÐ½")
                    .encode("utf-8", errors="ignore")
                    .decode("utf-8")
                )

                # ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÐ¼ Ð’Ð¡Ð• Ð·Ð°ÑÐ²ÐºÐ¸ Ð¼Ð°ÑÑ‚ÐµÑ€Ð° (Ð°ÐºÑ‚Ð¸Ð²Ð½Ñ‹Ðµ Ð¸ Ð·Ð°ÐºÑ€Ñ‹Ñ‚Ñ‹Ðµ)
                cursor = await connection.execute(
                    """
                    SELECT
                        o.id,
                        o.status,
                        o.equipment_type,
                        o.client_name,
                        o.client_address,
                        o.client_phone,
                        o.created_at,
                        o.updated_at,
                        o.total_amount,
                        o.materials_cost,
                        o.master_profit,
                        o.company_profit,
                        o.notes,
                        o.scheduled_time,
                        o.out_of_city,
                        o.has_review
                    FROM orders o
                    WHERE o.assigned_master_id = ?
                        AND o.status IN ('ASSIGNED', 'ACCEPTED', 'IN_PROGRESS', 'COMPLETED', 'CLOSED', 'REFUSED')
                        AND o.deleted_at IS NULL
                    ORDER BY
                        CASE o.status
                            WHEN 'IN_PROGRESS' THEN 1
                            WHEN 'ACCEPTED' THEN 2
                            WHEN 'ASSIGNED' THEN 3
                            WHEN 'COMPLETED' THEN 4
                            WHEN 'CLOSED' THEN 5
                            WHEN 'REFUSED' THEN 6
                            ELSE 7
                        END,
                        o.created_at DESC
                    """,
                    (master_id,),
                )

                orders_rows = await cursor.fetchall()

                if not orders_rows:
                    continue

                # Ð”Ð¾Ð±Ð°Ð²Ð»ÑÐµÐ¼ Ð·Ð°Ð³Ð¾Ð»Ð¾Ð²Ð¾Ðº Ð¼Ð°ÑÑ‚ÐµÑ€Ð°
                cell_master = ws4[f"A{row}"]
                cell_master.value = f"ðŸ‘¨â€ðŸ”§ {master_name}"
                cell_master.font = Font(bold=True, size=11, color="FFFFFF")
                cell_master.fill = PatternFill(
                    start_color="70AD47", end_color="70AD47", fill_type="solid"
                )
                cell_master.alignment = left_alignment
                ws4.merge_cells(f"A{row}:N{row}")
                ws4.row_dimensions[row].height = 20
                row += 1

                # Ð—Ð°ÑÐ²ÐºÐ¸ Ð¼Ð°ÑÑ‚ÐµÑ€Ð°
                for order_row in orders_rows:
                    # Ð­Ð¼Ð¾Ð´Ð·Ð¸ Ð´Ð»Ñ ÑÑ‚Ð°Ñ‚ÑƒÑÐ°
                    status_emoji = {
                        "ASSIGNED": "ðŸ†•",
                        "ACCEPTED": "âœ…",
                        "IN_PROGRESS": "âš™ï¸",
                        "COMPLETED": "âœ”ï¸",
                        "CLOSED": "ðŸ”’",
                        "REFUSED": "âŒ",
                    }.get(order_row["status"], "â“")

                    # Ð¤Ð¾Ñ€Ð¼Ð°Ñ‚Ð¸Ñ€ÑƒÐµÐ¼ Ð¿Ñ€Ð¸Ð¼ÐµÑ‡Ð°Ð½Ð¸Ñ
                    notes = []
                    if order_row["out_of_city"]:
                        notes.append("Ð’Ñ‹ÐµÐ·Ð´ Ð·Ð° Ð³Ð¾Ñ€Ð¾Ð´")
                    if order_row["has_review"]:
                        notes.append("Ð•ÑÑ‚ÑŒ Ð¾Ñ‚Ð·Ñ‹Ð²")
                    if order_row["scheduled_time"] and order_row["scheduled_time"] != "None":
                        scheduled_time = (
                            str(order_row["scheduled_time"] or "")
                            .encode("utf-8", errors="ignore")
                            .decode("utf-8")
                        )
                        notes.append(f"Ð’Ñ€ÐµÐ¼Ñ: {scheduled_time}")
                    if order_row["notes"]:
                        notes_text = (
                            str(order_row["notes"] or "")
                            .encode("utf-8", errors="ignore")
                            .decode("utf-8")
                        )
                        notes.append(notes_text[:50])

                    # Ð‘ÐµÐ·Ð¾Ð¿Ð°ÑÐ½Ð¾Ðµ Ñ„Ð¾Ñ€Ð¼Ð°Ñ‚Ð¸Ñ€Ð¾Ð²Ð°Ð½Ð¸Ðµ ÑÑ‚Ñ€Ð¾Ðº
                    equipment_type = (
                        str(order_row["equipment_type"] or "")
                        .encode("utf-8", errors="ignore")
                        .decode("utf-8")
                    )
                    client_name = (
                        str(order_row["client_name"] or "")
                        .encode("utf-8", errors="ignore")
                        .decode("utf-8")
                    )
                    client_address = (
                        str(order_row["client_address"] or "")
                        .encode("utf-8", errors="ignore")
                        .decode("utf-8")
                    )
                    client_phone = (
                        str(order_row["client_phone"] or "")
                        .encode("utf-8", errors="ignore")
                        .decode("utf-8")
                    )

                    # Ð‘ÐµÐ·Ð¾Ð¿Ð°ÑÐ½Ð¾Ðµ Ð¾Ð±Ñ€ÐµÐ·Ð°Ð½Ð¸Ðµ Ð°Ð´Ñ€ÐµÑÐ°
                    if len(client_address) > 30:
                        client_address = client_address[:30] + "..."

                    # Ð‘ÐµÐ·Ð¾Ð¿Ð°ÑÐ½Ð¾Ðµ Ñ„Ð¾Ñ€Ð¼Ð°Ñ‚Ð¸Ñ€Ð¾Ð²Ð°Ð½Ð¸Ðµ Ð´Ð°Ñ‚
                    created_at = ""
                    updated_at = ""
                    if order_row["created_at"]:
                        try:
                            created_at = str(order_row["created_at"])[:16]
                        except Exception:
                            created_at = ""
                    if order_row["updated_at"]:
                        try:
                            updated_at = str(order_row["updated_at"])[:16]
                        except Exception:
                            updated_at = ""

                    order_data = [
                        "",  # ÐŸÑƒÑÑ‚Ð°Ñ ÐºÐ¾Ð»Ð¾Ð½ÐºÐ° Ð´Ð»Ñ Ð¼Ð°ÑÑ‚ÐµÑ€Ð° (Ð¾Ð½ Ð² Ð·Ð°Ð³Ð¾Ð»Ð¾Ð²ÐºÐµ)
                        order_row["id"],
                        f"{status_emoji} {order_row['status']}",
                        equipment_type,
                        client_name,
                        client_address,
                        client_phone,
                        created_at,
                        updated_at,
                        float(order_row["total_amount"] or 0),
                        float(order_row["materials_cost"] or 0),
                        float(order_row["master_profit"] or 0),
                        float(order_row["company_profit"] or 0),  # Ð¡Ð´Ð°Ñ‡Ð° Ð² ÐºÐ°ÑÑÑƒ
                        "; ".join(notes) if notes else "-",
                    ]

                    for col_idx, value in enumerate(order_data, start=1):
                        cell = ws4.cell(row=row, column=col_idx, value=value)
                        cell.border = thin_border

                        # Ð¤Ð¾Ñ€Ð¼Ð°Ñ‚Ð¸Ñ€Ð¾Ð²Ð°Ð½Ð¸Ðµ
                        if col_idx == 2:  # ID
                            cell.alignment = center_alignment
                            cell.font = Font(bold=True)
                        elif col_idx == 3:  # Ð¡Ñ‚Ð°Ñ‚ÑƒÑ
                            cell.alignment = center_alignment
                            # Ð¦Ð²ÐµÑ‚ Ñ„Ð¾Ð½Ð° Ð² Ð·Ð°Ð²Ð¸ÑÐ¸Ð¼Ð¾ÑÑ‚Ð¸ Ð¾Ñ‚ ÑÑ‚Ð°Ñ‚ÑƒÑÐ°
                            status = order_row["status"]
                            if status == "IN_PROGRESS":
                                cell.fill = PatternFill(
                                    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                                )
                            elif status == "CLOSED":
                                cell.fill = PatternFill(
                                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                                )
                            elif status == "REFUSED":
                                cell.fill = PatternFill(
                                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                                )
                        elif col_idx in [4, 5, 6, 7, 8, 14]:  # Ð¢ÐµÐºÑÑ‚
                            cell.alignment = left_alignment
                        else:  # Ð§Ð¸ÑÐ»Ð° Ð¸ Ð´ÐµÐ½ÑŒÐ³Ð¸
                            cell.alignment = right_alignment
                            if col_idx >= 10 and col_idx <= 13:
                                cell.number_format = "#,##0.00 â‚½"

                    row += 1

                # Ð˜Ñ‚Ð¾Ð³Ð¸ Ð¿Ð¾ Ð¼Ð°ÑÑ‚ÐµÑ€Ñƒ
                cursor = await connection.execute(
                    """
                    SELECT
                        COUNT(*) as total,
                        SUM(CASE WHEN status = 'CLOSED' THEN total_amount ELSE 0 END) as sum_total,
                        SUM(CASE WHEN status = 'CLOSED' THEN materials_cost ELSE 0 END) as sum_materials,
                        SUM(CASE WHEN status = 'CLOSED' THEN master_profit ELSE 0 END) as sum_master,
                        SUM(CASE WHEN status = 'CLOSED' THEN company_profit ELSE 0 END) as sum_company
                    FROM orders
                    WHERE assigned_master_id = ?
                        AND deleted_at IS NULL
                    """,
                    (master_id,),
                )

                totals_row = await cursor.fetchone()
                totals: Mapping[str, Any] = (
                    dict(totals_row)
                    if totals_row is not None
                    else {
                        "sum_total": 0,
                        "sum_materials": 0,
                        "sum_master": 0,
                        "sum_company": 0,
                    }
                )

                cell_total = ws4[f"A{row}"]
                cell_total.value = f"Ð˜Ñ‚Ð¾Ð³Ð¾ Ð¿Ð¾ {master_name}:"
                cell_total.font = Font(bold=True, italic=True)
                ws4.merge_cells(f"A{row}:I{row}")

                cell_j = ws4[f"J{row}"]
                cell_j.value = float(totals["sum_total"] or 0)
                cell_j.font = Font(bold=True)
                cell_j.number_format = "#,##0.00 â‚½"

                cell_k = ws4[f"K{row}"]
                cell_k.value = float(totals["sum_materials"] or 0)
                cell_k.font = Font(bold=True)
                cell_k.number_format = "#,##0.00 â‚½"

                cell_l = ws4[f"L{row}"]
                cell_l.value = float(totals["sum_master"] or 0)
                cell_l.font = Font(bold=True)
                cell_l.number_format = "#,##0.00 â‚½"

                cell_m = ws4[f"M{row}"]
                cell_m.value = float(totals["sum_company"] or 0)
                cell_m.font = Font(bold=True)
                cell_m.number_format = "#,##0.00 â‚½"

                row += 2  # ÐŸÑƒÑÑ‚Ð°Ñ ÑÑ‚Ñ€Ð¾ÐºÐ° Ð¼ÐµÐ¶Ð´Ñƒ Ð¼Ð°ÑÑ‚ÐµÑ€Ð°Ð¼Ð¸

            # Ð¨Ð¸Ñ€Ð¸Ð½Ð° ÑÑ‚Ð¾Ð»Ð±Ñ†Ð¾Ð²
            ws4.column_dimensions["A"].width = 25
            ws4.column_dimensions["B"].width = 6
            ws4.column_dimensions["C"].width = 15
            ws4.column_dimensions["D"].width = 20
            ws4.column_dimensions["E"].width = 20
            ws4.column_dimensions["F"].width = 30
            ws4.column_dimensions["G"].width = 15
            ws4.column_dimensions["H"].width = 18
            ws4.column_dimensions["I"].width = 18
            ws4.column_dimensions["J"].width = 15
            ws4.column_dimensions["K"].width = 15
            ws4.column_dimensions["L"].width = 18
            ws4.column_dimensions["M"].width = 18
            ws4.column_dimensions["N"].width = 35

        # Ð¡Ð¾Ñ…Ñ€Ð°Ð½ÑÐµÐ¼ Ñ„Ð°Ð¹Ð»
        wb.save(file_path)
        logger.info(f"Excel Ð¾Ñ‚Ñ‡ÐµÑ‚ ÑÐ¾Ñ…Ñ€Ð°Ð½ÐµÐ½: {file_path}")

        return str(file_path)
