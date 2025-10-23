import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.database import Database
from app.utils.helpers import get_now, MOSCOW_TZ


logger = logging.getLogger(__name__)


class MasterArchiveService:
    """Сервис для архивирования заявок мастера"""
    
    def __init__(self):
        self.db = Database()
    
    async def archive_master_orders(self, master_id: int, reason: str = "deactivation") -> Optional[str]:
        """
        Архивирование всех заявок мастера
        
        Args:
            master_id: ID мастера
            reason: Причина архивирования (deactivation/firing)
            
        Returns:
            Путь к созданному архиву или None
        """
        try:
            await self.db.connect()
            
            # Получаем информацию о мастере
            master = await self.db.get_master_by_id(master_id)
            if not master:
                logger.error(f"Master with ID {master_id} not found")
                return None
            
            # Получаем все заявки мастера
            orders = await self.db.get_orders_by_master(master_id, exclude_closed=False)
            
            if not orders:
                logger.info(f"No orders found for master {master_id}")
                return None
            
            # Создаем архив
            archive_path = await self._create_archive(master, orders, reason)
            
            # Сохраняем информацию об архиве в базу
            await self._save_archive_info(master.id, archive_path, reason, len(orders))
            
            logger.info(f"Archived {len(orders)} orders for master {master.id} ({reason})")
            return archive_path
            
        except Exception as e:
            logger.error(f"Error archiving master orders: {e}")
            return None
        finally:
            await self.db.disconnect()
    
    async def _create_archive(self, master, orders: list, reason: str) -> str:
        """Создание Excel архива заявок мастера"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Архив заявок"
        
        # Стили
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        data_font = Font(size=11)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        # Заголовок
        # A1: "Архив заявок мастера:"
        cell_a1 = ws.cell(row=1, column=1)
        cell_a1.value = "Архив заявок мастера:"
        cell_a1.font = header_font
        cell_a1.fill = header_fill
        cell_a1.alignment = center_alignment
        
        # B1: имя мастера
        cell_b1 = ws.cell(row=1, column=2)
        cell_b1.value = master.get_display_name()
        cell_b1.font = header_font
        cell_b1.fill = header_fill
        cell_b1.alignment = center_alignment
        
        ws.row_dimensions[1].height = 30
        # Растягиваем заголовок на остальные столбцы
        for col in range(3, 9):  # C1:H1
            ws.cell(row=1, column=col).fill = header_fill
        
        # Информация о мастере
        ws.cell(row=2, column=1, value="Мастер:").font = Font(bold=True)
        ws.cell(row=2, column=2, value=master.get_display_name()).font = data_font
        ws.cell(row=3, column=1, value="Телефон:").font = Font(bold=True)
        ws.cell(row=3, column=2, value=master.phone).font = data_font
        ws.cell(row=4, column=1, value="Специализация:").font = Font(bold=True)
        ws.cell(row=4, column=2, value=master.specialization).font = data_font
        ws.cell(row=5, column=1, value="Причина архивирования:").font = Font(bold=True)
        reason_text = "Деактивация" if reason == "deactivation" else "Увольнение"
        ws.cell(row=5, column=2, value=reason_text).font = data_font
        ws.cell(row=6, column=1, value="Дата архивирования:").font = Font(bold=True)
        ws.cell(row=6, column=2, value=get_now().strftime("%d.%m.%Y %H:%M")).font = data_font
        
        # Заголовки столбцов
        headers = ["№ Заказа", "Статус", "Сумма", "Тип техники", "Дата создания", "Дата обновления", "Выезд", "Отзыв"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=8, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Данные заявок
        row = 9
        for order in orders:
            status_text = {
                "NEW": "Новая",
                "ASSIGNED": "Назначена",
                "IN_PROGRESS": "В работе",
                "COMPLETED": "Завершена",
                "CLOSED": "Закрыта",
                "CANCELLED": "Отменена"
            }.get(order.status, order.status)
            
            data = [
                order.id,
                status_text,
                order.total_amount or 0,
                order.equipment_type or "",
                order.created_at.strftime("%d.%m.%Y %H:%M") if order.created_at else "",
                order.updated_at.strftime("%d.%m.%Y %H:%M") if order.updated_at else "",
                "Да" if getattr(order, 'is_onsite', False) else "Нет",
                getattr(order, 'review', '') or ""
            ]
            
            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
            
            row += 1
        
        # Итоговая статистика
        total_orders = len(orders)
        completed_orders = len([o for o in orders if o.status == "CLOSED"])
        total_amount = sum(o.total_amount or 0 for o in orders)
        
        ws.cell(row=row+1, column=1, value="Итого заявок:").font = Font(bold=True)
        ws.cell(row=row+1, column=2, value=total_orders).font = data_font
        ws.cell(row=row+2, column=1, value="Завершено:").font = Font(bold=True)
        ws.cell(row=row+2, column=2, value=completed_orders).font = data_font
        ws.cell(row=row+3, column=1, value="Общая сумма:").font = Font(bold=True)
        ws.cell(row=row+3, column=2, value=f"{total_amount:,.0f} ₽").font = data_font
        
        # Настройка ширины столбцов
        column_widths = [12, 15, 15, 20, 18, 18, 10, 30]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
        
        # Создаем директорию для архивов
        archive_dir = Path("reports/archives")
        archive_dir.mkdir(parents=True, exist_ok=True)
        
        # Сохраняем файл
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"master_{master.id}_archive_{timestamp}.xlsx"
        filepath = archive_dir / filename
        
        wb.save(filepath)
        return str(filepath)
    
    async def _save_archive_info(self, master_id: int, filepath: str, reason: str, order_count: int):
        """Сохранение информации об архиве в базу данных"""
        try:
            # Проверяем, какая база данных используется
            if hasattr(self.db, 'connection'):
                # Обычная база данных
                await self.db.connection.execute("""
                    CREATE TABLE IF NOT EXISTS master_archives (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        master_id INTEGER NOT NULL,
                        file_path TEXT NOT NULL,
                        reason TEXT NOT NULL,
                        order_count INTEGER NOT NULL,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (master_id) REFERENCES masters (id)
                    )
                """)
                
                await self.db.connection.execute("""
                    INSERT INTO master_archives (master_id, file_path, reason, order_count)
                    VALUES (?, ?, ?, ?)
                """, (master_id, filepath, reason, order_count))
                
                await self.db.connection.commit()
            else:
                # ORM база данных - используем raw SQL
                from sqlalchemy import text
                async with self.db.get_session() as session:
                    await session.execute(text("""
                        CREATE TABLE IF NOT EXISTS master_archives (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            master_id INTEGER NOT NULL,
                            file_path TEXT NOT NULL,
                            reason TEXT NOT NULL,
                            order_count INTEGER NOT NULL,
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            FOREIGN KEY (master_id) REFERENCES masters (id)
                        )
                    """))
                    
                    await session.execute(text("""
                        INSERT INTO master_archives (master_id, file_path, reason, order_count)
                        VALUES (:master_id, :file_path, :reason, :order_count)
                    """), {
                        "master_id": master_id,
                        "file_path": filepath,
                        "reason": reason,
                        "order_count": order_count
                    })
                    
                    await session.commit()
            
            logger.info(f"Archive info saved for master {master_id}")
            
        except Exception as e:
            logger.error(f"Error saving archive info: {e}")
            if hasattr(self.db, 'connection'):
                await self.db.connection.rollback()
            else:
                await self.db.session.rollback()
