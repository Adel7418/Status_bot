"""
Утилиты для форматирования Excel-таблиц
"""

from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.services.excel.styles import ExcelStyles


class ExcelFormatter:
    """Утилиты для форматирования Excel-таблиц"""

    @staticmethod
    def format_header_row(
        ws: Worksheet, row_idx: int, columns: list[str], merge_range: str | None = None
    ) -> None:
        """
        Форматирование строки заголовка

        Args:
            ws: Рабочий лист Excel
            row_idx: Номер строки (1-based)
            columns: Список названий колонок
            merge_range: Диапазон ячеек для объединения (например, "A1:D1") или None
        """
        if merge_range:
            ws.merge_cells(merge_range)
            cell = ws[merge_range.split(":")[0]]
            cell.value = columns[0] if columns else ""
            cell.font = ExcelStyles.HEADER_FONT
            cell.fill = ExcelStyles.HEADER_FILL
            cell.alignment = ExcelStyles.CENTER_ALIGNMENT
            cell.border = ExcelStyles.THIN_BORDER
        else:
            for col_idx, column_name in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = column_name
                cell.font = ExcelStyles.HEADER_FONT
                cell.fill = ExcelStyles.HEADER_FILL
                cell.alignment = ExcelStyles.CENTER_ALIGNMENT
                cell.border = ExcelStyles.THIN_BORDER

    @staticmethod
    def format_subheader_row(
        ws: Worksheet, row_idx: int, text: str, merge_range: str | None = None
    ) -> None:
        """
        Форматирование строки подзаголовка

        Args:
            ws: Рабочий лист Excel
            row_idx: Номер строки (1-based)
            text: Текст подзаголовка
            merge_range: Диапазон ячеек для объединения (например, "A1:D1")
        """
        if merge_range:
            ws.merge_cells(merge_range)
            cell = ws[merge_range.split(":")[0]]
        else:
            cell = ws.cell(row=row_idx, column=1)

        cell.value = text
        cell.font = ExcelStyles.SUBHEADER_FONT
        cell.fill = ExcelStyles.SUBHEADER_FILL
        cell.alignment = ExcelStyles.CENTER_ALIGNMENT
        cell.border = ExcelStyles.THIN_BORDER

    @staticmethod
    def format_table_header(ws: Worksheet, row_idx: int, columns: list[str]) -> None:
        """
        Форматирование заголовка таблицы

        Args:
            ws: Рабочий лист Excel
            row_idx: Номер строки (1-based)
            columns: Список названий колонок
        """
        for col_idx, column_name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = column_name
            cell.font = ExcelStyles.TABLE_HEADER_FONT
            cell.fill = ExcelStyles.HEADER_FILL
            cell.alignment = ExcelStyles.CENTER_ALIGNMENT
            cell.border = ExcelStyles.THIN_BORDER

    @staticmethod
    def format_data_row(
        ws: Worksheet, row_idx: int, values: list[Any], centered: bool = False
    ) -> None:
        """
        Форматирование строки с данными

        Args:
            ws: Рабочий лист Excel
            row_idx: Номер строки (1-based)
            values: Список значений для ячеек
            centered: Использовать центрирование или выравнивание по левому краю
        """
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = ExcelStyles.DATA_FONT
            cell.alignment = (
                ExcelStyles.CENTER_ALIGNMENT if centered else ExcelStyles.LEFT_ALIGNMENT
            )
            cell.border = ExcelStyles.THIN_BORDER

    @staticmethod
    def auto_adjust_column_widths(ws: Worksheet, min_width: int = 10, max_width: int = 50) -> None:
        """
        Автоматическая подстройка ширины колонок

        Args:
            ws: Рабочий лист Excel
            min_width: Минимальная ширина колонки
            max_width: Максимальная ширина колонки
        """
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except Exception:
                    pass

            adjusted_width = max(min_width, min(max_length + 2, max_width))
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def set_column_widths(ws: Worksheet, widths: dict[str, int]) -> None:
        """
        Установить конкретную ширину для указанных колонок

        Args:
            ws: Рабочий лист Excel
            widths: Словарь {column_letter: width}, например {"A": 15, "B": 25}
        """
        for column_letter, width in widths.items():
            ws.column_dimensions[column_letter].width = width

    @staticmethod
    def set_row_height(ws: Worksheet, row_idx: int, height: int) -> None:
        """
        Установить высоту строки

        Args:
            ws: Рабочий лист Excel
            row_idx: Номер строки (1-based)
            height: Высота в пунктах
        """
        ws.row_dimensions[row_idx].height = height

    @staticmethod
    def apply_cell_style(ws: Worksheet, cell_ref: str, style_dict: dict) -> None:
        """
        Применить стиль к ячейке

        Args:
            ws: Рабочий лист Excel
            cell_ref: Ссылка на ячейку (например, "A1")
            style_dict: Словарь со стилями (font, fill, alignment, border)
        """
        cell = ws[cell_ref]
        if "font" in style_dict:
            cell.font = style_dict["font"]
        if "fill" in style_dict:
            cell.fill = style_dict["fill"]
        if "alignment" in style_dict:
            cell.alignment = style_dict["alignment"]
        if "border" in style_dict:
            cell.border = style_dict["border"]

    @staticmethod
    def create_merged_header(
        ws: Worksheet, row_idx: int, text: str, start_col: str, end_col: str, style: str = "header"
    ) -> None:
        """
        Создать объединенный заголовок

        Args:
            ws: Рабочий лист Excel
            row_idx: Номер строки (1-based)
            text: Текст заголовка
            start_col: Начальная колонка (например, "A")
            end_col: Конечная колонка (например, "D")
            style: Тип стиля ("header", "subheader", "table_header")
        """
        merge_range = f"{start_col}{row_idx}:{end_col}{row_idx}"
        ws.merge_cells(merge_range)
        cell = ws[f"{start_col}{row_idx}"]
        cell.value = text

        if style == "header":
            style_dict = ExcelStyles.get_header_style()
        elif style == "subheader":
            style_dict = ExcelStyles.get_subheader_style()
        elif style == "table_header":
            style_dict = ExcelStyles.get_table_header_style()
        else:
            style_dict = ExcelStyles.get_data_style()

        cell.font = style_dict["font"]
        cell.fill = style_dict["fill"]
        cell.alignment = style_dict["alignment"]
        cell.border = style_dict["border"]

    @staticmethod
    def format_currency_column(ws: Worksheet, column_letter: str, start_row: int = 2) -> None:
        """
        Форматировать колонку как валюту (с двумя знаками после запятой)

        Args:
            ws: Рабочий лист Excel
            column_letter: Буква колонки (например, "C")
            start_row: Начальная строка для форматирования (по умолчанию 2, пропускает заголовок)
        """
        for row in range(start_row, ws.max_row + 1):
            cell = ws[f"{column_letter}{row}"]
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00 ₽"

    @staticmethod
    def format_percentage_column(ws: Worksheet, column_letter: str, start_row: int = 2) -> None:
        """
        Форматировать колонку как проценты

        Args:
            ws: Рабочий лист Excel
            column_letter: Буква колонки (например, "D")
            start_row: Начальная строка для форматирования
        """
        for row in range(start_row, ws.max_row + 1):
            cell = ws[f"{column_letter}{row}"]
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0%"

    @staticmethod
    def highlight_row(ws: Worksheet, row_idx: int, fill_type: str = "highlight") -> None:
        """
        Подсветить строку цветом

        Args:
            ws: Рабочий лист Excel
            row_idx: Номер строки (1-based)
            fill_type: Тип заливки ("highlight", "success", "warning", "error")
        """
        fill_map = {
            "highlight": ExcelStyles.HIGHLIGHT_FILL,
            "success": ExcelStyles.SUCCESS_FILL,
            "warning": ExcelStyles.WARNING_FILL,
            "error": ExcelStyles.ERROR_FILL,
        }
        fill = fill_map.get(fill_type, ExcelStyles.HIGHLIGHT_FILL)

        for cell in ws[row_idx]:
            cell.fill = fill




