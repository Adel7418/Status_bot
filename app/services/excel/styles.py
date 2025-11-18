"""
Централизованные стили для Excel-отчетов
"""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class ExcelStyles:
    """Централизованные стили для Excel-отчетов"""

    # ======================
    # ШРИФТЫ
    # ======================

    HEADER_FONT = Font(bold=True, size=14, color="FFFFFF")
    """Шрифт для главных заголовков (белый, жирный, 14pt)"""

    SUBHEADER_FONT = Font(bold=True, size=12)
    """Шрифт для подзаголовков (черный, жирный, 12pt)"""

    TABLE_HEADER_FONT = Font(bold=True, size=12, color="000000")
    """Шрифт для заголовков таблиц (чёрный, жирный, 12pt)"""

    DATA_FONT = Font(size=10)
    """Шрифт для обычных данных (черный, 10pt)"""

    BOLD_FONT = Font(bold=True, size=11)
    """Жирный шрифт для выделения (11pt)"""

    SMALL_BOLD_FONT = Font(bold=True, size=10)
    """Малый жирный шрифт (10pt)"""

    ITALIC_FONT = Font(italic=True, size=10)
    """Курсивный шрифт (10pt)"""

    MASTER_NAME_FONT = Font(bold=True, size=11, color="FFFFFF")
    """Шрифт для имени мастера (белый, жирный, 11pt)"""

    SIMPLE_BOLD_FONT = Font(bold=True)
    """Простой жирный шрифт без указания размера"""

    SIMPLE_ITALIC_FONT = Font(italic=True)
    """Простой курсивный шрифт без указания размера"""

    BOLD_ITALIC_FONT = Font(bold=True, italic=True)
    """Жирный курсивный шрифт"""

    # ======================
    # ЗАЛИВКИ (ФОНЫ)
    # ======================

    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    """Синий фон для главных заголовков (#4472C4)"""

    SUBHEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    """Светло-синий фон для подзаголовков (#D9E1F2)"""

    HIGHLIGHT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    """Желтый фон для подсветки (#FFF2CC)"""

    SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    """Зеленый фон для успешных операций (#C6EFCE)"""

    WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    """Оранжевый фон для предупреждений (#FFEB9C)"""

    ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    """Красный фон для ошибок (#FFC7CE)"""

    TABLE_HEADER_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    """Серый фон для заголовков таблиц (#E7E6E6)"""

    MASTER_HEADER_FILL = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    """Зеленый фон для заголовков мастеров (#70AD47)"""

    # ======================
    # ВЫРАВНИВАНИЕ
    # ======================

    CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    """Центрирование по горизонтали и вертикали"""

    LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center")
    """Выравнивание по левому краю с вертикальным центрированием"""

    RIGHT_ALIGNMENT = Alignment(horizontal="right", vertical="center")
    """Выравнивание по правому краю с вертикальным центрированием"""

    WRAP_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    """Выравнивание по левому краю с переносом текста"""

    WRAP_TOP_ALIGNMENT = Alignment(horizontal="left", vertical="top", wrap_text=True)
    """Выравнивание по левому краю с переносом текста и выравниванием по верху"""

    # ======================
    # ГРАНИЦЫ
    # ======================

    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    """Тонкая граница со всех сторон"""

    THICK_BORDER = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick"),
    )
    """Толстая граница со всех сторон"""

    NO_BORDER = Border()
    """Без границы"""

    # ======================
    # РАЗМЕРЫ
    # ======================

    # Высоты строк
    HEADER_ROW_HEIGHT = 25
    """Высота строки заголовка (25)"""

    SUBHEADER_ROW_HEIGHT = 20
    """Высота строки подзаголовка (20)"""

    # Ширины колонок (стандартные)
    COL_WIDTH_ID = 12
    """Ширина колонки ID (12)"""

    COL_WIDTH_SHORT = 15
    """Ширина короткой колонки (15)"""

    COL_WIDTH_MEDIUM = 20
    """Ширина средней колонки (20)"""

    COL_WIDTH_LONG = 30
    """Ширина длинной колонки (30)"""

    COL_WIDTH_EXTRA_LONG = 40
    """Ширина очень длинной колонки (40)"""

    COL_WIDTH_PHONE = 18
    """Ширина колонки для телефона (18)"""

    COL_WIDTH_DATE = 20
    """Ширина колонки для даты (20)"""

    COL_WIDTH_MONEY = 15
    """Ширина колонки для денежных значений (15)"""

    # ======================
    # КОМБИНИРОВАННЫЕ СТИЛИ
    # ======================

    @staticmethod
    def get_header_style() -> dict:
        """
        Получить полный стиль для главного заголовка

        Returns:
            Dict со всеми свойствами стиля
        """
        return {
            "font": ExcelStyles.HEADER_FONT,
            "fill": ExcelStyles.HEADER_FILL,
            "alignment": ExcelStyles.CENTER_ALIGNMENT,
            "border": ExcelStyles.THIN_BORDER,
        }

    @staticmethod
    def get_subheader_style() -> dict:
        """
        Получить полный стиль для подзаголовка

        Returns:
            Dict со всеми свойствами стиля
        """
        return {
            "font": ExcelStyles.SUBHEADER_FONT,
            "fill": ExcelStyles.SUBHEADER_FILL,
            "alignment": ExcelStyles.CENTER_ALIGNMENT,
            "border": ExcelStyles.THIN_BORDER,
        }

    @staticmethod
    def get_table_header_style() -> dict:
        """
        Получить полный стиль для заголовка таблицы

        Returns:
            Dict со всеми свойствами стиля
        """
        return {
            "font": ExcelStyles.TABLE_HEADER_FONT,
            "fill": ExcelStyles.TABLE_HEADER_FILL,
            "alignment": ExcelStyles.CENTER_ALIGNMENT,
            "border": ExcelStyles.THIN_BORDER,
        }

    @staticmethod
    def get_data_style(centered: bool = False) -> dict:
        """
        Получить полный стиль для обычных данных

        Args:
            centered: Использовать центрирование или выравнивание по левому краю

        Returns:
            Dict со всеми свойствами стиля
        """
        return {
            "font": ExcelStyles.DATA_FONT,
            "alignment": ExcelStyles.CENTER_ALIGNMENT if centered else ExcelStyles.LEFT_ALIGNMENT,
            "border": ExcelStyles.THIN_BORDER,
        }
