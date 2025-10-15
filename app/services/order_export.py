"""
Сервис экспорта деталей заявки в Excel
"""

import logging
from io import BytesIO

from aiogram.types import BufferedInputFile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.config import OrderStatus
from app.database import Database
from app.utils import format_datetime

logger = logging.getLogger(__name__)


class OrderExportService:
    """Сервис для экспорта заявок в Excel"""

    @staticmethod
    async def export_order_to_excel(order_id: int) -> BufferedInputFile | None:
        """
        Экспорт деталей заявки в Excel

        Args:
            order_id: ID заявки

        Returns:
            BufferedInputFile с Excel файлом или None если заявка не найдена
        """
        db = Database()
        await db.connect()

        try:
            order = await db.get_order_by_id(order_id)

            if not order:
                return None

            # Создаем workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"Заявка #{order_id}"

            # Стили
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Заголовок
            ws.merge_cells("A1:B1")
            ws["A1"] = f"ЗАЯВКА #{order.id}"
            ws["A1"].font = Font(bold=True, size=14)
            ws["A1"].alignment = header_alignment
            ws["A1"].fill = header_fill

            row = 3

            # Основная информация
            data = [
                ("Статус", f"{OrderStatus.get_status_emoji(order.status)} {OrderStatus.get_status_name(order.status)}"),
                ("Тип техники", order.equipment_type),
                ("Описание", order.description),
                ("", ""),
                ("КЛИЕНТ", ""),
                ("Имя клиента", order.client_name),
                ("Адрес", order.client_address),
                ("Телефон", order.client_phone),
            ]

            if order.notes:
                data.append(("", ""))
                data.append(("Заметки", order.notes))

            if order.scheduled_time:
                data.append(("Время прибытия", order.scheduled_time))

            if order.dispatcher_name:
                data.append(("", ""))
                data.append(("Диспетчер", order.dispatcher_name))

            if order.master_name:
                data.append(("Мастер", order.master_name))

            # Длительный ремонт
            if order.status == OrderStatus.DR:
                data.append(("", ""))
                data.append(("ДЛИТЕЛЬНЫЙ РЕМОНТ", ""))
                if order.estimated_completion_date:
                    data.append(("Срок окончания", order.estimated_completion_date))
                if order.prepayment_amount:
                    data.append(("Предоплата", f"{order.prepayment_amount:.2f} ₽"))

            # Финансовая информация для закрытых заявок
            if order.status == OrderStatus.CLOSED and order.total_amount:
                net_profit = order.total_amount - (order.materials_cost or 0)
                data.append(("", ""))
                data.append(("ФИНАНСОВАЯ ИНФОРМАЦИЯ", ""))
                data.append(("Общая сумма", f"{order.total_amount:.2f} ₽"))
                if order.materials_cost:
                    data.append(("Расходный материал", f"{order.materials_cost:.2f} ₽"))
                data.append(("Чистая прибыль", f"{net_profit:.2f} ₽"))
                if order.master_profit:
                    data.append(("Прибыль мастера", f"{order.master_profit:.2f} ₽"))
                if order.company_profit:
                    data.append(("Прибыль компании", f"{order.company_profit:.2f} ₽"))
                if order.has_review:
                    data.append(("Отзыв", "✅ Взят (+10%)"))
                if order.out_of_city:
                    data.append(("Выезд за город", "✅ Да"))

            # Даты
            data.append(("", ""))
            data.append(("ДАТЫ", ""))
            if order.created_at:
                data.append(("Создана", format_datetime(order.created_at)))
            if order.updated_at:
                data.append(("Обновлена", format_datetime(order.updated_at)))

            # Заполняем данные
            for label, value in data:
                if label == "":
                    row += 1
                    continue

                ws[f"A{row}"] = label
                ws[f"A{row}"].font = Font(bold=True)
                ws[f"B{row}"] = value
                row += 1

            # Настраиваем ширину столбцов
            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 50

            # Сохраняем в BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            # Создаем BufferedInputFile
            filename = f"order_{order_id}.xlsx"
            return BufferedInputFile(excel_file.read(), filename=filename)

        finally:
            await db.disconnect()
