#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Тестовый скрипт для создания Excel отчета и проверки всех листов
"""

import asyncio
import sys
import io
from pathlib import Path
from datetime import datetime

# Фикс кодировки для Windows
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# Добавляем корневую директорию в путь
sys.path.insert(0, str(Path(__file__).parent.parent))

from app.services.reports_service import ReportsService


async def main():
    """Главная функция"""
    print("\n" + "=" * 60)
    print("🧪 ТЕСТ: Генерация Excel отчета")
    print("=" * 60)

    try:
        # Создаем сервис отчетов
        print("\n📋 Создание сервиса отчетов...")
        reports_service = ReportsService()

        # Генерируем отчет
        print("📊 Генерация ежедневного отчета...")
        report = await reports_service.generate_daily_report()

        print(f"✅ Отчет сгенерирован:")
        print(f"   Тип: {report['type']}")
        print(f"   Период: {report['period']}")
        print(f"   Заказов всего: {report['orders']['total_orders']}")
        print(f"   Закрытых: {report['orders']['closed_orders']}")
        print(f"   Мастеров: {len(report['masters'])}")

        # Переподключаемся для Excel (т.к. generate_daily_report закрыл соединение)
        await reports_service.db.connect()

        try:
            # Сохраняем в Excel
            print("\n💾 Сохранение в Excel...")
            excel_path = await reports_service.save_report_to_excel(
                report,
                filename=f"test_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
        finally:
            await reports_service.db.disconnect()

        print(f"✅ Excel файл создан: {excel_path}")

        # Проверяем листы в файле
        print("\n🔍 Проверка листов в Excel...")
        try:
            from openpyxl import load_workbook

            wb = load_workbook(excel_path)
            sheet_names = wb.sheetnames

            print(f"✅ Найдено листов: {len(sheet_names)}")
            for i, name in enumerate(sheet_names, 1):
                print(f"   {i}. {name}")

            wb.close()

            # Проверяем наличие всех ожидаемых листов
            expected_sheets = ["Статистика", "Заказы", "Мастера", "Заявки по мастерам"]
            found_all = all(s in sheet_names for s in expected_sheets)

            if found_all:
                print("\n🎉 ВСЕ 4 ЛИСТА НА МЕСТЕ!")
            else:
                print("\n⚠️  Некоторые листы отсутствуют:")
                for s in expected_sheets:
                    status = "✅" if s in sheet_names else "❌"
                    print(f"   {status} {s}")

        except ImportError:
            print("⚠️  openpyxl не установлен, проверка листов пропущена")
            print("   Файл создан, откройте его в Excel для проверки")

        print("\n" + "=" * 60)
        print("✅ ТЕСТ ЗАВЕРШЕН")
        print("=" * 60)
        print(f"\n📂 Откройте файл: {excel_path}")
        print("📊 Проверьте все 4 листа!")
        print()

    except Exception as e:
        print(f"\n❌ ОШИБКА: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    asyncio.run(main())
